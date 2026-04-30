from pydoc import doc

from .ctk_helpers import (
    CTkToplevel,
    create_frame,
    create_label,
    create_button,
    create_entry,
    create_text,
    create_listbox,
    create_label_frame,
    create_checkbutton,
    create_radiobutton,
    create_scrollbar,
    tk,
    ttk,
)
from tkinter import messagebox, simpledialog, filedialog

from datetime import datetime
import json
import os
import re
import subprocess
import sys

if __package__ is None:
    # Allow running python app/Ingenieria.py directly (not installed as package)
    root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if root_dir not in sys.path:
        sys.path.insert(0, root_dir)
    __package__ = 'app'

try:
    from app.services.ingenieria_service import (
        listar_maquinas,
        crear_maquina,
        actualizar_maquina,
        eliminar_maquina,
        buscar_info_toro,
        exportar_maquinas,
        on_maquina_actualizada,
    )
    from app.services.ingenieria_extras import (
        listar_componentes,
        agregar_componente,
        registrar_calibracion,
        listar_materiales,
        crear_material,
        listar_sistemas,
        agregar_sistema,
        eliminar_sistema,
        listar_dispositivos,
        crear_dispositivo,
        actualizar_dispositivo,
        eliminar_dispositivo,
    )

    from app.database import SessionLocal
    from app.models import Historial, Calibracion
    from app.treeview_utils import aplicar_ordenamiento
    from app.view_manager import ViewManager
    from app.planos_module import generar_plano_general, PlanoData, Polyline, MarcadorMaquina
    from app.icon_button import tooltip as _tooltip, icon_btn as _icon_btn
    from app import theme
except ModuleNotFoundError:
    from services.ingenieria_service import (
        listar_maquinas,
        crear_maquina,
        actualizar_maquina,
        eliminar_maquina,
        buscar_info_toro,
        exportar_maquinas,
        on_maquina_actualizada,
    )
    from services.ingenieria_extras import (
        listar_componentes,
        agregar_componente,
        registrar_calibracion,
        listar_materiales,
        crear_material,
        listar_sistemas,
        agregar_sistema,
        eliminar_sistema,
    )

    from database import SessionLocal
    from models import Historial, Calibracion
    from treeview_utils import aplicar_ordenamiento
    from view_manager import ViewManager
    from planos_module import generar_plano_general, PlanoData, Polyline, MarcadorMaquina
    from icon_button import tooltip as _tooltip, icon_btn as _icon_btn
    import theme

try:
    from zoneinfo import ZoneInfo
except ImportError:
    from backports.zoneinfo import ZoneInfo


_current_ingenieria_panel = None


class UndoManager:
    """Pila de deshacer (Ctrl+Z) para acciones del panel de Ingeniería."""

    def __init__(self, max_size=50):
        self._stack = []          # list of (descripcion, undo_fn)
        self._max = max_size

    def push(self, descripcion, undo_fn):
        self._stack.append((descripcion, undo_fn))
        if len(self._stack) > self._max:
            self._stack.pop(0)

    def pop(self):
        if self._stack:
            return self._stack.pop()
        return None

    def can_undo(self):
        return len(self._stack) > 0

    def last_description(self):
        return self._stack[-1][0] if self._stack else None


class IngenieriaPanel:
    """Panel principal de Ingeniería con pestañas de máquinas, componentes y
    búsqueda en Toro.

    Esta clase organiza la interfaz y mantiene un método ``refrescar`` que es
    llamado por otros módulos cuando hay un cambio en la lista de máquinas.
    """

    def __init__(self, root):
        panel = CTkToplevel(root)
        panel.title("Ingeniería - Gestión de Máquinas")
        panel.geometry("1000x700")
        self.panel = panel
        
        # Center window
        panel.update_idletasks()
        w = panel.winfo_width()
        h = panel.winfo_height()
        x = (panel.winfo_screenwidth() // 2) - (w // 2)
        y = (panel.winfo_screenheight() // 2) - (h // 2)
        panel.geometry(f'{w}x{h}+{x}+{y}')
        
        # Initialize maquinas list early to avoid AttributeError
        self.maquinas = listar_maquinas() or []
        # Sistemas usados para clasificar componentes (se puede editar con el diálogo)
        self.sistemas_lista = listar_sistemas() or []
        # Sistemas detectados dinámicamente en los componentes (no se guarda en DB)
        self._component_sistemas = set()
        if not self.sistemas_lista:
            # Primer arranque: inicializar con valores por defecto.
            for s in []:
                agregar_sistema(s)
            self.sistemas_lista = listar_sistemas() or []
        # Undo manager
        self.undo = UndoManager()

        # Flag to avoid recreating la máquina demo cuando el usuario elimina todo
        self._disable_demo_on_empty = False

        # Callback to update the UI when a machine changes elsewhere (e.g. Operario)
        self._detalle_maquina_id = None
        self._detalle_hor_entry = None
        on_maquina_actualizada(self._on_maquina_actualizada)
        panel.bind_all("<Control-z>", lambda e: self._deshacer())
        panel.bind_all("<Control-Z>", lambda e: self._deshacer())
        
        # register self as current singleton
        global _current_ingenieria_panel
        _current_ingenieria_panel = self
        panel.protocol("WM_DELETE_WINDOW", self._on_close)

        # Container + ViewManager para vistas embebidas (sin abrir nuevas ventanas)
        self._vm_container = create_frame(panel)
        self._vm_container.pack(fill=tk.BOTH, expand=True)
        self.view_mgr = ViewManager(self._vm_container)

        # Wrapper que contiene notebook, gestionado por ViewManager como "main"
        main_wrapper = create_frame(self._vm_container)

        notebook = ttk.Notebook(main_wrapper)
        notebook.pack(fill=tk.BOTH, expand=True)
        # keep handle for programmatic tab switching later
        self.notebook = notebook

        self.view_mgr.set_main(main_wrapper)

        # crear pestañas en el orden deseado (Programaciones primero para visión inicial)
        # catch any errors so panel inicialización no se rompe
        tabs = [
            self._crea_tab_maquinas,
            self._crea_tab_componentes,
            self._crea_tab_programaciones,
            self._crea_tab_clases_calibracion,
            self._crea_tab_personal,
            self._crea_tab_registro,
            self._crea_tab_datos_red,
            self._crea_tab_busqueda,
        ]

        for crear_tab in tabs:
            try:
                crear_tab(notebook)
            except Exception as e:
                print("Error creando tab:", crear_tab.__name__, e)

        # primer llenado de datos para varias pestañas
        self.refrescar_lista()
        self.refrescar_historial()
        # debug: mostrar nombres de pestañas creadas
        try:
            names = [notebook.tab(i, "text") for i in range(notebook.index("end"))]
            print("[DEBUG] pestañas en notebook:", names)
        except Exception:
            pass
        try:
            self.refrescar_personal()
        except Exception:
            pass
        try:
            self.refrescar_programaciones()
        except Exception:
            pass
        # initialize registro diario as well so user sees datos inmediatamente
        try:
            self._cargar_registro()
        except Exception:
            pass
        # actualizar al cambiar de pestaña
        notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)

        # Aplicar tema actual a este panel y a los widgets creados
        theme.apply_theme(self.panel)

    def _on_tab_changed(self, event):
        """Callback when the notebook tab changes; refresh associated data."""
        notebook = event.widget
        idx = notebook.index(notebook.select())
        text = notebook.tab(idx, option="text")
        if "Programaciones" in text or "✏️" in text or "🔮" in text:
            try:
                self.refrescar_programaciones()
            except Exception:
                pass
        elif "Componentes" in text:
            # Asegurarse de que el filtro de máquinas esté sincronizado
            try:
                self._refresh_filtro_maquinas()
                self._refresh_componentes()
            except Exception:
                pass
        elif "Registro" in text:
            try:
                self._cargar_registro()
            except Exception:
                pass
        elif "Buscar" in text:
            if hasattr(self, 'result_text'):
                self.result_text.delete("1.0", tk.END)
        elif "Espacial" in text:
            try:
                self._cargar_datos_espaciales()
            except Exception:
                pass
        elif "IA" in text:
            try:
                self._ia_actualizar_estado()
            except Exception:
                pass





    def _crear_selector_maquina(self, parent):
        from .services.ingenieria_service import listar_maquinas

        frm_maq = tk.LabelFrame(parent, text="Máquina")
        frm_maq.pack(fill=tk.X, padx=10, pady=4)

        maq_var = tk.StringVar()
        maquinas = listar_maquinas() or []
        maq_values = [f"{m.id} - {m.nombre}" for m in maquinas]

        maq_combo = ttk.Combobox(frm_maq, values=sorted(maq_values), textvariable=maq_var)
        maq_combo.pack(fill=tk.X)

        return maq_var, maq_combo

    def _crear_input_tarea(self, parent):
        frm_tarea = tk.LabelFrame(parent, text="Descripción de la tarea")
        frm_tarea.pack(fill=tk.X, padx=10, pady=4)

        tipo_entry = tk.Entry(frm_tarea)
        tipo_entry.pack(fill=tk.X)

        status_frame = tk.Frame(frm_tarea)
        status_frame.pack(fill=tk.X, pady=(4, 0))
        self.ia_status_label = tk.Label(status_frame, text="IA: sin datos indexados", anchor="w", justify="left")
        self.ia_status_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        def _update_ia_status():
            self.ia_status_label.config(text=self._get_ia_status_text())

        _update_ia_status()

        return tipo_entry

    def _crear_sugerencias(self, parent, maq_var, tipo_entry):
        sugerencias_frame = tk.Frame(parent)
        sugerencias_frame.pack(fill=tk.X, pady=(4, 0))

        tk.Label(
            sugerencias_frame,
            text="Sugerencias (manuales + historial):",
            font=("Arial", 9, "italic"),
        ).pack(anchor="w")

        sugerencias_box = tk.Listbox(sugerencias_frame, height=4)
        sugerencias_box.pack(fill=tk.X)

        def _get_selected_maquina_id():
            txt = maq_var.get() if maq_var else ""
            if isinstance(txt, str) and " - " in txt:
                try:
                    return int(txt.split(" - ")[0])
                except Exception:
                    return None
            return None

        def _cargar_sugerencias():
            from .services.ingenieria_extras import listar_programaciones

            tipos = set()
            progs = listar_programaciones() or []
            for p in progs:
                if getattr(p, "tipo", None):
                    tipos.add((getattr(p, "tipo") or "").strip())

            maq_id = _get_selected_maquina_id()
            if maq_id:
                try:
                    from .services.historial_service import obtener_historial

                    historiales = obtener_historial(maquina_id=maq_id)
                    for h in historiales or []:
                        if getattr(h, "tipo_mantenimiento", None):
                            tipos.add((h.tipo_mantenimiento or "").strip())
                        if getattr(h, "tipo_registro", None):
                            tipos.add((h.tipo_registro or "").strip())
                        if getattr(h, "descripcion", None):
                            parts = [p.strip() for p in h.descripcion.split(";") if p.strip()]
                            for part in parts:
                                if len(part) > 2:
                                    tipos.add(part)
                except Exception:
                    pass

                try:
                    from .services.ingenieria_extras import listar_operaciones_por_maquina

                    ops = listar_operaciones_por_maquina(maq_id) or []
                    for o in ops:
                        op_text = "Operación diaria"
                        if getattr(o, "operador", None):
                            op_text = f"Operación diaria ({o.operador})"
                        tipos.add(op_text)
                except Exception:
                    pass

            return sorted(tipos)

        sugerencias_cache = []
        _last_maquina_id = None

        def _actualizar_sugerencias(event=None):
            nonlocal sugerencias_cache, _last_maquina_id

            maq_id = _get_selected_maquina_id()
            if maq_id != _last_maquina_id:
                sugerencias_cache = _cargar_sugerencias()
                _last_maquina_id = maq_id

            texto = tipo_entry.get().strip().lower()
            sugerencias_box.delete(0, tk.END)
            if not texto:
                return

            coincidencias = [t for t in sugerencias_cache if texto in t.lower()]
            try:
                from .services.ingenieria_service import obtener_feedback_sugerencia
                coincidencias = sorted(
                    coincidencias,
                    key=lambda s: obtener_feedback_sugerencia(texto, s, tipo="programacion"),
                    reverse=True,
                )
            except Exception:
                pass

            for t in coincidencias[:10]:
                sugerencias_box.insert(tk.END, t)

            try:
                from .services.embeddings import search_embeddings, format_embedding_sugerencia

                def _filtrar_pdf(item, score):
                    meta = item.metadata or {}
                    tipo = (meta.get("tipo") or "").lower()
                    if tipo in ("manual", "training", "libro", "documento", "historial"):
                        return True
                    return bool(meta.get("fragmento"))

                maq_id = _get_selected_maquina_id()
                emb_sug = search_embeddings(texto, top_k=10, filter_fn=_filtrar_pdf)

                preferred, others = [], []
                for item in emb_sug:
                    md = item.get("metadata", {}) or {}
                    if maq_id and md.get("maquina_id") == maq_id:
                        preferred.append((item, md))
                    else:
                        others.append((item, md))

                def _insert_pdf_suggestion(md):
                    fragment = md.get("fragmento") or md.get("manual_nombre") or ""
                    if not fragment:
                        return
                    sugerencias_box.insert(tk.END, format_embedding_sugerencia(md, fragment))

                for item, md in preferred:
                    _insert_pdf_suggestion(md)
                for item, md in others:
                    _insert_pdf_suggestion(md)
            except Exception:
                pass

            if sugerencias_box.size() > 0:
                sugerencias_box.selection_set(0)

        def _completar_con_sugerencia(event=None):
            if sugerencias_box.size() == 0:
                return "break"

            sel = sugerencias_box.curselection()
            if not sel:
                sugerencias_box.selection_set(0)
                sel = (0,)

            idx = sel[0]
            val = sugerencias_box.get(idx)
            next_idx = (idx + 1) % sugerencias_box.size()
            sugerencias_box.selection_clear(0, tk.END)
            sugerencias_box.selection_set(next_idx)

            try:
                from .services.ingenieria_service import registrar_feedback_sugerencia

                query_text = tipo_entry.get().strip()
                suggestion_text = val
                context = None
                if suggestion_text.startswith("(PDF"):
                    idx2 = suggestion_text.find(") ")
                    if idx2 != -1:
                        suggestion_text = suggestion_text[idx2 + 2 :]
                    context = "pdf"
                registrar_feedback_sugerencia(
                    query=query_text,
                    suggestion=suggestion_text,
                    tipo="programacion",
                    context=context,
                    utilidad=True,
                )
            except Exception:
                pass

            tipo_entry.delete(0, tk.END)
            tipo_entry.insert(0, suggestion_text)
            return "break"

        return sugerencias_box, _actualizar_sugerencias, _completar_con_sugerencia, sugerencias_cache

    def _crear_programacion_ui(self, parent):
        frm_sched = tk.LabelFrame(parent, text="Programado por")
        frm_sched.pack(fill=tk.X, padx=10, pady=4)

        sched_var = tk.StringVar(value="fecha")

        ttk.Radiobutton(frm_sched, text="Fecha", variable=sched_var, value="fecha").pack(side=tk.LEFT)
        ttk.Radiobutton(frm_sched, text="Horómetro", variable=sched_var, value="horometro").pack(side=tk.LEFT)
        ttk.Radiobutton(frm_sched, text="Kilometraje", variable=sched_var, value="kilometraje").pack(side=tk.LEFT)

        details_frame = tk.Frame(parent)
        details_frame.pack(fill=tk.X, padx=10, pady=2)

        fecha_frame = tk.Frame(details_frame)
        fecha_var = tk.StringVar()

        def elegir_fecha():
            cal_win = tk.Toplevel(self.panel)
            from tkcalendar import Calendar

            cal = Calendar(cal_win)
            cal.pack()

            def ok():
                fecha_var.set(cal.selection_get().isoformat())
                cal_win.destroy()

            tk.Button(cal_win, text="OK", command=ok).pack()

        _icon_btn(fecha_frame, "📅", "Seleccionar", elegir_fecha).pack(side=tk.LEFT)
        tk.Label(fecha_frame, textvariable=fecha_var, relief="sunken", width=12).pack(side=tk.LEFT)

        hor_frame = tk.Frame(details_frame)
        hor_label = tk.Label(hor_frame, text="Horómetro:")
        hor_label.pack(side=tk.LEFT)
        hor_entry = tk.Entry(hor_frame, width=10)
        hor_entry.pack(side=tk.LEFT)

        fecha_frame.grid(row=0, column=0, sticky="ew")
        hor_frame.grid(row=0, column=0, sticky="ew")

        def _toggle_sched(*args):
            mode = sched_var.get()
            if mode == "fecha":
                fecha_frame.tkraise()
            else:
                hor_frame.tkraise()
                hor_label.config(text="Kilometraje:" if mode == "kilometraje" else "Horómetro:")

        sched_var.trace_add("write", _toggle_sched)
        _toggle_sched()

        return sched_var, fecha_var, hor_entry

    def _crear_componente_ui(self, parent, maq_var):
        frm_comp = tk.LabelFrame(parent, text="Componente asociado")
        frm_comp.pack(fill=tk.X, padx=10, pady=4)

        comp_var = tk.StringVar()
        comp_combo = ttk.Combobox(frm_comp, textvariable=comp_var)
        comp_combo.pack(fill=tk.X)

        from .services.ingenieria_extras import listar_componentes

        def actualizar_componentes(event=None):
            sel = maq_var.get()
            if " - " not in sel:
                return

            mid = int(sel.split(" - ")[0])
            comps = listar_componentes(mid) or []
            values = []
            for c in comps:
                material = getattr(c, "material", None)
                if material:
                    values.append(f"{c.id} - {material.nombre}")
                else:
                    values.append(f"{c.id} - {getattr(c, 'nombre', 'Componente')}")

            comp_combo["values"] = values
            comp_var.set("")

        return comp_var, comp_combo, actualizar_componentes

    def _get_ia_status_text(self):
        """Texto rápido para mostrar el estado del índice semántico (embeddings)."""
        try:
            from .services.embeddings import get_default_vector_store
            store = get_default_vector_store()
            total = len(store._items)
            docs = []
            for it in store._items:
                md = it.metadata or {}
                doc = md.get("documento") or md.get("manual_nombre") or md.get("source")
                if doc and doc not in docs:
                    docs.append(doc)
            docs_text = ", ".join(docs[:3])
            if len(docs) > 3:
                docs_text += ", ..."
            return f"IA: {total} fragmentos indexados (docs: {docs_text or 'ninguno'})"
        except Exception:
            return "IA: estado no disponible"


    def _crea_tab_clases_calibracion(self, notebook):
        tab = ttk.Frame(notebook)
        notebook.add(tab, text="Calibraciones")

        frame = tk.Frame(tab)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        # =========================
        # CLASES DE CALIBRACIÓN
        # =========================

        self.clases_calibracion = [
            {"nombre": "GREENES", "valor": "3,2 mm", "maquinas": []},
            {"nombre": "TEES", "valor": "10 mm", "maquinas": []},
            {"nombre": "FAIRWAYS", "valor": "11 mm", "maquinas": []},
            {"nombre": "PREROUGH", "valor": "30 mm", "maquinas": []},
            {"nombre": "LOMAS", "valor": "25 mm", "maquinas": []},
            {"nombre": "BOSQUE", "valor": "30 mm", "maquinas": []},
        ]

        clases_frame = tk.LabelFrame(frame, text="Clases de calibración (actividades)", padx=8, pady=8)
        clases_frame.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)

        self.clases_listbox = tk.Listbox(clases_frame, height=8, exportselection=False)
        self.clases_listbox.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        btns = tk.Frame(clases_frame)
        btns.pack(fill=tk.X, pady=4)

        _icon_btn(btns, "➕", "Agregar clase", self._agregar_clase_calibracion).pack(side=tk.LEFT, padx=2)
        _icon_btn(btns, "🗑", "Eliminar clase", self._eliminar_clase_calibracion).pack(side=tk.LEFT, padx=2)

        tk.Label(clases_frame, text="Máquinas asociadas:").pack(anchor="w", padx=4)

        self.detalle_listbox = tk.Listbox(clases_frame, height=6)
        self.detalle_listbox.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        _icon_btn(clases_frame, "📁", "Exportar CSV", self._exportar_clase_calibracion).pack(pady=2)

        # cargar clases
        self._actualiza_clases_listbox()

        self.clases_listbox.bind("<<ListboxSelect>>", lambda e: self._mostrar_clase_detalle())
        self.clases_listbox.bind("<Double-1>", lambda e: self._on_clase_double_click())

        # =========================
        # PANEL ASOCIACIÓN
        # =========================

        asociar_frame = tk.LabelFrame(frame, text="Asociar máquinas a clase", padx=8, pady=8)
        asociar_frame.pack(fill=tk.BOTH, expand=True, side=tk.RIGHT)

        tk.Label(asociar_frame, text="Seleccione una clase:").pack(anchor="w")

        self.asoc_clase_var = tk.StringVar()

        self.asoc_clase_combo = ttk.Combobox(
        asociar_frame,
        values=[c["nombre"] for c in self.clases_calibracion],
        textvariable=self.asoc_clase_var,
        state="readonly"
        )
        self.asoc_clase_combo.pack(fill=tk.X, pady=2)

        tk.Label(asociar_frame, text="Máquinas disponibles:").pack(anchor="w")

        self.asoc_maquinas_listbox = tk.Listbox(
        asociar_frame,
        selectmode=tk.MULTIPLE,
        height=8
        )
        self.asoc_maquinas_listbox.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        # asegurar que exista la lista de máquinas
        if not hasattr(self, "maquinas") or self.maquinas is None:
            try:
                self.maquinas = listar_maquinas() or []
            except Exception:
                self.maquinas = []

        self._actualiza_maquinas_listbox()

        _icon_btn(
            asociar_frame,
            "✔",
            "Asociar máquinas",
            self._asociar_maquinas_a_clase
        ).pack(fill=tk.X, pady=4)

        self.asoc_clase_combo.bind(
             "<<ComboboxSelected>>",
        lambda e: self._actualiza_maquinas_asociadas()
        )

        # =========================
        # INICIALIZACIÓN
        # =========================

        if self.clases_calibracion:
            primera = self.clases_calibracion[0]["nombre"]
            self.asoc_clase_var.set(primera)
            self._actualiza_maquinas_asociadas()
        # seleccionar primera clase para mostrar detalle
        if self.clases_listbox.size() > 0:
            self.clases_listbox.selection_set(0)
        self._mostrar_clase_detalle()





    def _actualiza_clases_listbox(self):
        self.clases_listbox.delete(0, tk.END)
        for clase in self.clases_calibracion:
            maquinas_txt = f" ({len(clase['maquinas'])} máquinas)" if clase['maquinas'] else ""
            self.clases_listbox.insert(tk.END, f"{clase['nombre']} - {clase['valor']}{maquinas_txt}")
        # actualizar detalle si la clase seleccionada sigue presente
        self._mostrar_clase_detalle()

    def _agregar_clase_calibracion(self):
        nombre = simpledialog.askstring("Agregar clase", "Nombre de la clase:", parent=self.panel)
        if not nombre:
            return
        valor = simpledialog.askstring("Valor de calibración", "Valor (ej: 10 mm):", parent=self.panel)
        if not valor:
            return
        self.clases_calibracion.append({"nombre": nombre.strip(), "valor": valor.strip(), "maquinas": []})
        self.asoc_clase_combo['values'] = [c["nombre"] for c in self.clases_calibracion]
        self._actualiza_clases_listbox()

    def _eliminar_clase_calibracion(self):
        sel = self.clases_listbox.curselection()
        if not sel:
            messagebox.showwarning("Aviso", "Seleccione una clase para eliminar", parent=self.panel)
            return
        idx = sel[0]
        nombre = self.clases_calibracion[idx]["nombre"]
        if not messagebox.askyesno("Confirmar", f"¿Eliminar la clase '{nombre}'?", parent=self.panel):
            return
        del self.clases_calibracion[idx]
        self.asoc_clase_combo['values'] = [c["nombre"] for c in self.clases_calibracion]
        self._actualiza_clases_listbox()
        self.detalle_listbox.delete(0, tk.END)

    def _actualiza_maquinas_listbox(self):
        self.asoc_maquinas_listbox.delete(0, tk.END)
        for maq in self.maquinas:
            self.asoc_maquinas_listbox.insert(tk.END, f"{maq.id} - {maq.nombre}")

    def _guardar_csv_clase(self, clase, path):
        """Dumps the given calibration class and its machines to CSV file."""
        import csv
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, delimiter=";")
            w.writerow(["Máquina", "Calibración", "Clase de Calibración"])
            for maq_id in clase.get("maquinas", []):
                maq = next((m for m in self.maquinas if m.id == maq_id), None)
                if maq:
                    w.writerow([f"{maq.id} - {maq.nombre}", clase.get("valor", ""), clase.get("nombre", "")])

    def _mostrar_clase_detalle(self, event=None):
        """Muestra en el panel lateral las máquinas asociadas a la clase seleccionada."""
        sel = self.clases_listbox.curselection()
        self.detalle_listbox.delete(0, tk.END)
        if not sel:
            return
        idx = sel[0]
        clase = self.clases_calibracion[idx]
        for maq_id in clase.get("maquinas", []):
            maq = next((m for m in self.maquinas if m.id == maq_id), None)
            if maq:
                self.detalle_listbox.insert(tk.END, f"{maq.id} - {maq.nombre}")

    def _exportar_clase_calibracion(self):
        sel = self.clases_listbox.curselection()
        if not sel:
            messagebox.showwarning("Exportar", "Seleccione una clase para exportar", parent=self.panel)
            return
        idx = sel[0]
        clase = self.clases_calibracion[idx]
        maquinas = clase.get("maquinas", [])
        if not maquinas:
            messagebox.showinfo("Exportar", "No hay máquinas asociadas a esta clase.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv")])
        if not path:
            return
        self._guardar_csv_clase(clase, path)
        messagebox.showinfo("Exportado", f"CSV guardado en:\n{path}")

    def _asociar_maquinas_a_clase(self):
        clase_nombre = self.asoc_clase_var.get()
        if not clase_nombre:
            messagebox.showwarning("Aviso", "Seleccione una clase", parent=self.panel)
            return
        sel_maqs = self.asoc_maquinas_listbox.curselection()
        if not sel_maqs:
            messagebox.showwarning("Aviso", "Seleccione una o más máquinas", parent=self.panel)
            return
        clase = next((c for c in self.clases_calibracion if c["nombre"] == clase_nombre), None)
        if not clase:
            messagebox.showerror("Error", "Clase no encontrada", parent=self.panel)
            return
        clase["maquinas"] = [self.maquinas[i].id for i in sel_maqs]
        self._actualiza_clases_listbox()
        self._mostrar_clase_detalle()
        messagebox.showinfo("Asociación", f"Máquinas asociadas a '{clase_nombre}'")

    def _actualiza_maquinas_asociadas(self):
        clase_nombre = self.asoc_clase_var.get()
        clase = next((c for c in self.clases_calibracion if c["nombre"] == clase_nombre), None)

        if not clase:
            return
            return
        # limpiar selección
        self.asoc_maquinas_listbox.selection_clear(0, tk.END)

        # marcar máquinas que pertenecen a la clase
        for idx, maq in enumerate(self.maquinas):
            if maq.id in clase["maquinas"]:
                self.asoc_maquinas_listbox.selection_set(idx)

        # refrescar contadores y panel de detalle
        self._actualiza_clases_listbox()
        self._mostrar_clase_detalle()
    def _on_clase_double_click(self):
        """Al hacer doble click en una clase muestra un panel con las máquinas.

        Además de actualizar el panel lateral, se abre una ventana "detalle" con
        la lista completa para que el usuario pueda verla claramente. Este método
        ya no cambia otras pestañas; esa lógica se mantiene separada en caso de
        que la navegación sea necesaria más adelante.
        """
        sel = self.clases_listbox.curselection()
        if not sel:
            return
        idx = sel[0]
        clase = self.clases_calibracion[idx]
        maq_ids = list(clase.get("maquinas", []))
        # update side-detail
        self._mostrar_clase_detalle()
        if not maq_ids:
            return
        # abrir ventana para mostrar máquinas asociadas
        win = tk.Toplevel(self.panel)
        win.title(f"Máquinas en clase {clase['nombre']}")
        lb = tk.Listbox(win)
        lb.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        for maq_id in maq_ids:
            maq = next((m for m in self.maquinas if m.id == maq_id), None)
            if maq:
                lb.insert(tk.END, f"{maq.id} - {maq.nombre}")
        _icon_btn(win, "❌", "Cerrar", win.destroy).pack(pady=5)

    def _on_close(self):
        global _current_ingenieria_panel
        _current_ingenieria_panel = None
        self.panel.destroy()

    def _deshacer(self):
        """Ejecuta Ctrl+Z: deshace la última acción registrada."""
        item = self.undo.pop()
        if not item:
            messagebox.showinfo("Deshacer", "No hay acciones para deshacer.", parent=self.panel)
            return
        desc, fn = item
        try:
            fn()
            messagebox.showinfo("Deshacer", f"Acción deshecha: {desc}", parent=self.panel)
        except Exception as e:
            messagebox.showerror("Error al deshacer", str(e), parent=self.panel)

    def _snapshot_maquina(self, maq):
        """Captura un dict con todos los campos de una máquina para restaurar."""
        return {
            'id': maq.id, 'nombre': maq.nombre, 'categoria': maq.categoria,
            'estado': maq.estado, 'ubicacion': maq.ubicacion,
            'fabricante': maq.fabricante, 'modelo': maq.modelo,
            'tipo_unidad': maq.tipo_unidad, 'año': maq.año,
            'codigo_serie': maq.codigo_serie, 'motor_marca': maq.motor_marca,
            'motor_modelo': maq.motor_modelo, 'motor_serie': maq.motor_serie,
            'tipo_combustible': maq.tipo_combustible,
            'altura_corte_mm': maq.altura_corte_mm,
            'material_id': maq.material_id, 'observaciones': maq.observaciones,
            'horometro_inicial': maq.horometro_inicial,
            'horometro_actual': maq.horometro_actual,
            'fecha_operacion': maq.fecha_operacion,
        }

    def _restore_maquina(self, snap):
        """Recrea una máquina a partir de un snapshot."""
        crear_maquina(
            nombre=snap['nombre'],
            horometro_inicial=snap.get('horometro_inicial', 0),
            categoria=snap.get('categoria'),
            estado=snap.get('estado'),
            ubicacion=snap.get('ubicacion'),
            fabricante=snap.get('fabricante'),
            modelo=snap.get('modelo'),
            tipo_unidad=snap.get('tipo_unidad'),
            año=snap.get('año'),
            codigo_serie=snap.get('codigo_serie'),
            motor_marca=snap.get('motor_marca'),
            motor_modelo=snap.get('motor_modelo'),
            motor_serie=snap.get('motor_serie'),
            tipo_combustible=snap.get('tipo_combustible'),
            altura_corte_mm=snap.get('altura_corte_mm'),
            material_id=snap.get('material_id'),
            observaciones=snap.get('observaciones'),
            fecha_operacion=snap.get('fecha_operacion'),
            id=snap.get('id'),
        )
        # Restaurar horómetro actual si difiere del inicial
        if snap.get('horometro_actual') and snap['horometro_actual'] != snap.get('horometro_inicial', 0):
            actualizar_maquina(snap['id'], horometro_actual=snap['horometro_actual'])
        self.refrescar_lista()

    def _ensure_demo(self):
        """Create demo machine if no machines exist."""
        if self._disable_demo_on_empty:
            return False
        if not self.maquinas:
            try:
                crear_maquina(
                    nombre="REELMASTER 5510",
                    horometro_inicial=0,
                    categoria="Cortacésped",
                    estado="Operativa",
                    ubicacion="Almacén",
                    fabricante="Toro",
                    modelo="5510",
                    tipo_unidad="Cortacésped",
                    año=2020,
                    codigo_serie="SN1234",
                    motor_marca="Honda",
                    motor_modelo="GXV160",
                    motor_serie="MOTOR001",
                    tipo_combustible="Gasolina",
                    altura_corte_mm=25,
                    observaciones="Máquina de ejemplo para demostración"
                )
                # Reload machines list
                self.maquinas = listar_maquinas() or []
                return True
            except Exception as e:
                print(f"[ERROR] Could not create demo machine: {e}")
                return False
        return False

    def _crea_tab_maquinas(self, notebook):
        self.tab_maquinas = ttk.Frame(notebook)
        notebook.add(self.tab_maquinas, text="Máquinas")

        # Barra de herramientas exclusiva de Máquinas
        toolbar = tk.Frame(self.tab_maquinas)
        toolbar.pack(fill=tk.X, padx=10, pady=(4, 0))
        _icon_btn(toolbar, "➕", "Agregar máquina", self.agregar_maquina).pack(side=tk.LEFT, padx=2)
        _icon_btn(toolbar, "🗑", "Eliminar máquina", self.eliminar).pack(side=tk.LEFT, padx=2)
        _icon_btn(toolbar, "📤", "Exportar", self._dialogo_exportar).pack(side=tk.RIGHT, padx=2)

        frame_lista = tk.Frame(self.tab_maquinas)
        frame_lista.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=10)

        scrollbar = tk.Scrollbar(frame_lista)
        # enable extended selection (Ctrl/Shift) so user can pick multiple
        self.listbox = tk.Listbox(frame_lista, width=70, yscrollcommand=scrollbar.set, selectmode=tk.EXTENDED)
        scrollbar.config(command=self.listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        # actualizar componentes/historial al cambiar selección
        # (usar la vista de componentes por sistema, no la antigua matriz de máquinas)
        self.listbox.bind('<<ListboxSelect>>', lambda e: self._refresh_componentes())
        # doble clic abre vista detallada (incluye imágenes/componentes)
        self.listbox.bind('<Double-1>', lambda e: self.detallar_maquina())

        # section for historial timeline below
        frame_hist = tk.Frame(self.tab_maquinas)
        frame_hist.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # ── Row 1: machine & action filters ──
        ctrl1 = tk.Frame(frame_hist)
        ctrl1.pack(fill=tk.X, pady=(0, 3))

        tk.Label(ctrl1, text="Máquina:").pack(side=tk.LEFT, padx=(0, 2))
        self.hist_maq_var = tk.StringVar(value="Todas")
        self.hist_maq_combo = ttk.Combobox(ctrl1, textvariable=self.hist_maq_var,
                                            values=["Todas"], state="readonly", width=25)
        self.hist_maq_combo.pack(side=tk.LEFT, padx=2)

        tk.Label(ctrl1, text="  Origen:").pack(side=tk.LEFT, padx=(6, 2))
        self.hist_origen_var = tk.StringVar(value="Todos")
        self._hist_origen_opciones = [
            "Todos",
            "Registrar Mantenimiento",
            "Registrar Calibración",
            "Solicitar Reparación",
            "Cambio de Componente",
            "Gestión de Materiales",
            "Gestión de Personal",
        ]
        self.hist_origen_combo = ttk.Combobox(ctrl1, textvariable=self.hist_origen_var,
                                              values=self._hist_origen_opciones,
                                              state="readonly", width=24)
        self.hist_origen_combo.pack(side=tk.LEFT, padx=2)

        tk.Label(ctrl1, text="  Acción:").pack(side=tk.LEFT, padx=(6, 2))
        self.hist_accion_var = tk.StringVar(value="Todas")
        self.hist_accion_combo = ttk.Combobox(ctrl1, textvariable=self.hist_accion_var,
                                              values=["Todas"], state="readonly", width=22)
        self.hist_accion_combo.pack(side=tk.LEFT, padx=2)

        # ── Row 2: date filters + buttons ──
        ctrl2 = tk.Frame(frame_hist)
        ctrl2.pack(fill=tk.X, pady=(0, 3))

        self.hist_start_var = tk.StringVar()
        self.hist_end_var = tk.StringVar()
        tk.Label(ctrl2, text="Desde:").pack(side=tk.LEFT)
        tk.Label(ctrl2, textvariable=self.hist_start_var, width=10, relief="sunken").pack(side=tk.LEFT)
        _icon_btn(ctrl2, "📅", "Elegir fecha inicio", lambda: self._elegir_hist_fecha(self.hist_start_var)).pack(side=tk.LEFT)
        tk.Label(ctrl2, text="Hasta:").pack(side=tk.LEFT, padx=(10, 0))
        tk.Label(ctrl2, textvariable=self.hist_end_var, width=10, relief="sunken").pack(side=tk.LEFT)
        _icon_btn(ctrl2, "📅", "Elegir fecha fin", lambda: self._elegir_hist_fecha(self.hist_end_var)).pack(side=tk.LEFT)

        _icon_btn(ctrl2, "🔍", "Filtrar historial", self.refrescar_historial).pack(side=tk.LEFT, padx=(12, 4))
        _icon_btn(ctrl2, "📤", "Exportar historial", self._exportar_historial_panel).pack(side=tk.LEFT, padx=4)

        hist_cols = ("fecha", "maquina", "tipo", "descripcion")
        style = ttk.Style()
        style.configure("Treeview", rowheight=60)

        self.hist_tree = ttk.Treeview(frame_hist, columns=hist_cols, show="headings")
        self.hist_tree.heading("fecha", text="Fecha")
        self.hist_tree.heading("maquina", text="Máquina")
        self.hist_tree.heading("tipo", text="Tipo / Acción")
        self.hist_tree.heading("descripcion", text="Descripción")
        self.hist_tree.column("fecha", width=130)
        self.hist_tree.column("maquina", width=120)
        self.hist_tree.column("tipo", width=150)
        self.hist_tree.column("descripcion", width=250)
        hsv = ttk.Scrollbar(frame_hist, orient="vertical", command=self.hist_tree.yview)
        self.hist_tree.configure(yscrollcommand=hsv.set)
        self.hist_tree.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)
        hsv.pack(fill=tk.Y, side=tk.RIGHT)
        aplicar_ordenamiento(self.hist_tree)
        # allow user to double-click a row to see full details (historial or calibración)
        self.hist_tree.bind('<Double-1>', self._mostrar_detalle_hist)

    def _mostrar_detalle_hist(self, event=None):
        sel = self.hist_tree.selection()
        if not sel:
            return
        iid = sel[0]
        kind = iid[0]
        rec_id = int(iid[1:])
        db = SessionLocal()
        try:
            if kind == 'h':
                rec = db.query(Historial).filter(Historial.id == rec_id).first()
                texto = f"Registro: {rec.tipo_registro}\n"
                texto += f"Mantenimiento: {rec.tipo_mantenimiento}\n"
                texto += f"Técnico: {rec.tecnico or ''}\n"
                texto += f"Horómetro: {rec.horometro}\n"
                texto += f"Descripción:\n{rec.descripcion or ''}\n"
                fotos = []
                if rec.descripcion and 'fotos:' in rec.descripcion:
                    after = rec.descripcion.split('fotos:')[1]
                    fotos = [p for p in after.split(';') if p]
            else:
                rec = db.query(Calibracion).filter(Calibracion.id == rec_id).first()
                texto = f"Calibración tipo: {rec.tipo}\nDuración: {rec.duracion_minutos} minutos\nObservaciones:\n{rec.observaciones or ''}\n"
                fotos = []
        finally:
            db.close()
        def _build_detail(content):
            tk.Label(content, text=texto, justify='left').pack(padx=10, pady=10)
            for fpath in fotos:
                def _open(p=fpath):
                    import webbrowser
                    try:
                        webbrowser.open(p)
                    except Exception:
                        pass
                _icon_btn(content, "📁", f'Ver imagen {fpath}', _open).pack(pady=2)
        self.view_mgr.push('Detalle historial', _build_detail)

    def _crea_tab_componentes(self, notebook):
        """Crea la pestaña de Componentes con árboles por sistema."""

        from .services.ingenieria_extras import listar_componentes

        # ------------------------------------------------
        # TAB PRINCIPAL
        # ------------------------------------------------
        self.tab_componentes = ttk.Frame(notebook)
        notebook.add(self.tab_componentes, text="Componentes")

        # ------------------------------------------------
        # BARRA SUPERIOR
        # ------------------------------------------------
        upper = tk.Frame(self.tab_componentes)
        upper.pack(fill=tk.X, padx=10, pady=5)

        # Asociar máquina + material (crea un componente)
        _icon_btn(upper, "🔗", "Asociar Máquina-Material", self._dialogo_asociar_maquina_material).pack(side=tk.LEFT)

        # Materiales (contabilidad integrada)
        _icon_btn(upper, "🧱", "Materiales", self._dialogo_materiales).pack(side=tk.LEFT, padx=10)
        _icon_btn(upper, "🧾", "Contabilidad de componentes reemplazados", self._dialogo_contabilidad_componentes).pack(side=tk.LEFT, padx=10)
        _icon_btn(upper, "📤", "", self._exportar_componentes).pack(side=tk.RIGHT, padx=10)
        _icon_btn(upper, "📥", "Importar componentes", self._importar_componentes).pack(side=tk.RIGHT, padx=10)

        # Filtro de máquinas (desplegable)
        self._filtro_maquinas_ids = None  # None = todas
        self._filtro_maquina_var = tk.StringVar(value="Todos")
        self._filtro_maquina_combo = ttk.Combobox(
            upper,
            textvariable=self._filtro_maquina_var,
            values=[""],
            state="readonly",
            width=30,
        )
        self._filtro_maquina_combo.pack(side=tk.LEFT, padx=10)
        self._filtro_maquina_combo.bind("<<ComboboxSelected>>", lambda e: self._on_filtro_maquina_change())

        # Etiqueta que muestra en qué máquinas estamos filtrando
        self._filtro_maquinas_label = tk.Label(upper, text="Máquinas: todas")
        self._filtro_maquinas_label.pack(side=tk.LEFT, padx=10)

        # ------------------------------------------------
        # LISTA DE COMPONENTES (unificada)
        # ------------------------------------------------
        self.comp_tree = ttk.Treeview(
            self.tab_componentes,
            columns=("maquina", "nombre", "codigo", "tipo", "sistema", "descripcion"),
            show="headings",
            selectmode="extended",
        )
        self.comp_tree.heading("maquina", text="Máquina")
        self.comp_tree.heading("nombre", text="Nombre")
        self.comp_tree.heading("codigo", text="Código")
        self.comp_tree.heading("tipo", text="Tipo")
        self.comp_tree.heading("sistema", text="Sistema")
        self.comp_tree.heading("descripcion", text="Descripción")

        self.comp_tree.column("maquina", width=140)
        self.comp_tree.column("nombre", width=180)
        self.comp_tree.column("codigo", width=120)
        self.comp_tree.column("tipo", width=100)
        self.comp_tree.column("sistema", width=120)
        self.comp_tree.column("descripcion", width=250)

        hscroll = ttk.Scrollbar(self.tab_componentes, orient="horizontal", command=self.comp_tree.xview)
        vscroll = ttk.Scrollbar(self.tab_componentes, orient="vertical", command=self.comp_tree.yview)
        self.comp_tree.configure(xscrollcommand=hscroll.set, yscrollcommand=vscroll.set)
        self.comp_tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=(10, 0))
        hscroll.pack(fill=tk.X, side=tk.BOTTOM, padx=10)
        vscroll.pack(fill=tk.Y, side=tk.RIGHT, padx=(0, 10))

        # Selección de componentes y botón para generar sistema
        self.comp_selection_label = tk.Label(
            self.tab_componentes,
            text="Selecciona uno o más componentes para generar sistema",
            bg=theme.current.get("label_bg", "white"),
            fg=theme.current.get("label_fg", "black"),
        )
        self.comp_selection_label.pack(fill=tk.X, padx=10, pady=(5, 0))

        # Botones de acción (orden horizontal para optimizar espacio)
        btn_frame = tk.Frame(self.tab_componentes)
        btn_frame.pack(fill=tk.X, padx=10, pady=(2, 10))

        self._generar_sistema_btn = tk.Button(
            btn_frame,
            text="Generar Sistema",
            state=tk.DISABLED,
            command=self._generar_sistema_desde_seleccion,
        )
        self._generar_sistema_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)

        self._editar_sistema_btn = tk.Button(
            btn_frame,
            text="Editar sistema",
            state=tk.DISABLED,
            command=self._editar_sistema_seleccion,
        )
        self._editar_sistema_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)

        self._añadir_a_sistema_existente_btn = tk.Button(
            btn_frame,
            text="Añadir a sistema existente",
            state=tk.DISABLED,
            command=self._agregar_seleccion_a_sistema_existente,
        )
        self._añadir_a_sistema_existente_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)

        self.comp_tree.bind("<<TreeviewSelect>>", self._on_select_componentes)
        self.comp_tree.bind("<Double-1>", self._on_comp_double_click)

        self.comp_tree.bind("<<TreeviewSelect>>", self._on_select_componentes)
        self.comp_tree.bind("<Double-1>", self._on_comp_double_click)

        # Keep compatibility with old selection / delete logic.
        self.comp_trees = {"Componentes": self.comp_tree}
        self.comp_selection_labels = {"Componentes": self.comp_selection_label}

        # Carga inicial de componentes
        try:
            self.componentes = listar_componentes() or []
        except Exception as e:
            print("Error cargando componentes:", e)
            self.componentes = []

        self.comp_map = {c.id: c for c in self.componentes}

        self._refresh_componentes()
        self._refresh_filtro_maquinas()

    def _dialogo_filtrar_maquinas(self):
        """Dialogo para seleccionar máquinas (multi-selección)."""
        # Si ya existe el diálogo, simplemente mostrarlo y traerlo al frente.
        if hasattr(self, '_filtrar_maquinas_win') and self._filtrar_maquinas_win.winfo_exists():
            win = self._filtrar_maquinas_win
            win.deiconify()
            win.lift()
            return

        win = tk.Toplevel(self.panel)
        self._filtrar_maquinas_win = win
        win.title("Filtrar máquinas")
        win.geometry("400x400")
        win.config(bg=theme.current.get("bg"))
        theme.apply_theme(win)

        lbl = tk.Label(
            win,
            text="Selecciona una o más máquinas:",
            bg=theme.current.get("label_bg"),
            fg=theme.current.get("label_fg"),
        )
        lbl.pack(padx=10, pady=10)

        listbox = tk.Listbox(
            win,
            selectmode=tk.MULTIPLE,
            bg=theme.current.get("listbox_bg"),
            fg=theme.current.get("listbox_fg"),
            selectbackground=theme.current.get("select_bg"),
            selectforeground=theme.current.get("select_fg"),
            highlightthickness=0,
            bd=0,
        )
        listbox.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # populate listbox
        for m in self.maquinas:
            listbox.insert(tk.END, f"{m.id} - {m.nombre}")

        if self._filtro_maquinas_ids:
            for idx, m in enumerate(self.maquinas):
                if m.id in self._filtro_maquinas_ids:
                    listbox.selection_set(idx)

        def aplicar():
            sel = listbox.curselection()
            if not sel:
                self._filtro_maquinas_ids = None
                self._filtro_maquinas_label.config(text="Máquinas: todas")
            else:
                ids = [self.maquinas[i].id for i in sel]
                self._filtro_maquinas_ids = ids
                names = [self.maquinas[i].nombre for i in sel]
                self._filtro_maquinas_label.config(text=f"Máquinas: {', '.join(names)}")
            self._refresh_componentes()
            win.withdraw()

        _icon_btn(win, "✔", "Aplicar filtro", aplicar).pack(pady=10)

        # cerrar sólo oculta para mantener la instancia
        win.protocol("WM_DELETE_WINDOW", win.withdraw)

    def _dialogo_contabilidad(self):
        """Panel de contabilidad usando los datos de materiales."""
        # Reusar si ya existe
        if hasattr(self, '_contabilidad_win') and self._contabilidad_win.winfo_exists():
            win = self._contabilidad_win
            win.deiconify()
            win.lift()
            return

        from .services.ingenieria_extras import listar_materiales

        win = tk.Toplevel(self.panel)
        self._contabilidad_win = win
        win.title("Contabilidad de materiales")
        win.geometry("700x400")
        win.config(bg=theme.current.get("bg"))
        theme.apply_theme(win)

        frame = tk.Frame(win, bg=theme.current.get("frame_bg"))
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        cols = ("nombre", "codigo", "tipo", "stock_actual", "stock_minimo", "sistema", "fecha_actualizacion")
        tree = ttk.Treeview(frame, columns=cols, show="headings")
        for c in cols:
            tree.heading(c, text=c.replace("_", " ").capitalize())
            tree.column(c, width=100, anchor="w")

        vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tree.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)
        vsb.pack(fill=tk.Y, side=tk.RIGHT)
        hsb.pack(fill=tk.X, side=tk.BOTTOM)

        def refresh():
            tree.delete(*tree.get_children())
            for m in listar_materiales():
                tree.insert("", tk.END, values=(
                    m.nombre,
                    m.codigo,
                    m.tipo,
                    m.stock_actual,
                    m.stock_minimo,
                    getattr(m, 'sistema', ''),
                    getattr(m, 'fecha_actualizacion', None) and m.fecha_actualizacion.strftime("%Y-%m-%d %H:%M"),
                ))

        btn_frame = tk.Frame(win, bg=theme.current.get("frame_bg"))
        btn_frame.pack(fill=tk.X, padx=10, pady=5)
        _icon_btn(btn_frame, "🔄", "Refrescar", refresh).pack(side=tk.LEFT)
        _icon_btn(btn_frame, "❌", "Cerrar", win.withdraw).pack(side=tk.RIGHT)

        refresh()

        win.protocol("WM_DELETE_WINDOW", win.withdraw)

    def _dialogo_contabilidad_componentes(self):
        """Panel de contabilidad de componentes extraídos de historial."""
        if hasattr(self, '_contabilidad_componentes_win') and self._contabilidad_componentes_win.winfo_exists():
            win = self._contabilidad_componentes_win
            win.deiconify()
            win.lift()
            return

        from .services.ingenieria_extras import obtener_uso_componentes, agrupar_uso_componentes

        win = tk.Toplevel(self.panel)
        self._contabilidad_componentes_win = win
        win.title("Contabilidad de componentes reemplazados")
        win.geometry("900x450")
        win.config(bg=theme.current.get("bg"))
        theme.apply_theme(win)

        filtro_frame = tk.Frame(win, bg=theme.current.get("frame_bg"))
        filtro_frame.pack(fill=tk.X, padx=10, pady=6)

        tk.Label(filtro_frame, text="Máquina:", bg=theme.current.get("frame_bg"), fg=theme.current.get("label_fg")).pack(side=tk.LEFT)
        self._comp_maquina_var = tk.StringVar(value="Todas")
        maquinas_values = ["Todas"] + [f"{m.id} - {m.nombre}" for m in (self.maquinas or [])]
        self._comp_maquina_combo = ttk.Combobox(filtro_frame, values=maquinas_values, textvariable=self._comp_maquina_var, state="readonly", width=30)
        self._comp_maquina_combo.pack(side=tk.LEFT, padx=5)

        tk.Label(filtro_frame, text="Fecha inicio:", bg=theme.current.get("frame_bg"), fg=theme.current.get("label_fg")).pack(side=tk.LEFT, padx=(10,0))
        self._comp_fecha_inicio_var = tk.StringVar(value="")
        fecha_inicio_entry = tk.Entry(filtro_frame, textvariable=self._comp_fecha_inicio_var, width=18)
        fecha_inicio_entry.pack(side=tk.LEFT, padx=5)

        tk.Label(filtro_frame, text="Fecha fin:", bg=theme.current.get("frame_bg"), fg=theme.current.get("label_fg")).pack(side=tk.LEFT, padx=(10,0))
        self._comp_fecha_fin_var = tk.StringVar(value="")
        fecha_fin_entry = tk.Entry(filtro_frame, textvariable=self._comp_fecha_fin_var, width=18)
        fecha_fin_entry.pack(side=tk.LEFT, padx=5)

        btns_frame = tk.Frame(win, bg=theme.current.get("frame_bg"))
        btns_frame.pack(fill=tk.X, padx=10, pady=4)

        def _refrescar():
            tree.delete(*tree.get_children())
            start = None
            end = None
            try:
                if self._comp_fecha_inicio_var.get().strip():
                    start = datetime.fromisoformat(self._comp_fecha_inicio_var.get().strip())
            except Exception:
                start = None
            try:
                if self._comp_fecha_fin_var.get().strip():
                    end = datetime.fromisoformat(self._comp_fecha_fin_var.get().strip())
            except Exception:
                end = None

            maquina_id = None
            sel_maq = self._comp_maquina_var.get().strip()
            if sel_maq and sel_maq != "Todas":
                try:
                    maquina_id = int(sel_maq.split(maxsplit=1)[0])
                except Exception:
                    maquina_id = None

            rows = obtener_uso_componentes(fecha_inicio=start, fecha_fin=end, maquina_id=maquina_id)
            for r in rows:
                tree.insert("", tk.END, values=(
                    r.get("fecha") and r.get("fecha").strftime("%Y-%m-%d %H:%M"),
                    r.get("maquina_id"),
                    r.get("componente"),
                    r.get("codigo") or "",
                    r.get("cantidad") if r.get("cantidad") is not None else "",
                    r.get("unidad") or "",
                    r.get("hora_inicio") and r.get("hora_inicio").strftime("%Y-%m-%d %H:%M"),
                    r.get("hora_fin") and r.get("hora_fin").strftime("%Y-%m-%d %H:%M"),
                    r.get("horometro") if r.get("horometro") is not None else "",
                    r.get("historial_id"),
                ))

            resumen = agrupar_uso_componentes(fecha_inicio=start, fecha_fin=end)
            resumen_text = "  ".join([f"{x['componente']} ({x['unidad'] or 'u'}): {x['total']}" for x in resumen])
            summary_label.config(text=f"Totales: {len(rows)} registros. {resumen_text}")

        def _exportar_csv():
            rows = tree.get_children()
            if not rows:
                messagebox.showinfo("Exportar", "No hay registros para exportar.", parent=win)
                return

            path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV", "*.csv")],
                title="Guardar contabilidad como...",
                parent=win,
            )
            if not path:
                return

            import csv
            try:
                with open(path, "w", newline="", encoding="utf-8-sig") as fh:
                    writer = csv.writer(fh, delimiter=";")
                    headers = [tree.heading(c, option="text") for c in tree["columns"]]
                    writer.writerow(headers)
                    for iid in rows:
                        writer.writerow(tree.item(iid, "values"))
            except PermissionError:
                messagebox.showerror("Error", f"No tiene permisos para guardar en la ruta seleccionada:\n{path}", parent=win)
                return
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo exportar CSV:\n{e}", parent=win)
                return

            messagebox.showinfo("Exportado", f"CSV guardado en:\n{path}", parent=win)

        _icon_btn(btns_frame, "🔄", "Refrescar", _refrescar).pack(side=tk.LEFT)
        _icon_btn(btns_frame, "📤", "Exportar CSV", _exportar_csv).pack(side=tk.LEFT, padx=4)
        _icon_btn(btns_frame, "⚙", "Sincronizar Programaciones", lambda: self._sincronizar_programaciones_componentes()).pack(side=tk.LEFT, padx=4)
        _icon_btn(btns_frame, "❌", "Cerrar", win.withdraw).pack(side=tk.RIGHT)

        def _vincular_programacion():
            # Funcionalidad unificada por botón, se redirige a sincronización completa
            self._sincronizar_programaciones_componentes()
            return

            uso_rows = []
            maq_ids = set()
            for row_id in sel:
                vals = tree.item(row_id, 'values')
                if not vals or len(vals) < 3:
                    continue
                try:
                    cur_maq_id = int(vals[1])
                except Exception:
                    continue
                maq_ids.add(cur_maq_id)
                uso_rows.append({
                    'fecha': vals[0] if vals[0] else None,
                    'maquina_id': cur_maq_id,
                    'componente': (vals[2] or '').strip(),
                    'cantidad': float(vals[3]) if len(vals) > 3 and vals[3] not in (None, '') else None,
                    'unidad': vals[4] if len(vals) > 4 and vals[4] not in (None, '') else None,
                    'historial_id': int(vals[5]) if len(vals) > 5 and vals[5] not in (None, '') else None,
                })

            if not uso_rows:
                messagebox.showerror("Error", "No hay registros válidos seleccionados", parent=win)
                return

            if len(maq_ids) > 1:
                messagebox.showwarning(
                    "Aviso",
                    "Debe seleccionar registros de contabilidad de la misma máquina en una sola vez",
                    parent=win
                )
                return

            maq_id = next(iter(maq_ids))
            componente = uso_rows[0]['componente']
            if not componente:
                messagebox.showwarning("Aviso", "El componente asociado está vacío", parent=win)
                return

            from .services.ingenieria_extras import listar_programaciones
            progs = listar_programaciones(maquina_id=maq_id)
            componente_lower = componente.lower()

            candidatos = []
            for p in progs:
                if not p.componente:
                    continue
                comp_name = p.componente.material.nombre if p.componente.material else p.componente.nombre
                if not comp_name:
                    continue
                comp_lower = comp_name.strip().lower()
                if componente_lower == comp_lower or componente_lower in comp_lower or comp_lower in componente_lower:
                    candidatos.append(p)

            if not candidatos:
                messagebox.showinfo(
                    "No encontrado",
                    f"No se encontró programación para Máquina {maq_id} y Componente '{componente}'",
                    parent=win
                )
                return

            target = None
            if len(candidatos) == 1:
                target = candidatos[0]
            else:
                sel_win = tk.Toplevel(win)
                sel_win.title("Elija la programación a actualizar")
                sel_win.geometry("700x380")
                sel_win.config(bg=theme.current.get("bg"))

                tk.Label(sel_win, text=f"Máquina {maq_id} - componente '{componente}'", bg=theme.current.get("frame_bg")).pack(fill=tk.X, padx=8, pady=6)

                sel_tree = ttk.Treeview(sel_win, columns=("id","tipo","estado","horometro"), show="headings", selectmode="browse")
                for key, head, width in [("id","ID",70), ("tipo","Tarea",260), ("estado","Estado",100), ("horometro","Horómetro objetivo",120)]:
                    sel_tree.heading(key, text=head)
                    sel_tree.column(key, width=width, anchor="w")
                sel_tree.pack(fill=tk.BOTH, expand=True, padx=8, pady=6)

                for p in candidatos:
                    freq = getattr(p, 'frecuencia_horas', '') or ''
                    horo = getattr(p, 'horometro_objetivo', '') or ''
                    sel_tree.insert('', tk.END, iid=str(p.id), values=(p.id, getattr(p,'tipo',''), getattr(p,'estado_repuesto',''), horo or freq))

                chosen = {'id': None}

                def elegir():
                    s = sel_tree.selection()
                    if not s:
                        messagebox.showwarning("Aviso", "Seleccione una programación", parent=sel_win)
                        return
                    chosen['id'] = int(s[0])
                    sel_win.destroy()

                btn_frame = tk.Frame(sel_win, bg=theme.current.get("frame_bg"))
                btn_frame.pack(fill=tk.X, padx=8, pady=8)
                _icon_btn(btn_frame, "✅", "Aceptar", elegir).pack(side=tk.LEFT, padx=8)
                _icon_btn(btn_frame, "❌", "Cancelar", sel_win.destroy).pack(side=tk.RIGHT, padx=8)

                sel_win.wait_window()
                if not chosen['id']:
                    return
                target = next((p for p in candidatos if p.id == chosen['id']), None)

            if not target:
                return

            appended = self._aplicar_uso_a_programacion(target, uso_rows)
            if appended:
                messagebox.showinfo("Actualizado", f"Programación {target.id} actualizada con {len(uso_rows)} registros", parent=win)
            else:
                messagebox.showinfo("Sin cambios", f"No se aplicaron registros nuevos en programación {target.id}", parent=win)

            self._safe_refrescar_programaciones()

        cols = ("fecha", "maquina", "componente", "codigo", "cantidad", "unidad", "hora_inicio", "hora_fin", "horometro", "historial")
        tree = ttk.Treeview(win, columns=cols, show="headings")
        for c in cols:
            heading = c.capitalize()
            if c == "hora_inicio":
                heading = "Hora de inicio"
            elif c == "hora_fin":
                heading = "Hora de fin"
            elif c == "horometro":
                heading = "Horómetro"
            elif c == "maquina":
                heading = "Máquina"
            elif c == "historial":
                heading = "ID Historial"
            elif c == "codigo":
                heading = "Código"
            tree.heading(c, text=heading)
            tree.column(c, width=100, anchor="w")

        vsb = ttk.Scrollbar(win, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(win, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tree.pack(fill=tk.BOTH, expand=True, side=tk.LEFT, padx=10, pady=5)
        vsb.pack(fill=tk.Y, side=tk.RIGHT)
        hsb.pack(fill=tk.X, side=tk.BOTTOM)

        summary_label = tk.Label(win, text="", anchor="w", bg=theme.current.get("frame_bg"))
        summary_label.pack(fill=tk.X, padx=10)

        _refrescar()

        win.protocol("WM_DELETE_WINDOW", win.withdraw)

    def _buscar_uso_componentes_para_programacion(self, programacion):
        import re
        from .services.ingenieria_extras import obtener_uso_componentes

        def _normalize_component(comp):
            if not comp:
                return ""
            c = str(comp).strip().lower()
            m = re.search(r"\[([^\]]+)\]", c)
            if m:
                val = m.group(1).strip()
                if val:
                    c = val.lower()
            c = re.sub(r"^\d+\s*-\s*", "", c)
            c = re.sub(r"\s*\(.*$", "", c)
            return c.strip()

        if not programacion or not getattr(programacion, "maquina_id", None):
            return []

        comp_asoc = ""
        if getattr(programacion, "componente", None):
            if getattr(programacion.componente, "material", None):
                comp_asoc = (programacion.componente.material.nombre or "").strip()
            else:
                comp_asoc = (programacion.componente.nombre or "").strip()

        if not comp_asoc:
            return []

        uso = obtener_uso_componentes(maquina_id=programacion.maquina_id)

        comp_norm = _normalize_component(comp_asoc)
        if not comp_norm:
            comp_norm = comp_asoc.strip().lower()

        # match parcial con la descripción de componente de contabilidad
        results = []
        for u in uso:
            ucomp = _normalize_component(u.get("componente") or "")
            if not ucomp:
                continue
            if comp_norm == ucomp or comp_norm in ucomp or ucomp in comp_norm:
                results.append(u)
        return results

    def _aplicar_uso_a_programacion(self, programacion, uso_items):
        if not programacion or not uso_items:
            return False

        import json
        from datetime import datetime
        from zoneinfo import ZoneInfo
        from .services.ingenieria_extras import actualizar_programacion

        registros = []
        try:
            registros = json.loads(programacion.mantenimientos_realizados or "[]")
        except Exception:
            registros = []

        nuevo_registros = list(registros)
        aplicados = 0
        for u in uso_items:
            fecha_val = u.get("fecha")
            if fecha_val is None:
                fecha_str = None
            elif isinstance(fecha_val, str):
                try:
                    fecha_str = datetime.fromisoformat(fecha_val).isoformat()
                except Exception:
                    fecha_str = fecha_val
            elif hasattr(fecha_val, "isoformat"):
                fecha_str = fecha_val.isoformat()
            else:
                fecha_str = str(fecha_val)

            registro = {
                "fecha": fecha_str,
                "componente": u.get("componente"),
                "cantidad": u.get("cantidad"),
                "unidad": u.get("unidad"),
                "historial_id": u.get("historial_id"),
            }
            if registro not in nuevo_registros:
                nuevo_registros.append(registro)
                aplicados += 1

        if aplicados == 0:
            return False

        actualizar_programacion(
            programacion.id,
            mantenimientos_realizados=json.dumps(nuevo_registros, ensure_ascii=False),
            ultima_fecha=datetime.now(ZoneInfo("America/Lima"))
        )
        self._safe_refrescar_programaciones()
        return True

    def _dialogo_vincular_programacion_desde_programaciones(self):
        sel = self.prog_tree.selection()
        if not sel:
            messagebox.showwarning("Aviso", "Seleccione al menos una programación en la pestaña Programaciones", parent=self.panel)
            return

        try:
            prog_id = int(sel[0])
        except Exception:
            messagebox.showerror("Error", "Id de programación inválida", parent=self.panel)
            return

        from .services.ingenieria_extras import listar_programaciones
        progs = listar_programaciones()
        target = next((p for p in progs if getattr(p, 'id', None) == prog_id), None)

        if not target:
            messagebox.showinfo("No encontrado", "Programación seleccionada no existe", parent=self.panel)
            return

        matches = self._buscar_uso_componentes_para_programacion(target)
        if not matches:
            messagebox.showinfo("Sin coincidencias", "No se encontró uso de componentes en contabilidad para esta programación", parent=self.panel)
            return

        dlg = tk.Toplevel(self.panel)
        dlg.title("Vincular Programación con Contabilidad")
        dlg.geometry("750x420")
        dlg.config(bg=theme.current.get("bg"))

        tk.Label(dlg, text=f"Programación: {prog_id} - {getattr(target, 'tipo', '-')}", bg=theme.current.get("frame_bg")).pack(fill=tk.X, padx=10, pady=6)

        frame = tk.Frame(dlg)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=4)

        tree_cols = ("fecha", "cantidad", "unidad", "historial")
        usages_tree = ttk.Treeview(frame, columns=tree_cols, show="headings", selectmode="extended")
        for c in tree_cols:
            usages_tree.heading(c, text=c.capitalize())
            usages_tree.column(c, width=150, anchor="w")

        usages_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=usages_tree.yview)
        usages_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        for u in matches:
            usages_tree.insert("", tk.END, values=(
                u.get("fecha") and u.get("fecha").strftime("%Y-%m-%d %H:%M") or "",
                u.get("cantidad") or "",
                u.get("unidad") or "",
                u.get("historial_id") or "",
            ))

        auto_var = tk.BooleanVar(value=True)
        tk.Checkbutton(dlg, text="Aplicar automáticamente todos los registros", variable=auto_var, bg=theme.current.get("frame_bg")).pack(anchor="w", padx=10, pady=6)

        def aplicar():
            from datetime import datetime as _dt

            if auto_var.get():
                ok = self._aplicar_uso_a_programacion(target, matches)
                if ok:
                    messagebox.showinfo("Actualizado", f"Programación {prog_id} actualizada con {len(matches)} registros de contabilidad", parent=dlg)
                else:
                    messagebox.showinfo("Sin cambios", "No se aplicó ningún registro nuevo", parent=dlg)
                dlg.destroy()
                return

            sel_items = usages_tree.selection()
            if not sel_items:
                messagebox.showwarning("Aviso", "Seleccione al menos un registro de uso o active la opción automática", parent=dlg)
                return

            selected = []
            for iid in sel_items:
                vals = usages_tree.item(iid, 'values')
                selected.append({
                    "fecha": _dt.fromisoformat(vals[0]) if vals[0] else None,
                    "cantidad": float(vals[1]) if vals[1] else None,
                    "unidad": vals[2] or None,
                    "historial_id": int(vals[3]) if vals[3] else None,
                    "componente": (getattr(target, 'componente', None) and (getattr(target.componente, 'material', None) and getattr(target.componente.material, 'nombre', '') or getattr(target.componente, 'nombre', ''))),
                })

            ok = self._aplicar_uso_a_programacion(target, selected)
            if ok:
                messagebox.showinfo("Actualizado", f"Programación {prog_id} actualizada", parent=dlg)
            else:
                messagebox.showinfo("Sin cambios", "No se aplicó ningún registro nuevo", parent=dlg)
            dlg.destroy()

        btn_frame = tk.Frame(dlg, bg=theme.current.get("frame_bg"))
        btn_frame.pack(fill=tk.X, padx=10, pady=8)
        _icon_btn(btn_frame, "✅", "Aplicar", aplicar).pack(side=tk.LEFT)
        _icon_btn(btn_frame, "❌", "Cerrar", dlg.destroy).pack(side=tk.RIGHT)

        dlg.protocol("WM_DELETE_WINDOW", dlg.destroy)

    def _sincronizar_programaciones_componentes(self):
        import re
        import json
        from datetime import datetime
        from zoneinfo import ZoneInfo
        from .services.ingenieria_extras import listar_programaciones, obtener_uso_componentes, actualizar_programacion, _get_realizados
        from .models import Historial
        from .database import SessionLocal

        def _normalize_component(comp):
            if not comp:
                return ""
            c = str(comp).strip().lower()
            m = re.search(r"\[([^\]]+)\]", c)
            if m:
                val = m.group(1).strip()
                if val:
                    c = val.lower()
            c = re.sub(r"^\d+\s*-\s*", "", c)
            c = re.sub(r"\s*\(.*$", "", c)
            return c.strip()

        usos = obtener_uso_componentes()
        if not usos:
            messagebox.showinfo("Sin datos", "No hay uso de componentes disponibles", parent=self.panel)
            return

        uso_map = {}
        for u in usos:
            maq_id = u.get('maquina_id')
            comp = _normalize_component(u.get('componente'))
            horo = u.get('horometro')
            if maq_id is None or not comp or horo is None:
                continue
            key = (maq_id, comp)
            uso_map.setdefault(key, set()).add(float(horo))

        progs = listar_programaciones()
        if not progs:
            messagebox.showinfo("Sin programaciones", "No hay programaciones registradas", parent=self.panel)
            return

        updated = 0
        for p in progs:
            if not getattr(p, 'maquina_id', None) or not getattr(p, 'componente', None):
                continue

            comp_name = ''
            if p.componente:
                if getattr(p.componente, 'material', None):
                    comp_name = (p.componente.material.nombre or '').strip()
                else:
                    comp_name = (p.componente.nombre or '').strip()
            comp_name = comp_name.lower()
            comp_name = _normalize_component(comp_name)
            if not comp_name:
                continue

            key = (p.maquina_id, comp_name)
            contab_horometros = uso_map.get(key, set())
            if not contab_horometros:
                continue

            # Actualizar mantenimientos_realizados con horómetros exactos de contabilidad
            actuales = set(_get_realizados(p) or [])
            nuevos = actuales.union(contab_horometros)
            if nuevos != actuales:
                try:
                    actualizar_programacion(
                        p.id,
                        mantenimientos_realizados=json.dumps(sorted(nuevos, key=float), ensure_ascii=False),
                        ultima_fecha=datetime.now(ZoneInfo("America/Lima"))
                    )
                    updated += 1
                except Exception:
                    continue

        if updated:
            messagebox.showinfo("Sincronizado", f"Aplicados {updated} mantenimientos de contabilidad a programaciones", parent=self.panel)
        else:
            messagebox.showinfo("Sincronizado", "No se encontraron coincidencias con horómetros aplicables.", parent=self.panel)

        self._safe_refrescar_programaciones()

    def _refresh_filtro_maquinas(self):
        """Actualiza el combo de filtro de máquinas (y su lista de opciones)."""
        if not hasattr(self, '_filtro_maquina_combo'):
            return

        # Asegurarse de tener una lista válida de máquinas y sincronizar con la DB.
        try:
            from .services.ingenieria_service import listar_maquinas
            self.maquinas = listar_maquinas() or []
        except Exception:
            # Si falla la consulta, no romper la UI.
            if not hasattr(self, 'maquinas') or self.maquinas is None:
                self.maquinas = []

        vals = ["Todos"]
        for m in self.maquinas:
            try:
                if not m:
                    continue
                vals.append(f"{m.id} - {m.nombre}")
            except Exception:
                # Ignorar entradas corruptas en la lista de máquinas
                continue

        self._filtro_maquina_combo['values'] = vals

        # Sincroniza también el filtro de Registro Diario
        if hasattr(self, 'reg_maquina_combo'):
            self.reg_maquina_combo['values'] = vals
            if self.reg_maquina_var.get() not in vals:
                self.reg_maquina_var.set('Todos')

        cur = self._filtro_maquina_var.get()
        if cur not in vals:
            self._filtro_maquina_var.set("Todos")

        self._on_filtro_maquina_change()

    def _on_filtro_maquina_change(self):
        """Callback cuando cambia el filtro de máquinas (combo)."""
        sel = self._filtro_maquina_var.get()
        if sel and sel != "Todos":
            try:
                mid = int(sel.split(maxsplit=1)[0])
                self._filtro_maquinas_ids = [mid]
            except Exception:
                self._filtro_maquinas_ids = None
        else:
            self._filtro_maquinas_ids = None

        self._refresh_componentes()

    def _refresh_componentes(self):
        """Actualiza la vista única de componentes (filtrada por máquina si se indicó)."""
        from .services.ingenieria_extras import listar_componentes

        self.componentes = listar_componentes(None) or []
        self.comp_map = {c.id: c for c in self.componentes}

        # Aplicar filtro por máquinas (si se definió)
        filtro_ids = self._filtro_maquinas_ids

        self.comp_tree.delete(*self.comp_tree.get_children())

        # Llenar con los componentes filtrados
        for comp in self.componentes:
            if filtro_ids is not None and getattr(comp, 'maquina_id', None) not in filtro_ids:
                continue

            maq = next((m for m in self.maquinas if m.id == comp.maquina_id), None)
            maquina_nombre = maq.nombre if maq else f"ID {comp.maquina_id}"

            mat = getattr(comp, 'material', None)
            codigo = getattr(mat, 'codigo', '') if mat else ''
            tipo = getattr(mat, 'tipo', '') if mat else ''
            sistema = getattr(comp, 'sistema', '') or ''

            self.comp_tree.insert(
                "",
                "end",
                iid=str(getattr(comp, 'id', '')),
                values=(
                    maquina_nombre,
                    comp.nombre or '',
                    codigo,
                    tipo,
                    sistema,
                    comp.descripcion or '',
                ),
            )

        # Actualizar etiqueta de filtro si es necesario
        if filtro_ids is None:
            self._filtro_maquinas_label.config(text="Máquinas: todas")
        else:
            names = [m.nombre for m in self.maquinas if m.id in filtro_ids]
            self._filtro_maquinas_label.config(text=f"Máquinas: {', '.join(names)}")

        # Reset selection UI
        self.comp_selection_label.config(text="Selecciona uno o más componentes para generar sistema")
        self._generar_sistema_btn.config(state=tk.DISABLED)
        if hasattr(self, '_editar_sistema_btn'):
            self._editar_sistema_btn.config(state=tk.DISABLED)

    def _refresh_componentes_listado(self):
        """Refresh the components list, choosing the correct view mode.

        The UI supports two different components list modes:
        - Modern list view (columns: maquina, nombre, codigo, tipo, sistema, descripcion)
        - Legacy matrix view (columns include "_num" + one col per machine)

        When editing components we want to refresh only the view currently in use.
        """
        if hasattr(self, 'comp_tree'):
            cols = self.comp_tree.cget('columns')
            if isinstance(cols, (list, tuple)) and "_num" in cols:
                try:
                    self._actualiza_componente_listado()
                except Exception:
                    pass
                return

        # Default to modern list view
        try:
            self._refresh_componentes()
        except Exception:
            pass

    def _on_select_componentes(self, event=None):
        """Callback al seleccionar componentes: habilita el botón adecuado."""
        sel = self.comp_tree.selection()
        if not sel:
            self.comp_selection_label.config(text="Selecciona uno o más componentes para generar sistema")
            self._generar_sistema_btn.config(state=tk.DISABLED)
            if hasattr(self, '_editar_sistema_btn'):
                self._editar_sistema_btn.config(state=tk.DISABLED)
            return

        self.comp_selection_label.config(text=f"{len(sel)} componente(s) seleccionado(s)")

        # Si todos los componentes seleccionados ya tienen sistema, permitir editar sistema.
        all_have_sistema = True
        for iid in sel:
            comp = self.comp_map.get(int(iid))
            if not comp or not getattr(comp, 'sistema', None):
                all_have_sistema = False
                break

        # Cuando existe al menos un sistema definido, permitir agregar a sistema existente.
        # Considerar también los sistemas que ya están asignados a componentes (generados por botón).
        sistemas_existentes = bool(getattr(self, 'sistemas_lista', None) or
                                   any(getattr(c, 'sistema', None) for c in self.componentes))

        if all_have_sistema:
            self._generar_sistema_btn.config(state=tk.DISABLED)
            if hasattr(self, '_editar_sistema_btn'):
                self._editar_sistema_btn.config(state=tk.NORMAL)
            if hasattr(self, '_añadir_a_sistema_existente_btn'):
                self._añadir_a_sistema_existente_btn.config(state=tk.DISABLED)
        else:
            has_sin_sistema = any(
                not getattr(self.comp_map.get(int(iid)), 'sistema', None)
                for iid in sel
                if iid and iid.isdigit() and int(iid) in self.comp_map
            )
            self._generar_sistema_btn.config(state=tk.NORMAL if has_sin_sistema else tk.DISABLED)

            if hasattr(self, '_editar_sistema_btn'):
                self._editar_sistema_btn.config(state=tk.DISABLED)

            if hasattr(self, '_añadir_a_sistema_existente_btn'):
                self._añadir_a_sistema_existente_btn.config(
                    state=tk.NORMAL if has_sin_sistema and sistemas_existentes else tk.DISABLED
                )

    def _generar_sistema_desde_seleccion(self):
        """Asigna un nombre de sistema a los componentes seleccionados."""
        sel = self.comp_tree.selection()
        if not sel:
            return

        sistema = simpledialog.askstring("Generar sistema", "Nombre del sistema:", parent=self.panel)
        if not sistema:
            return

        from .services.ingenieria_extras import actualizar_componente

        count = 0
        for iid in sel:
            try:
                comp_id = int(iid)
            except Exception:
                continue
            try:
                actualizar_componente(comp_id, sistema=sistema)
                count += 1
            except Exception:
                pass

        if count:
            messagebox.showinfo("Generar sistema", f"Sistema '{sistema}' asignado a {count} componente(s).", parent=self.panel)
            from .services.ingenieria_extras import listar_sistemas
            self.sistemas_lista = listar_sistemas() or []
            self._refresh_componentes()

    def _editar_sistema_seleccion(self):
        """Edita o elimina el sistema de los componentes seleccionados."""
        sel = self.comp_tree.selection()
        if not sel:
            return

        # Determinar sistema común de selección (si existe)
        sistemas_sel = {
            getattr(self.comp_map.get(int(iid)), 'sistema', None)
            for iid in sel
            if iid and iid.isdigit() and int(iid) in self.comp_map
        }
        sistemas_sel = {s for s in sistemas_sel if s}

        if not sistemas_sel:
            messagebox.showwarning(
                "Editar sistema",
                "No hay sistema asignado a los elementos seleccionados.",
                parent=self.panel,
            )
            return

        sistema_actual = ', '.join(sorted(sistemas_sel)) if len(sistemas_sel) > 1 else next(iter(sistemas_sel))

        # Dialogo con opciones: renombrar o eliminar sistema
        dlg = tk.Toplevel(self.panel)
        dlg.title("Editar sistema")
        dlg.transient(self.panel)
        dlg.grab_set()

        tk.Label(dlg, text=f"Sistema seleccionado: {sistema_actual}").pack(padx=10, pady=8)

        accion_var = tk.StringVar(value="rename")
        tk.Radiobutton(dlg, text="Cambiar nombre", variable=accion_var, value="rename").pack(anchor='w', padx=10)
        tk.Radiobutton(dlg, text="Eliminar sistema", variable=accion_var, value="clear").pack(anchor='w', padx=10)

        nombre_var = tk.StringVar(value=sistema_actual if len(sistemas_sel) == 1 else "")
        entry = tk.Entry(dlg, textvariable=nombre_var)
        entry.pack(fill=tk.X, padx=10, pady=(8, 10))

        resultado = {'accion': None, 'nombre': None}

        def _ok():
            resultado['accion'] = accion_var.get()
            resultado['nombre'] = nombre_var.get().strip()
            dlg.destroy()

        def _cancel():
            dlg.destroy()

        btns = tk.Frame(dlg)
        btns.pack(pady=8)
        tk.Button(btns, text="Aceptar", command=_ok).pack(side=tk.LEFT, padx=5)
        tk.Button(btns, text="Cancelar", command=_cancel).pack(side=tk.LEFT, padx=5)

        dlg.wait_window()

        if not resultado.get('accion'):
            return

        from .services.ingenieria_extras import actualizar_componente

        count = 0
        if resultado['accion'] == 'rename':
            nuevo_sistema = resultado['nombre']
            if not nuevo_sistema:
                messagebox.showwarning(
                    "Editar sistema",
                    "Debe ingresar un nombre de sistema para renombrar.",
                    parent=self.panel,
                )
                return

            for iid in sel:
                try:
                    comp_id = int(iid)
                except Exception:
                    continue
                try:
                    actualizar_componente(comp_id, sistema=nuevo_sistema)
                    count += 1
                except Exception:
                    pass

            if count:
                messagebox.showinfo(
                    "Editar sistema",
                    f"Sistema renombrado a '{nuevo_sistema}' en {count} componente(s).",
                    parent=self.panel,
                )

        else:  # clear
            for iid in sel:
                try:
                    comp_id = int(iid)
                except Exception:
                    continue
                try:
                    actualizar_componente(comp_id, sistema=None)
                    count += 1
                except Exception:
                    pass

            if count:
                messagebox.showinfo(
                    "Editar sistema",
                    f"Se quitó el sistema de {count} componente(s).",
                    parent=self.panel,
                )

        if count:
            from .services.ingenieria_extras import listar_sistemas
            self.sistemas_lista = listar_sistemas() or []
            self._refresh_componentes()

    def _agregar_seleccion_a_sistema_existente(self):
        """Añade los componentes seleccionados a un sistema existente."""
        # Los sistemas válidos se derivan de los componentes que ya tienen un sistema.
        try:
            from .services.ingenieria_extras import listar_componentes
            comps = listar_componentes(None) or []
        except Exception:
            comps = []

        systems = sorted({(getattr(c, 'sistema', None) or '').strip() for c in comps if getattr(c, 'sistema', None)})

        if not systems:
            messagebox.showwarning(
                "Aviso",
                "No hay sistemas existentes (no hay componentes con sistema asignado).",
                parent=self.panel,
            )
            return

        dialog = tk.Toplevel(self.panel)
        dialog.title("Seleccionar sistema")
        dialog.transient(self.panel)
        dialog.grab_set()

        tk.Label(dialog, text="Seleccione un sistema existente:").pack(padx=10, pady=(10, 4))
        sys_var = tk.StringVar(value=systems[0])
        sys_combo = ttk.Combobox(dialog, values=systems, textvariable=sys_var, state="readonly")
        sys_combo.pack(fill=tk.X, padx=10, pady=4)

        result = {"system": None}

        def _ok():
            result["system"] = sys_var.get().strip()
            dialog.destroy()

        def _cancel():
            dialog.destroy()

        btns = tk.Frame(dialog)
        btns.pack(pady=10)
        tk.Button(btns, text="Aceptar", command=_ok).pack(side=tk.LEFT, padx=5)
        tk.Button(btns, text="Cancelar", command=_cancel).pack(side=tk.LEFT, padx=5)

        dialog.wait_window()

        sistema_sel = result.get("system")
        if not sistema_sel:
            return

        from .services.ingenieria_extras import actualizar_componente

        sel = self.comp_tree.selection()
        count = 0
        for iid in sel:
            try:
                comp_id = int(iid)
            except Exception:
                continue
            try:
                actualizar_componente(comp_id, sistema=sistema_sel)
                count += 1
            except Exception:
                pass

        if count:
            messagebox.showinfo(
                "Asignado",
                f"{count} componente(s) asignado(s) al sistema '{sistema_sel}'.",
                parent=self.panel,
            )
            self._refresh_componentes()

    def _exportar_componentes(self):
        """Exporta componentes a un CSV."""
        from .services.ingenieria_extras import exportar_componentes_csv

        path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv")])
        if not path:
            return

        try:
            exportar_componentes_csv(path)
            messagebox.showinfo("Exportado", f"CSV guardado en:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _importar_componentes(self):
        """Importa componentes desde un CSV."""
        from .services.ingenieria_extras import importar_componentes_csv

        path = filedialog.askopenfilename(filetypes=[("CSV", "*.csv")])
        if not path:
            return

        try:
            resultado = importar_componentes_csv(path)
            messagebox.showinfo(
                "Importado",
                f"Importados {resultado.get('created', 0)} nuevos, actualizados {resultado.get('updated', 0)}, omitidos {resultado.get('skipped', 0)}.",
                parent=self.panel,
            )
            self._refresh_componentes()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo importar:\n{e}", parent=self.panel)


    def _actualiza_componente_listado_por_sistema(self):

        from .services.ingenieria_extras import listar_materiales

        materiales = listar_materiales()

        # Mantener tipo de material para colorear filas, pero no usamos su "sistema" para
        # clasificar componentes globalmente (el sistema es una clasificación interna).
        material_tipo = {m.id: m.tipo for m in materiales}

        NO_SISTEMA_TAB = "Sin sistema"

        # Detectar qué sistemas aparecen en los componentes (no los guardamos en DB).
        self._component_sistemas = {
            getattr(comp, "sistema", None) or NO_SISTEMA_TAB
            for comp in self.componentes
        }

        # Asegurar que la vista incluya tab para cada sistema definido y una tab "Sin sistema".
        desired_tabs = set(getattr(self, "sistemas_lista", []) or [])
        desired_tabs |= self._component_sistemas
        desired_tabs.add(NO_SISTEMA_TAB)

        existing_tabs = set(self.sistemas_tabs.keys())
        if desired_tabs != existing_tabs:
            # Si hay sistemas nuevos (incluyendo los que surgen en componentes), reconstruir pestañas.
            self._rebuild_tab_componentes()
            # Después de reconstruir, volver a poblar componentes.
            self._refresh_componentes()
            return

        # limpiar árboles
        for tree in self.comp_trees.values():
            tree.delete(*tree.get_children())

        contador = {k: 0 for k in self.comp_trees}

    # poblar
        # aplicar filtro por máquinas (si está definido)
        filtro_ids = self._filtro_maquinas_ids

        # Track materials usados por sistema (para mostrar mensaje relativo a inventario)
        materiales_usados_por_sistema = {k: set() for k in self.comp_trees}

        for comp in self.componentes:

            if filtro_ids is not None and getattr(comp, 'maquina_id', None) not in filtro_ids:
                continue

            mat_id = getattr(comp, 'material_id', None)
            sistema = getattr(comp, 'sistema', None) or "Sin sistema"

            tree = self.comp_trees.get(sistema)
            if not tree:
                continue

            tipo = None
            if mat_id is not None:
                tipo = material_tipo.get(mat_id)

            # Mostrar datos basados en el material asociado (nombre/tipo/código)
            tree.insert(
                "",
                "end",
                iid=str(getattr(comp, 'id', '')),
                values=(
                    getattr(comp, 'material', None) and getattr(comp.material, 'nombre', '') or comp.nombre or "",
                    getattr(comp, 'material', None) and getattr(comp.material, 'codigo', ''),
                    getattr(comp, 'material', None) and getattr(comp.material, 'tipo', ''),
                    sistema or "",
                    getattr(comp, 'material', None) and getattr(comp.material, 'descripcion', '') or comp.descripcion or "",
                ),
                tags=(tipo,)
            )

            contador[sistema] += 1
            materiales_usados_por_sistema.setdefault(sistema, set()).add(mat_id)

        # Si no hay componentes, mostrar mensaje (y opcionalmente cuántos materiales hay en inventario)
        materiales_por_sistema = {}
        for mat in materiales:
            sistemas = materiales_por_sistema.setdefault(getattr(mat, 'sistema', None), set())
            sistemas.add(mat.id)

        for sistema, tree in self.comp_trees.items():
            if tree.get_children():
                continue

            total_mat = len(materiales_por_sistema.get(sistema, set()))
            usados = len(materiales_usados_por_sistema.get(sistema, set()))
            disponibles = max(total_mat - usados, 0)
            msg = "No hay componentes instalados"
            if disponibles > 0:
                msg += f" (materiales disponibles: {disponibles})"
            tree.insert(
                "",
                "end",
                values=("", "", "", msg),
                tags=("no_componentes",)
            )

    # actualizar contador en pestañas
        for sistema, count in contador.items():

            idx = self.sistemas_tabs.get(sistema)

            if idx is not None:

                self.sistemas_notebook.tab(
                    idx,
                    text=f"Sistema {sistema} ({count})"
                )






    # Nota: la clasificación de sistemas por IA se ha eliminado de la UI.
    # Los sistemas se generan únicamente desde el botón "Generar Sistema".

    def _rebuild_tab_componentes(self):
        """Reconstruye la pestaña de Componentes para reflejar cambios en los sistemas."""
        if hasattr(self, "tab_componentes"):
            try:
                self.notebook.forget(self.tab_componentes)
            except Exception:
                pass
            self.tab_componentes.destroy()
        self._crea_tab_componentes(self.notebook)

        # Asegurar que las nuevas widgets usen el tema actual (p. ej. modo oscuro)
        try:
            from . import theme
            theme.apply_theme(self.tab_componentes)
        except Exception:
            pass

        try:
            self.notebook.select(self.tab_componentes)
        except Exception:
            pass


    # NOTE: Gestionar sistemas / clasificar sistemas se eliminaron de la UI.
    # Los sistemas ahora se generan únicamente desde la acción "Generar Sistema"
    # a partir de componentes seleccionados en la lista.

        # Use la lista persistida de sistemas (base de datos). No volver a valores por defecto.
        sistemas_actuales = list(getattr(self, "sistemas_lista", []) or listar_sistemas() or [])

        def _ask_string(title: str, prompt: str, initialvalue: str = "") -> str | None:
            """Pide un string usando un diálogo que respeta el tema (oscuro/claro)."""
            dlg = tk.Toplevel(self.panel)
            dlg.title(title)
            dlg.transient(self.panel)
            dlg.grab_set()
            dlg.resizable(False, False)

            frame = tk.Frame(dlg, padx=10, pady=10)
            frame.pack(fill=tk.BOTH, expand=True)

            tk.Label(frame, text=prompt, anchor="w").pack(fill=tk.X)
            entry = tk.Entry(frame)
            entry.insert(0, initialvalue or "")
            entry.pack(fill=tk.X, pady=6)
            entry.select_range(0, tk.END)
            entry.focus_set()

            result = {"value": None}

            def _ok():
                result["value"] = entry.get()
                dlg.destroy()

            def _cancel():
                dlg.destroy()

            btn_frame = tk.Frame(frame)
            btn_frame.pack(fill=tk.X, pady=4)
            tk.Button(btn_frame, text="Aceptar", command=_ok).pack(side=tk.LEFT)
            tk.Button(btn_frame, text="Cancelar", command=_cancel).pack(side=tk.LEFT, padx=6)

            # Ensure dialog uses the current theme
            try:
                from . import theme

                theme.apply_theme(dlg)
            except Exception:
                pass

            dlg.wait_window()
            return result["value"]

        def _build(content):
            tk.Label(content, text="Sistemas existentes:", font=("Arial", 10, "bold")).pack(anchor="w", padx=10, pady=5)
            lv = tk.Listbox(content, height=6)
            lv.pack(fill=tk.BOTH, padx=10, pady=2)
            for s in sistemas_actuales:
                lv.insert(tk.END, s)

            def _ui_agregar_sistema():
                nombre = _ask_string("Nuevo Sistema", "Nombre del sistema:")
                if nombre and nombre.strip():
                    nombre = nombre.strip()
                    sistemas_actuales.append(nombre)
                    lv.insert(tk.END, nombre)

            def _ui_eliminar_sistema():
                sel = lv.curselection()
                if sel:
                    idx = sel[0]
                    lv.delete(idx)
                    sistemas_actuales.pop(idx)

            def _ui_editar_sistema(event=None):
                sel = lv.curselection()
                if not sel:
                    return
                idx = sel[0]
                old = lv.get(idx)
                nuevo = _ask_string("Editar sistema", "Nombre del sistema:", initialvalue=old)
                if not nuevo:
                    return
                nuevo = nuevo.strip()
                if not nuevo or nuevo == old:
                    return
                sistemas_actuales[idx] = nuevo
                lv.delete(idx)
                lv.insert(idx, nuevo)

            btns = tk.Frame(content)
            btns.pack(pady=5)
            tk.Button(btns, text="➕ Agregar", command=_ui_agregar_sistema).pack(side=tk.LEFT, padx=5)
            tk.Button(btns, text="🗑 Eliminar", command=_ui_eliminar_sistema).pack(side=tk.LEFT, padx=5)

            # Doble click para editar sistema
            lv.bind("<Double-1>", _ui_editar_sistema)

            def guardar_cambios():
                # Guardar la lista de sistemas donde corresponda
                actuales = set(listar_sistemas())
                nuevos = set(sistemas_actuales)

                # Agregar nuevos sistemas
                for s in sorted(nuevos - actuales):
                    try:
                        agregar_sistema(s)
                    except Exception:
                        pass

                # Eliminar sistemas removidos (solo si no hay componentes asociados)
                for s in sorted(actuales - nuevos):
                    try:
                        eliminar_sistema(s)
                    except Exception:
                        pass

                # Refrescar lista local y reconstruir pestañas
                self.sistemas_lista = listar_sistemas() or []
                try:
                    self._rebuild_tab_componentes()
                except Exception:
                    pass
                self.view_mgr.pop()

            _icon_btn(content, "💾", "Guardar cambios", guardar_cambios).pack(fill=tk.X, padx=10, pady=5)

        self.view_mgr.push("Gestionar Sistemas", _build, scrollable=True)


    def _dialogo_entrenar_ia(self):
        """Diálogo simplificado para entrenar la IA con un libro/documento (PDF)."""

        # Variables necesarias por la lógica heredada (no se usan en el flujo principal).
        classifier_model = None
        classifier_labels = None
        extracted = []

        def _entrenar_con_libro():
            path = filedialog.askopenfilename(
                filetypes=[("PDF", "*.pdf")],
                parent=self.panel,
            )
            if not path:
                return

            nombre = simpledialog.askstring("Nombre del documento", "Ingrese un nombre para este documento (opcional):", parent=self.panel)
            from .services.ingenieria_service import entrenar_con_documento

            count = 0
            try:
                count = entrenar_con_documento(path, nombre=nombre, tag="libro")
            except Exception as e:
                import traceback

                tb = traceback.format_exc()
                messagebox.showerror(
                    "Error",
                    f"No se pudo indexar el documento:\n{e}\n\nDetalles:\n{tb}",
                    parent=self.panel,
                )
                return

            messagebox.showinfo(
                "Entrenamiento completo",
                f"Se indexaron {count} fragmentos del documento.\nAhora estarán disponibles para sugerencias semánticas.",
                parent=self.panel,
            )

            # Actualizar la etiqueta de estado de la IA si está presente en la ventana de programaciones
            if hasattr(self, "ia_status_label"):
                self.ia_status_label.config(text=self._get_ia_status_text())

        def _entrenar_con_csv():
            path = filedialog.askopenfilename(
                filetypes=[("CSV", "*.csv")],
                parent=self.panel,
            )
            if not path:
                return

            nombre = simpledialog.askstring("Nombre del CSV", "Ingrese un nombre para este CSV (opcional):", parent=self.panel)
            from .services.ingenieria_service import entrenar_con_csv

            count = 0
            try:
                count = entrenar_con_csv(path, nombre=nombre, tag="csv")
            except Exception as e:
                import traceback

                tb = traceback.format_exc()
                messagebox.showerror(
                    "Error",
                    f"No se pudo indexar el CSV:\n{e}\n\nDetalles:\n{tb}",
                    parent=self.panel,
                )
                return

            messagebox.showinfo(
                "Entrenamiento completo",
                f"Se indexaron {count} fragmentos del CSV.\nAhora estarán disponibles para sugerencias semánticas.",
                parent=self.panel,
            )

            # Actualizar la etiqueta de estado de la IA si está presente en la ventana de programaciones
            if hasattr(self, "ia_status_label"):
                self.ia_status_label.config(text=self._get_ia_status_text())

        def _entrenar_desde_historial():
            from .services.ingenieria_service import entrenar_desde_historial

            count = 0
            try:
                count = entrenar_desde_historial()
            except Exception as e:
                import traceback

                tb = traceback.format_exc()
                messagebox.showerror(
                    "Error",
                    f"No se pudo indexar el historial:\n{e}\n\nDetalles:\n{tb}",
                    parent=self.panel,
                )
                return

            messagebox.showinfo(
                "Entrenamiento completo",
                f"Se indexaron {count} fragmentos del historial.\nAhora estarán disponibles para sugerencias semánticas.",
                parent=self.panel,
            )

            # Actualizar la etiqueta de estado de la IA si está presente en la ventana de programaciones
            if hasattr(self, "ia_status_label"):
                self.ia_status_label.config(text=self._get_ia_status_text())

        def _build(content):
            tk.Label(
                content,
                text="Entrenar IA",
                font=("Arial", 11, "bold"),
            ).pack(anchor="w", padx=10, pady=5)

            tk.Label(
                content,
                text=(
                    "Para indexar manuales de una máquina, use el botón "
                    "\"Adjuntar manual (PDF)\" en la vista de detalles de la máquina.\n\n"
                    "Para indexar un libro o documento general sin asignar a una máquina, "
                    "use el siguiente botón."
                ),
                justify="left",
                wraplength=500,
            ).pack(anchor="w", padx=10, pady=5)

            tk.Button(
                content,
                text="📚 Entrenar IA con libro (PDF)",
                command=_entrenar_con_libro,
                bg=theme.current.get("button_bg"),
                fg=theme.current.get("button_fg"),
            ).pack(fill=tk.X, padx=10, pady=(8, 5))

            tk.Button(
                content,
                text="📄 Entrenar IA con CSV",
                command=_entrenar_con_csv,
                bg=theme.current.get("button_bg"),
                fg=theme.current.get("button_fg"),
            ).pack(fill=tk.X, padx=10, pady=(0, 5))

            tk.Button(
                content,
                text="🧠 Entrenar IA desde historial",
                command=_entrenar_desde_historial,
                bg=theme.current.get("button_bg"),
                fg=theme.current.get("button_fg"),
            ).pack(fill=tk.X, padx=10, pady=(0, 5))

            # ------------------------------------------------
            # Consulta inteligente (búsqueda semántica rápida)
            # ------------------------------------------------
            sep = tk.Frame(content, height=1, bg="#ccc")
            sep.pack(fill=tk.X, padx=10, pady=10)

            query_var = tk.StringVar()
            tk.Label(content, text="Consulta inteligente (palabras clave/códigos):", font=("Arial", 10, "bold")).pack(anchor="w", padx=10)
            query_entry = tk.Entry(content, textvariable=query_var)
            query_entry.pack(fill=tk.X, padx=10, pady=4)

            results_box = tk.Listbox(content, height=8)
            results_box.pack(fill=tk.BOTH, padx=10, pady=4, expand=True)

            def _buscar_inteligente(event=None):
                q = query_var.get().strip()
                results_box.delete(0, tk.END)
                if not q:
                    return
                try:
                    from .services.embeddings import search_embeddings, format_embedding_suggestion

                    def _filtrar(item, score):
                        meta = item.metadata or {}
                        tipo = (meta.get("tipo") or "").lower()
                        if tipo in ("manual", "training", "libro", "documento", "historial"):
                            return True
                        return bool(meta.get("fragmento"))

                    res = search_embeddings(q, top_k=10, filter_fn=_filtrar)
                    for r in res:
                        md = r.get("metadata", {})
                        frag = md.get("fragmento") or md.get("descripcion") or ""
                        if not frag:
                            continue
                        label = format_embedding_suggestion(md, frag[:180])
                        results_box.insert(tk.END, f"{label} (score {r.get('score', 0):.2f})")
                    if res:
                        results_box.selection_set(0)
                except Exception:
                    pass

            query_entry.bind("<Return>", _buscar_inteligente)

            tk.Button(
                content,
                text="🔎 Buscar",
                command=_buscar_inteligente,
                bg=theme.current.get("button_bg"),
                fg=theme.current.get("button_fg"),
            ).pack(fill=tk.X, padx=10, pady=(0, 5))

        self.view_mgr.push("Entrenamiento IA", _build, scrollable=True)

        def _load_classifier():
            nonlocal classifier_model, classifier_labels
            if classifier_model is not None:
                return
            base = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "ml"))
            model_path = os.path.join(base, "system_classifier.h5")
            labels_path = os.path.join(base, "system_classifier_labels.json")
            if not os.path.exists(model_path) or not os.path.exists(labels_path):
                return
            try:
                import tensorflow as tf
                classifier_model = tf.keras.models.load_model(model_path)
                with open(labels_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                classifier_labels = data.get("labels", [])
            except Exception:
                classifier_model = None
                classifier_labels = None

        def _predict_system(text):
            if classifier_model is None:
                _load_classifier()
            if not classifier_model:
                return None, None
            try:
                preds = classifier_model.predict([text], verbose=0)[0]
                idx = int(preds.argmax())
                label = classifier_labels[idx] if classifier_labels and idx < len(classifier_labels) else str(idx)
                return label, float(preds[idx])
            except Exception:
                return None, None

        def _get_or_create_material(code, sistema=None):
            # Favor reutilizar materiales existentes por código o nombre.
            from .services.ingenieria_extras import listar_materiales, crear_material, actualizar_material

            mats = listar_materiales() or []
            mat = next((m for m in mats if (getattr(m, 'codigo', None) or '').strip() == code.strip() or (getattr(m, 'nombre', None) or '').strip() == code.strip()), None)
            if mat:
                # Si no tiene sistema y tenemos uno sugerido, actualizamos.
                if sistema and not getattr(mat, 'sistema', None):
                    try:
                        actualizar_material(mat.id, sistema=sistema)
                    except Exception:
                        pass
                return mat.id

            try:
                new_mat = crear_material(
                    nombre=code,
                    codigo=code,
                    tipo="pieza",
                    unidad="",
                    descripcion="Generado automáticamente desde manual.",
                    stock_inicial=0.0,
                    stock_minimo=0.0,
                    sistema=sistema,
                )
                return new_mat.id
            except Exception:
                return None

            return None

        def _extraer_de_pdf(path):
            # Intenta usar pypdf (recomendado) y, si no está disponible, prueba PyPDF2.
            reader_cls = None
            try:
                from pypdf import PdfReader
                reader_cls = PdfReader
            except ImportError:
                try:
                    from PyPDF2 import PdfReader
                    reader_cls = PdfReader
                except ImportError:
                    messagebox.showerror(
                        "Dependencia faltante",
                        "La librería 'pypdf' o 'PyPDF2' no está instalada. Instale con:\n\n    pip install pypdf\n    (o) pip install PyPDF2",
                        parent=self.panel,
                    )
                    return []

            try:
                reader = reader_cls(path)
                text = "\n".join([p.extract_text() or "" for p in reader.pages])
            except Exception as e:
                messagebox.showerror("Error al leer PDF", str(e), parent=self.panel)
                return []

            # Extraer códigos de componente típicos (por ejemplo: ABC-1234, 1234-ABC, etc.)
            codes = re.findall(r"\b([A-Z]{1,4}-?\d{2,6}[A-Z]?)\b", text)
            seen = set()
            items = []
            for code in codes:
                if code in seen:
                    continue
                seen.add(code)
                idx = text.find(code)
                snippet = text[max(0, idx - 40): idx + len(code) + 40].replace("\n", " ").strip()
                items.append({"code": code, "context": snippet})
            return items

        def _build(content):
            tk.Label(content, text="Entrenamiento / Importación desde manual (PDF)", font=("Arial", 11, "bold")).pack(anchor="w", padx=10, pady=5)

            file_frame = tk.Frame(content)
            file_frame.pack(fill=tk.X, padx=10, pady=5)
            pdf_path_var = tk.StringVar()
            tk.Entry(file_frame, textvariable=pdf_path_var, state="readonly").pack(side=tk.LEFT, fill=tk.X, expand=True)

            selected_pdf_path = None

            def _seleccionar_pdf():
                nonlocal selected_pdf_path
                path = filedialog.askopenfilename(filetypes=[("PDF", "*.pdf")], parent=self.panel)
                if not path:
                    return
                selected_pdf_path = path
                pdf_path_var.set(path)
                nonlocal extracted
                extracted = _extraer_de_pdf(path)

                # Predecir sistema para cada componente (si existe un modelo entrenado)
                for item in extracted:
                    text = f"{item.get('code','')} {item.get('context','')}"
                    sistema, confid = _predict_system(text)
                    item['predicted_system'] = sistema
                    item['pred_conf'] = confid

                lb.delete(0, tk.END)
                for item in extracted:
                    sys_text = f"[{item.get('predicted_system')}] " if item.get('predicted_system') else ""
                    lb.insert(
                        tk.END,
                        f"{sys_text}{item['code']}  –  {item['context'][:80]}..."
                    )
                txt_context.delete("1.0", tk.END)

                # Habilitar el botón de entrenamiento ahora que hay un PDF cargado
                train_btn.config(state=tk.NORMAL)

            _icon_btn(file_frame, "📄", "Abrir PDF", _seleccionar_pdf).pack(side=tk.LEFT, padx=5)

            def _entrenar_con_libro():
                nonlocal selected_pdf_path
                path = selected_pdf_path or pdf_path_var.get()
                if not path:
                    messagebox.showwarning("Aviso", "Primero seleccione un PDF mediante 'Abrir PDF'", parent=self.panel)
                    return

                nombre = simpledialog.askstring("Nombre del documento", "Ingrese un nombre para este documento (opcional):", parent=self.panel)
                from .services.ingenieria_service import entrenar_con_documento

                count = 0
                try:
                    count = entrenar_con_documento(path, nombre=nombre, tag="libro")
                except Exception as e:
                    import traceback

                    tb = traceback.format_exc()
                    messagebox.showerror(
                        "Error",
                        f"No se pudo indexar el documento:\n{e}\n\nDetalles:\n{tb}",
                        parent=self.panel,
                    )
                    return

                messagebox.showinfo(
                    "Entrenamiento completo",
                    f"Se indexaron {count} fragmentos del documento.\nAhora estarán disponibles para sugerencias semánticas.",
                    parent=self.panel,
                )

            train_btn = tk.Button(
                content,
                text="📚 Entrenar IA con libro (PDF)",
                command=_entrenar_con_libro,
                bg=theme.current.get("button_bg"),
                fg=theme.current.get("button_fg"),
                state=tk.DISABLED,
            )
            train_btn.pack(fill=tk.X, padx=10, pady=(8, 5))

            def _exportar_excel():
                nonlocal selected_pdf_path
                path = selected_pdf_path or pdf_path_var.get()
                if not path:
                    messagebox.showwarning("Aviso", "Seleccione primero un PDF con 'Abrir PDF'", parent=self.panel)
                    return

                # Elegir archivo de salida
                out_path = filedialog.asksaveasfilename(
                    title="Guardar resumen como Excel",
                    defaultextension=".xlsx",
                    filetypes=[("Excel", "*.xlsx")],
                    parent=self.panel,
                )
                if not out_path:
                    return

                try:
                    from .services.manual_to_excel import generar_excel_desde_manual

                    # Si el usuario tiene un CSV de repuestos en la misma carpeta, lo usamos.
                    parts_csv = None
                    candidate = Path(path).with_suffix("")
                    for ext in (".csv",):
                        cand = candidate.with_suffix(ext)
                        if cand.exists():
                            parts_csv = str(cand)
                            break

                    generar_excel_desde_manual(
                        path,
                        out_path,
                        repuestos_csv=parts_csv,
                        incluir_historial=True,
                    )
                    messagebox.showinfo("Listo", f"Resumen exportado a:\n{out_path}", parent=self.panel)
                except Exception as e:
                    import traceback

                    tb = traceback.format_exc()
                    messagebox.showerror(
                        "Error",
                        f"No se pudo exportar Excel:\n{e}\n\n{tb}",
                        parent=self.panel,
                    )

            tk.Button(
                content,
                text="📄 Exportar resumen a Excel",
                command=_exportar_excel,
                bg=theme.current.get("button_bg"),
                fg=theme.current.get("button_fg"),
            ).pack(fill=tk.X, padx=10, pady=(0, 10))

            tk.Label(content, text="Máquina destino (para los componentes):", font=("Arial", 10, "bold")).pack(anchor="w", padx=10, pady=(10,0))
            maq_names = [f"{m.nombre} (ID {m.id})" for m in self.maquinas]
            maq_var = tk.StringVar()
            maq_combo = ttk.Combobox(content, values=maq_names, textvariable=maq_var, width=60)
            maq_combo.pack(fill=tk.X, padx=10, pady=2)

            tk.Label(content, text="Componentes encontrados (seleccione para ver contexto):", font=("Arial", 10, "bold")).pack(anchor="w", padx=10, pady=(10,0))
            lb = tk.Listbox(content, selectmode=tk.MULTIPLE, height=10)
            lb.pack(fill=tk.BOTH, padx=10, pady=2, expand=True)

            def _on_select(evt):
                sel = lb.curselection()
                if not sel:
                    return
                item = extracted[sel[0]]
                txt_context.delete("1.0", tk.END)
                txt_context.insert(tk.END, item.get("context") or "")

            lb.bind("<<ListboxSelect>>", _on_select)

            tk.Label(content, text="Contexto / descripción", font=("Arial", 10, "bold")).pack(anchor="w", padx=10, pady=(10,0))
            txt_context = tk.Text(content, height=5)
            txt_context.pack(fill=tk.BOTH, padx=10, pady=2, expand=True)

            def _load_training_files():
                base = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "ml"))
                return (
                    os.path.join(base, "system_classifier_keywords.json"),
                    os.path.join(base, "system_classifier_metrics.json"),
                    os.path.join(base, "components_summary.json"),
                )

            def _run_synthesis():
                # Ejecuta el script python que sintetiza los datos extraídos
                base = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "ml"))
                script = os.path.join(base, "summarize_extracted.py")
                input_json = os.path.join(base, "components.json")
                output_json = os.path.join(base, "components_summary.json")
                if not os.path.exists(script):
                    messagebox.showerror("Error", f"No se encontró el script: {script}", parent=self.panel)
                    return
                try:
                    proc = subprocess.run(
                        [sys.executable, script, "--input", input_json, "--output", output_json],
                        capture_output=True,
                        text=True,
                        check=True
                    )
                    msg = proc.stdout.strip() or proc.stderr.strip()
                    if not msg:
                        msg = "Síntesis completada"
                    messagebox.showinfo("Síntesis", msg, parent=self.panel)
                except subprocess.CalledProcessError as e:
                    messagebox.showerror("Error", e.stderr + "\n" + e.stdout, parent=self.panel)
                except Exception as e:
                    messagebox.showerror("Error", str(e), parent=self.panel)

            def _mostrar_sintesis():
                _, _, summary_path = _load_training_files()
                if not os.path.exists(summary_path):
                    messagebox.showerror("Error", "No se ha generado el resumen aún", parent=self.panel)
                    return

                try:
                    with open(summary_path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                except Exception as e:
                    messagebox.showerror("Error", str(e), parent=self.panel)
                    return

                # Construir texto de resumen
                s = []
                sumo = data.get("summary", {})
                s.append(f"Componentes totales: {sumo.get('total_components', 0)}")
                s.append(f"Códigos únicos: {sumo.get('unique_codes', 0)}")
                s.append("")

                s.append("Top códigos:")
                for c, cnt in sumo.get('top_codes', [])[:10]:
                    s.append(f"  {c} x{cnt}")
                s.append("")

                s.append("Top keywords:")
                for kw in sumo.get('top_keywords', []):
                    s.append(f"  {kw}")
                s.append("")

                s.append("Sistemas detectados:")
                for sys_name, cnt in sumo.get('systems', [])[:10]:
                    s.append(f"  {sys_name}: {cnt}")
                s.append("")

                s.append("Resumen por PDF:")
                by_pdf = sumo.get('by_pdf', {})
                for pdf, info in by_pdf.items():
                    s.append(f"  {os.path.basename(pdf)}: {info.get('count', 0)} componentes")

                s.append("")
                s.append("Ejemplos por sistema:")
                for sys_name, examples in sumo.get('examples_by_system', {}).items():
                    s.append(f"  {sys_name} ({len(examples)})")
                    for ex in examples[:2]:
                        s.append(f"    - {ex.get('code')} | {ex.get('context')}")

                # Mostrar en un cuadro de texto scrollable
                win = tk.Toplevel(self.panel)
                win.title("Síntesis de datos extraídos")
                txt = tk.Text(win, wrap=tk.WORD)
                txt.insert(tk.END, "\n".join(s))
                txt.config(state=tk.DISABLED)
                txt.pack(fill=tk.BOTH, expand=True)
                scrollbar = ttk.Scrollbar(win, orient=tk.VERTICAL, command=txt.yview)
                txt.configure(yscrollcommand=scrollbar.set)
                scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            def _compute_keywords_from_components():
                # Fallback cuando no hay archivo de keywords: usar components.json para generar algo rápido.
                base = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "ml"))
                components_path = os.path.join(base, "components.json")
                if not os.path.exists(components_path):
                    return None
                try:
                    with open(components_path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    comps = data.get("components", [])
                    # extraer texto de contexto + código
                    textos = [f"{c.get('code','')} {c.get('context','')}" for c in comps]
                    # simple tokenización
                    import re
                    stopwords = {
                        "de", "la", "el", "y", "a", "en", "los", "del", "se", "las", "por",
                        "un", "para", "con", "no", "una", "su", "al", "lo", "como", "más", "pero",
                        "sus", "le", "ya", "o", "este", "si", "porque", "esta", "entre", "cuando",
                        "muy", "sin", "sobre", "también", "me", "hasta", "hay", "donde", "quien",
                        "desde", "todo", "nos", "durante", "todos", "uno", "les", "ni", "contra",
                        "otros", "ese", "eso", "ante", "ellos",
                    }
                    from collections import Counter
                    counter = Counter()
                    for t in textos:
                        if not isinstance(t, str):
                            continue
                        words = re.findall(r"\b\w+\b", t.lower())
                        for w in words:
                            if len(w) < 3 or w.isdigit() or w in stopwords:
                                continue
                            counter[w] += 1
                    return {"(generado)": [w for w, _ in counter.most_common(15)]}
                except Exception:
                    return None

            def _mostrar_palabras_clave():
                keywords_path, _, _ = _load_training_files()
                data = None
                if os.path.exists(keywords_path):
                    try:
                        with open(keywords_path, "r", encoding="utf-8") as f:
                            data = json.load(f)
                    except Exception as e:
                        messagebox.showerror("Error", str(e), parent=self.panel)
                        return
                else:
                    data = _compute_keywords_from_components()
                    if data is None:
                        messagebox.showinfo("Palabras clave", f"No se encontró el archivo de palabras clave:\n{keywords_path}", parent=self.panel)
                        return

                lines = []
                for sistema, kws in data.items():
                    if isinstance(kws, list):
                        lines.append(f"{sistema}: {', '.join(kws[:10])}")
                messagebox.showinfo("Palabras clave", "\n".join(lines) or "No hay datos", parent=self.panel)

            def _generar_latex():
                # Genera un archivo .tex con el procedimiento y componentes extraídos
                base = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "ml"))
                out_path = os.path.join(base, "proceso_fourier.tex")

                parts = [
                    "% Documento generado automáticamente",
                    "\\section*{Procedimiento de extracción de contornos}",
                    "Este documento describe cómo extraer contornos y obtener la serie de Fourier a partir de un PDF/imagen.",
                    "\\subsection*{Pasos}",
                    "\\begin{enumerate}",
                    "  \\item Convertir a escala de grises y binarizar (THRESH_BINARY_INV).",
                    "  \\item Extraer contornos con findContours y seleccionar el mayor.",
                    "  \\item Representar el contorno como señal compleja $z[n]=x_n + i y_n$.",
                    "  \\item Calcular FFT $Z[k]=\frac{1}{N}\\sum_{n=0}^{N-1} z[n] e^{-2\\pi i kn/N}$.",
                    "  \\item Reconstruir con la serie inversa: $z[n]\\approx \\sum_{k=-K}^{K} Z[k] e^{2\\pi i kn/N}$.",
                    "\\end{enumerate}",
                    "\\subsection*{Componentes detectados (ejemplo)}",
                ]

                for item in extracted[:20]:
                    sistema = item.get("predicted_system") or "(sin sistema)"
                    parts.append(f"\\paragraph{{{item.get('code','')}}} Sistema: {sistema} \\nContexto: {item.get('context','')[:200]}")

                with open(out_path, "w", encoding="utf-8") as f:
                    f.write("\n\n".join(parts))

                try:
                    os.startfile(out_path)
                except Exception:
                    pass

                messagebox.showinfo("LaTeX", f"Archivo generado: {out_path}", parent=self.panel)

            def _mostrar_metricas():
                _, metrics_path, _ = _load_training_files()
                if not os.path.exists(metrics_path):
                    messagebox.showinfo("Métricas", f"No se encontró el archivo de métricas:\n{metrics_path}", parent=self.panel)
                    return
                try:
                    with open(metrics_path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                except Exception as e:
                    messagebox.showerror("Error", str(e), parent=self.panel)
                    return
                train_acc = data.get("train_accuracy")
                val_acc = data.get("val_accuracy")
                msg = []
                if train_acc is not None:
                    msg.append(f"Precisión entrenamiento: {train_acc:.1%}")
                if val_acc is not None:
                    msg.append(f"Precisión validación: {val_acc:.1%}")
                errors = data.get("sample_errors", [])
                if errors:
                    msg.append("\nErrores de muestra:")
                    for e in errors[:5]:
                        txt = e.get("text", "")
                        msg.append(f"- {e.get('true')} → {e.get('pred')} : {txt[:80].replace('\n',' ')}...")
                messagebox.showinfo("Métricas del modelo", "\n".join(msg), parent=self.panel)

            btns_info = tk.Frame(content)
            btns_info.pack(fill=tk.X, padx=10, pady=(5, 0))
            _icon_btn(btns_info, "🔍", "Ver palabras clave (entrenamiento)", _mostrar_palabras_clave).pack(side=tk.LEFT, padx=5)
            _icon_btn(btns_info, "📊", "Ver métricas de entrenamiento", _mostrar_metricas).pack(side=tk.LEFT, padx=5)
            _icon_btn(btns_info, "🧾", "Mostrar síntesis (JSON)", _mostrar_sintesis).pack(side=tk.LEFT, padx=5)
            _icon_btn(btns_info, "📄", "Generar archivo LaTeX", _generar_latex).pack(side=tk.LEFT, padx=5)

            def _get_machine_id():
                maq_name = maq_var.get().strip()
                if not maq_name:
                    messagebox.showwarning("Aviso", "Seleccione una máquina destino", parent=self.panel)
                    return None
                for m in self.maquinas:
                    if f"{m.nombre} (ID {m.id})" == maq_name:
                        return m.id
                messagebox.showwarning("Aviso", "Máquina no válida", parent=self.panel)
                return None

            def _agregar_componentes(indices):
                maq_id = _get_machine_id()
                if maq_id is None:
                    return 0
                added = 0
                for idx in indices:
                    item = extracted[idx]
                    # Crear/obtener material asociado según el sistema sugerido
                    material_id = None
                    sistema = item.get("predicted_system")
                    if sistema:
                        material_id = _get_or_create_material(item.get("code") or "", sistema=sistema)
                    try:
                        agregar_componente(
                            maquina_id=maq_id,
                            nombre=item.get("code") or "Componente",
                            descripcion=item.get("context") or "",
                            material_id=material_id,
                        )
                        added += 1
                    except Exception:
                        pass
                if added > 0:
                    self._refresh_componentes()
                return added

            def _aplicar_a_maquina():
                sel = lb.curselection()
                if not sel:
                    messagebox.showwarning("Aviso", "Seleccione al menos un componente de la lista", parent=self.panel)
                    return
                added = _agregar_componentes(sel)
                if added:
                    messagebox.showinfo("Éxito", f"Se agregaron {added} componentes a la máquina seleccionada", parent=self.panel)

            def _agregar_todos_a_maquina():
                if not extracted:
                    messagebox.showwarning("Aviso", "No hay componentes para agregar", parent=self.panel)
                    return
                added = _agregar_componentes(range(len(extracted)))
                if added:
                    messagebox.showinfo("Éxito", f"Se agregaron {added} componentes a la máquina seleccionada", parent=self.panel)

            btns_apply = tk.Frame(content)
            btns_apply.pack(fill=tk.X, padx=10, pady=10)
            _icon_btn(btns_apply, "✅", "Aplicar seleccionados", _aplicar_a_maquina).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=2)
            _icon_btn(btns_apply, "🧩", "Agregar todos (sugerido)", _agregar_todos_a_maquina).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=2)

        self.view_mgr.push("Entrenamiento IA", _build, scrollable=True)


    def _on_comp_click(self, event):
        # determine which cell (row + column) was clicked
        tree = getattr(event, "widget", None) or getattr(self, "comp_tree", None)
        if not tree:
            return
        col = tree.identify_column(event.x)
        row_id = tree.identify_row(getattr(event, 'y', 0))
        if not col or col == "#0":
            return
        try:
            idx = int(col.lstrip("#")) - 1
        except ValueError:
            return
        cols = getattr(self, '_comp_cols', [])
        # idx 0 is the "#" counter column — skip it
        if idx <= 0 or idx >= len(cols):
            return
        col_key = cols[idx]
        maq = next((m for m in self.maquinas if f"m{m.id}" == col_key), None)
        if not maq:
            return
        self.comp_selected_machine_id = maq.id
        name = maq.nombre or f"ID {maq.id}"
        self.comp_selection_label.config(text=name, bg="lightblue")
        # highlight only the single clicked cell
        for iid in tree.get_children():
            vals = list(tree.item(iid, "values"))
            # strip previous markers from all machine columns
            for j in range(1, len(cols)):
                v = str(vals[j]) if j < len(vals) else ""
                if v.startswith("► ") and v.endswith(" ◄"):
                    vals[j] = v[2:-2]
            # add marker only to the exact clicked cell (same row AND column)
            if iid == row_id:
                ci = idx
                if ci < len(vals) and vals[ci]:
                    vals[ci] = f"► {vals[ci]} ◄"
                tree.item(iid, values=vals)
            else:
                tree.item(iid, values=vals)

    def _on_select_componente(self, event, sistema):
        """Actualiza etiqueta de selección al elegir un componente en la pestaña."""
        tree = getattr(event, "widget", None)
        if not tree:
            return
        sel = tree.selection()
        label = self.comp_selection_labels.get(sistema) or getattr(self, 'comp_selection_label', None)
        if not label:
            return
        if not sel:
            label.config(text="", bg=theme.current.get("label_bg", "white"))
            return
        vals = tree.item(sel[0]).get('values') or []
        nombre = vals[0] if len(vals) > 0 else ""
        codigo = vals[1] if len(vals) > 1 else ""
        texto = f"Seleccionado: {nombre}"
        if codigo:
            texto += f" (Código: {codigo})"
        label.config(text=texto, bg=theme.current.get("select_bg", "#cde"))

    def _on_comp_double_click(self, event):
        """Abre un diálogo de edición al hacer doble click sobre un componente."""
        tree = getattr(event, "widget", None) or getattr(self, "comp_tree", None)
        if not tree:
            return

        # En nuestra vista moderna, el Treeview usa el ID del componente como iid.
        if tree is getattr(self, "comp_tree", None):
            sel = tree.selection()
            if not sel:
                return
            try:
                comp_id = int(sel[0])
            except Exception:
                return
            self._dialogo_editar_componente(comp_id)
            return

        # Fallback: comportamiento antiguo para matrices (legacy)
        # (mantenerlo para compatibilidad con viejas vistas si se usan)
        col = tree.identify_column(event.x)
        row_id = tree.identify_row(event.y)
        if not col or col == "#0" or not row_id:
            return
        try:
            col_idx = int(col.lstrip("#")) - 1
        except ValueError:
            return
        cols = getattr(self, '_comp_cols', [])
        if col_idx <= 0 or col_idx >= len(cols):
            return
        col_key = cols[col_idx]
        # Determine the row index within this column
        all_children = tree.get_children()
        row_index = list(all_children).index(row_id) if row_id in all_children else -1
        if row_index < 0:
            return
        # Check cell has content
        cell_val = tree.set(row_id, col_key)
        if not cell_val or cell_val.strip() == "":
            # Celda vacía: preguntar si desea añadir componente
            maq = next((m for m in self.maquinas if f"m{m.id}" == col_key), None)
            if maq and messagebox.askyesno("Componente", "¿Desea añadir Componente?", parent=self.panel):
                self._dialogo_agregar_componente(maquina=maq)
            return
        # Lookup component id
        comp_id_map = getattr(self, '_comp_id_map', {})
        comp_id = comp_id_map.get((col_key, row_index))
        if comp_id is None:
            return
        self._dialogo_editar_componente(comp_id)

    def _eliminar_componente_seleccionado(self):
        """Elimina el/los componente(s) seleccionado(s) en la tabla."""
        from .services.ingenieria_extras import eliminar_componente, agregar_componente

        # Recopilar IDs seleccionados en las tablas de componentes (por sistema)
        comp_ids = []
        for tree in getattr(self, 'comp_trees', {}).values():
            for sel in tree.selection():
                try:
                    comp_ids.append(int(sel))
                except Exception:
                    continue
        # Fallback a la vista de matriz antigua (marcadores ► ◄)
        if not comp_ids and hasattr(self, 'comp_tree'):
            for iid in self.comp_tree.get_children():
                vals = list(self.comp_tree.item(iid, "values"))
                cols = getattr(self, '_comp_cols', [])
                row_index = list(self.comp_tree.get_children()).index(iid)
                for j in range(1, len(cols)):
                    v = str(vals[j]) if j < len(vals) else ""
                    if v.startswith("► ") and v.endswith(" ◄"):
                        comp_id = self._comp_id_map.get((cols[j], row_index))
                        if comp_id:
                            comp_ids.append(comp_id)
                if comp_ids:
                    break

        comp_ids = list(dict.fromkeys(comp_ids))  # unicos, en orden

        if not comp_ids:
            messagebox.showwarning("Aviso", "Seleccione al menos un componente para eliminar.", parent=self.panel)
            return

        count = len(comp_ids)
        msg = f"¿Está seguro de eliminar {'estos' if count > 1 else 'este'} {count} componente{'s' if count > 1 else ''}?"
        if not messagebox.askyesno("Confirmar", msg, parent=self.panel):
            return

        try:
            all_comps = listar_componentes(None)
            snaps = []
            for comp_id in comp_ids:
                comp_obj = next((c for c in all_comps if c.id == comp_id), None)
                if comp_obj:
                    snaps.append({
                        'maquina_id': comp_obj.maquina_id,
                        'nombre': comp_obj.nombre,
                        'descripcion': comp_obj.descripcion or '',
                        'material_id': comp_obj.material_id,
                    })
                try:
                    eliminar_componente(comp_id)
                except Exception:
                    pass

            if snaps:
                def _undo_delete(snaps=snaps):
                    for s in snaps:
                        agregar_componente(s['maquina_id'], s['nombre'], s['descripcion'], material_id=s['material_id'])
                    self._refresh_componentes_listado()

                self.undo.push(
                    f"Eliminar {len(snaps)} componente{'s' if len(snaps) != 1 else ''}",
                    _undo_delete,
                )

            messagebox.showinfo("Eliminado", f"Se eliminaron {count} componente{'s' if count > 1 else ''}.", parent=self.panel)
            self._refresh_componentes_listado()
        except Exception as e:
            messagebox.showerror("Error", str(e), parent=self.panel)

    def _dialogo_editar_componente(self, comp_id):
        """Abre un diálogo para editar un componente existente."""
        from .services.ingenieria_extras import listar_componentes, agregar_componente, actualizar_componente, listar_materiales, eliminar_componente

            # Find the component
        all_comps = listar_componentes(None)
        comp = next((c for c in all_comps if c.id == comp_id), None)
        if not comp:
            messagebox.showerror("Error", "Componente no encontrado")
            return

        maq = next((m for m in self.maquinas if m.id == comp.maquina_id), None)

        def _build(content):
            tk.Label(content, text=f"Máquina: {maq.nombre if maq else comp.maquina_id}", font=("Arial", 10, "bold")).pack(pady=5)

            tk.Label(content, text="Nombre").pack(anchor="w", padx=10)
            name_entry = tk.Entry(content, width=40)
            name_entry.pack(fill=tk.X, padx=10, pady=2)
            name_entry.insert(0, comp.nombre or "")

            tk.Label(content, text="Descripción (opcional)").pack(anchor="w", padx=10)
            desc_entry = tk.Entry(content, width=40)
            desc_entry.pack(fill=tk.X, padx=10, pady=2)
            desc_entry.insert(0, comp.descripcion or "")

            tk.Label(content, text="Material asociado (opcional)").pack(anchor="w", padx=10)
            mats = listar_materiales()
            mat_names = [m.nombre for m in mats]
            comp_mat_var = tk.StringVar()
            if getattr(comp, 'material', None):
                comp_mat_var.set(comp.material.nombre)
            comp_mat_combo = ttk.Combobox(content, values=mat_names, textvariable=comp_mat_var, width=38)
            def _val_mat(event=None):
                v = comp_mat_var.get().strip()
                if v and v not in mat_names:
                    comp_mat_var.set("")
            comp_mat_combo.bind("<FocusOut>", _val_mat)
            comp_mat_combo.pack(fill=tk.X, padx=10, pady=2)


            def guardar_edit():
                nombre = name_entry.get().strip()
                if not nombre:
                    messagebox.showwarning("Aviso", "El nombre es obligatorio")
                    return
                desc = desc_entry.get().strip() or ""

                mat_map = {m.nombre: m for m in mats}

                mat_id = None
                mat_name = comp_mat_var.get().strip()
                if mat_name:
                    found = mat_map.get(mat_name)
                    if found:
                        mat_id = found.id

                # Snapshot antes de actualizar
                old_vals = {
                    'nombre': comp.nombre,
                    'descripcion': comp.descripcion or '',
                    'material_id': comp.material_id,
                    'sistema': getattr(comp, 'sistema', None),
                }

                old_nombre = comp.nombre
                try:

                    data = {
                        "nombre": nombre,
                        "descripcion": desc,
                        "material_id": mat_id
                        }

                    actualizar_componente(comp_id, nombre=nombre, descripcion=desc, material_id=mat_id)
                except Exception as e:
                    messagebox.showerror("Error", str(e) )
                    return
                
                
                self.undo.push(f"Editar componente '{old.nombre}'",
                    lambda cid=comp_id, ov=old_vals: (
                        actualizar_componente(cid, **ov), 
                        self._refresh_componentes_listado())
                        )
                
                messagebox.showinfo("Actualizado", "Componente actualizado con éxito")

                # Force-refresh the component list to reflect the change.
                # Some UI paths may not redraw immediately, so also update the
                # specific row directly.
                self._refresh_componentes_listado()
                try:
                    from .services.ingenieria_extras import listar_componentes
                    updated = next((c for c in listar_componentes(None) if c.id == comp_id), None)
                    if updated and hasattr(self, 'comp_tree'):
                        maq = next((m for m in self.maquinas if m.id == updated.maquina_id), None)
                        maquina_nombre = maq.nombre if maq else f"ID {updated.maquina_id}"
                        mat = getattr(updated, 'material', None)
                        codigo = getattr(mat, 'codigo', '') if mat else ''
                        tipo = getattr(mat, 'tipo', '') if mat else ''
                        sistema = getattr(updated, 'sistema', '') or ''
                        self.comp_tree.item(
                            str(comp_id),
                            values=(
                                maquina_nombre,
                                updated.nombre or '',
                                codigo,
                                tipo,
                                sistema,
                                updated.descripcion or '',
                            ),
                        )
                except Exception:
                    pass

                self.view_mgr.pop()

            _icon_btn(content, "💾", "Guardar cambios", guardar_edit, font=("Arial", 10, "bold")).pack(fill=tk.X, padx=10, pady=8)

            def _desasociar():
                if not messagebox.askyesno("Desasociar", "¿Desea eliminar este componente (desasociar Material - Máquina)?", parent=self.panel):
                    return
                try:
                    eliminar_componente(comp_id)
                    messagebox.showinfo("Desasociado", "Componente eliminado.", parent=self.panel)
                    self._refresh_componentes()
                    self.view_mgr.pop()
                except Exception as e:
                    messagebox.showerror("Error", str(e), parent=self.panel)

            _icon_btn(content, "🧩", "Desasociar Material - Máquina", _desasociar, font=("Arial", 10, "bold")).pack(fill=tk.X, padx=10, pady=(0, 8))

        self.view_mgr.push("Editar componente", _build)

    def _crea_tab_busqueda(self, notebook):
        tab = ttk.Frame(notebook)
        notebook.add(tab, text="Buscar en Toro")

        tk.Label(tab, text="Código o nombre:").pack(padx=5, pady=5, anchor="w")
        self.search_entry = tk.Entry(tab)
        self.search_entry.pack(fill=tk.X, padx=5)
        _icon_btn(tab, "🔍", "Buscar en Toro", self._buscar_toro).pack(pady=5)
        self.result_text = tk.Text(tab, height=10)
        self.result_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

    def _crea_tab_registro(self, notebook):
        tab = ttk.Frame(notebook)
        notebook.add(tab, text="Registro Diario")

        top = tk.Frame(tab)
        top.pack(fill=tk.X, padx=10, pady=5)
        tk.Label(top, text="Fecha:").pack(side=tk.LEFT, padx=5)

        self.reg_fecha_var = tk.StringVar(value=datetime.now().date().isoformat())
        try:
            from tkcalendar import DateEntry
            self.reg_fecha_entry = DateEntry(top, textvariable=self.reg_fecha_var, date_pattern="yyyy-mm-dd", width=12)
        except Exception:
            # fallback si no está instalado; usar entrada de texto
            self.reg_fecha_entry = tk.Entry(top, textvariable=self.reg_fecha_var, width=14)

        self.reg_fecha_entry.pack(side=tk.LEFT)

        _icon_btn(top, "📅", "Todos", lambda: (self.reg_fecha_var.set('Todos'), self._cargar_registro())).pack(side=tk.LEFT, padx=3)

        tk.Label(top, text="Máquina:").pack(side=tk.LEFT, padx=(12, 5))
        self.reg_maquina_var = tk.StringVar(value="Todos")
        self.reg_maquina_combo = ttk.Combobox(top, textvariable=self.reg_maquina_var, values=["Todos"], state="readonly", width=26)
        self.reg_maquina_combo.pack(side=tk.LEFT)

        _icon_btn(top, "🔄", "Cargar registro", self._cargar_registro).pack(side=tk.LEFT, padx=5)
        _icon_btn(top, "📁", "Exportar CSV", self._exportar_registro).pack(side=tk.LEFT, padx=5)

        cols = ("fecha","maq","h_ini","h_fin","h_inicial","h_final","horas","oper","comb","fallo")
        self.reg_tree = ttk.Treeview(tab, columns=cols, show="headings")
        # more descriptive headings: first column is report date, then horómetro, times, and report metadata
        headings = {
            "fecha": "Fecha",
            "maq": "Máquina",
            "h_ini": "Horómetro Inicial",
            "h_fin": "Horómetro Final",
            "h_inicial": "Inicio",
            "h_final": "Fin",
            "horas": "Hora informe",
            "oper": "Operario",
            "comb": "Combustible",
            "fallo": "Falla"
        }
        for c in cols:
            self.reg_tree.heading(c, text=headings.get(c, c))
            self.reg_tree.column(c, width=80)
        self.reg_tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        aplicar_ordenamiento(self.reg_tree)

    def _crea_tab_datos_red(self, notebook):
        tab = ttk.Frame(notebook)
        notebook.add(tab, text="Datos Red")

        top = tk.Frame(tab)
        top.pack(fill=tk.X, padx=10, pady=5)

        _icon_btn(top, "📶", "Modo Celulares", lambda: self._set_datos_red_estado('celulares')).pack(side=tk.LEFT, padx=5)
        _icon_btn(top, "🗺️", "Modo Datos Espaciales", lambda: self._set_datos_red_estado('espacial')).pack(side=tk.LEFT, padx=5)
        _icon_btn(top, "�", "Actualizar datos", self._cargar_datos_red).pack(side=tk.LEFT, padx=5)
        _icon_btn(top, "📊", "Exportar CSV", self._exportar_datos_red_csv).pack(side=tk.LEFT, padx=5)
        _icon_btn(top, "🧠", "Análisis Avanzado", self._abrir_analisis_avanzado).pack(side=tk.LEFT, padx=5)

        self.red_estado_label = tk.Label(top, text='Estado: Celulares', fg='blue')
        self.red_estado_label.pack(side=tk.RIGHT, padx=5)

        # Frame para los distintos estados
        self.red_celulares_frame = tk.Frame(tab)
        self.red_espacial_frame = tk.Frame(tab)

        # Celulares Treeview
        cols = ("id", "nombre", "rubro", "ip", "ultima_actualizacion", "estado")
        self.celulares_tree = ttk.Treeview(self.red_celulares_frame, columns=cols, show="headings", height=10)
        for c in cols:
            self.celulares_tree.heading(c, text=c.replace("_", " ").capitalize())
            self.celulares_tree.column(c, width=140, anchor="w")
        hsb = ttk.Scrollbar(self.red_celulares_frame, orient=tk.HORIZONTAL, command=self.celulares_tree.xview)
        vsb = ttk.Scrollbar(self.red_celulares_frame, orient=tk.VERTICAL, command=self.celulares_tree.yview)
        self.celulares_tree.configure(xscrollcommand=hsb.set, yscrollcommand=vsb.set)
        self.celulares_tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10, side=tk.LEFT)
        vsb.pack(fill=tk.Y, side=tk.RIGHT)
        hsb.pack(fill=tk.X, side=tk.BOTTOM)

        # Tooltip dinámico para nombres de celulares al pasar el mouse
        self._celulares_tip = None
        self._celulares_tip_item = None

        def _show_celulares_tooltip(event):
            item = self.celulares_tree.identify_row(event.y)
            if not item:
                _hide_tooltip(event)
                return

            nombre = self.celulares_tree.item(item, 'values')[1] if len(self.celulares_tree.item(item, 'values')) > 1 else ''
            if not nombre:
                _hide_tooltip(event)
                return

            if self._celulares_tip_item == item and self._celulares_tip:
                # Si ya está mostrando el tooltip para este item, sólo moverlo
                try:
                    self._celulares_tip.wm_geometry(f'+{event.x_root + 10}+{event.y_root + 10}')
                except Exception:
                    pass
                return

            _hide_tooltip(event)
            self._celulares_tip_item = item
            self._celulares_tip = tk.Toplevel(self.celulares_tree)
            self._celulares_tip.wm_overrideredirect(True)
            self._celulares_tip.attributes('-topmost', True)
            try:
                self._celulares_tip.transient(self.celulares_tree)
            except Exception:
                pass
            self._celulares_tip.wm_geometry(f'+{event.x_root + 10}+{event.y_root + 10}')
            tk.Label(
                self._celulares_tip,
                text=f"Nombre: {nombre}",
                background=theme.current.get('menu_bg', '#ffffe0'),
                foreground=theme.current.get('menu_fg', '#000000'),
                relief='solid',
                borderwidth=1,
                font=('Arial', 9),
                padx=4,
                pady=2,
            ).pack()

        def _hide_tooltip(event=None):
            if self._celulares_tip:
                try:
                    self._celulares_tip.destroy()
                except Exception:
                    pass
                self._celulares_tip = None
            self._celulares_tip_item = None

        self.celulares_tree.bind('<Motion>', _show_celulares_tooltip)
        self.celulares_tree.bind('<Leave>', _hide_tooltip)

        # Espacial texto
        self.espacial_text = tk.Text(self.red_espacial_frame, height=15)
        self.espacial_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Pack frames
        self.red_celulares_frame.pack(fill=tk.BOTH, expand=True)
        self.red_espacial_frame.pack(fill=tk.BOTH, expand=True)

        self.red_estado = 'celulares'
        self._set_datos_red_estado(self.red_estado)

        # Cargar contenido inicial
        self._cargar_celulares()
        self._cargar_datos_espaciales()

    def _abrir_analisis_avanzado(self):
        """Lanza el Sistema de Análisis Avanzado (PyQt6)."""
        import subprocess, sys
        raiz = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
        script = os.path.join(raiz, "run_analisis.py")
        if not os.path.isfile(script):
            messagebox.showerror('Análisis', f'No se encontró: {script}', parent=self.panel)
            return
        subprocess.Popen([sys.executable, script], cwd=raiz)

    def _ia_actualizar_estado(self):
        if hasattr(self, 'ia_status_label'):
            self.ia_status_label.config(text=self._get_ia_status_text())

    def _ia_mostrar_estado_resumen(self):
        """Muestra resumen de estado IA en modo solo lectura."""
        try:
            from .ai import config as _config
            data = _config.get_ai_status()
            estado = (
                "═" * 50 + "\n"
                "  ESTADO DE INTELIGENCIA ARTIFICIAL\n"
                "═" * 50 + "\n\n"
                f"Backend activo:           {data.active_backend}\n"
                f"Modelo GPT:               {data.gpt_model}\n"
                f"OpenAI disponible:        {'Sí' if data.openai_available else 'No'}\n"
                f"DeepSeek disponible:      {'Sí' if data.deepseek_available else 'No'}\n"
                f"Transformers:             {'Sí' if data.transformers_available else 'No'}\n"
                f"Sentence Transformers:    {'Sí' if data.sentence_transformers_available else 'No'}\n"
                f"AaaS URL:                 {data.aaas_url}\n"
                f"Modo offline:             {'Sí' if data.offline_mode else 'No'}\n"
                f"Poppler:                  {data.poppler_path or 'No encontrado'}\n\n"
                "Para consultas avanzadas y entrenamiento de IA,\n"
                "use el botón «Abrir Análisis Avanzado» (arriba).\n"
            )
            self.ia_result_text.config(state='normal')
            self.ia_result_text.delete('1.0', tk.END)
            self.ia_result_text.insert(tk.END, estado)
            self.ia_result_text.config(state='disabled')
        except Exception as e:
            self.ia_result_text.config(state='normal')
            self.ia_result_text.delete('1.0', tk.END)
            self.ia_result_text.insert(tk.END, f"Error obteniendo estado: {e}")
            self.ia_result_text.config(state='disabled')

    def _ia_actualizar_estado(self):
        if hasattr(self, 'ia_status_label'):
            self.ia_status_label.config(text=self._get_ia_status_text())

    # ── Métodos IA legacy (redirigen a Análisis Avanzado) ──

    def _ia_recargar_indice(self):
        self._abrir_analisis_avanzado()

    def _ia_indexar_historial(self):
        self._abrir_analisis_avanzado()

    def _ia_buscar(self):
        self._abrir_analisis_avanzado()

    def _ia_ask_aaas(self, query):
        return []

    def _ia_ask_gpt(self):
        self._abrir_analisis_avanzado()

    def _ia_ask_aaas_ui(self):
        self._abrir_analisis_avanzado()

    def _ia_generate_lisp_from_pdf(self):
        self._abrir_analisis_avanzado()

    def _ia_mostrar_estado(self):
        try:
            self._ia_mostrar_estado_resumen()
        except Exception as e:
            messagebox.showerror('Estado IA', f'No se pudo obtener estado: {e}', parent=self.panel)

    def _ia_web_search(self, query):
        try:
            import requests
            resp = requests.get(
                'https://api.duckduckgo.com/',
                params={'q': query, 'format': 'json', 'no_html': 1, 'skip_disambig': 1},
                timeout=6,
            )
            if resp.ok:
                data = resp.json() or {}
                abstract = data.get('Abstract', '')
                related = data.get('RelatedTopics', [])
                filas = []
                if abstract:
                    filas.append(f"DuckDuckGo Abstract: {abstract}")
                for item in (related or [])[:5]:
                    text = ''
                    if isinstance(item, dict):
                        text = item.get('Text', '')
                    if text:
                        filas.append(f"DuckDuckGo: {text}")
                return filas
        except Exception:
            pass
        return []

    def _ia_feedback(self, payload):
        try:
            import requests
            requests.post('http://127.0.0.1:8000/assistant/feedback', json=payload, timeout=5)
        except Exception:
            pass

    def _set_datos_red_estado(self, estado):
        self.red_estado = estado
        if hasattr(self, 'red_estado_label'):
            self.red_estado_label.config(text=f"Estado: {estado.capitalize()}")

        if estado == 'celulares':
            self.red_celulares_frame.lift()
            self.red_celulares_frame.tkraise()
            self.red_celulares_frame.pack(fill=tk.BOTH, expand=True)
            self.red_espacial_frame.pack_forget()
            self._cargar_celulares()
        elif estado == 'espacial':
            self.red_espacial_frame.lift()
            self.red_espacial_frame.tkraise()
            self.red_espacial_frame.pack(fill=tk.BOTH, expand=True)
            self.red_celulares_frame.pack_forget()
            self._cargar_datos_espaciales()
        else:
            # fallback
            self.red_celulares_frame.pack(fill=tk.BOTH, expand=True)
            self.red_espacial_frame.pack_forget()

    def _siguiente_estado_datos_red(self):
        next_estado = 'espacial' if self.red_estado == 'celulares' else 'celulares'
        self._set_datos_red_estado(next_estado)

    def _cargar_datos_red(self):
        # Refresca ambas secciones, pero mantiene el estado actual
        self._cargar_celulares()
        self._cargar_datos_espaciales()

        # Actualiza la vista visible según el estado
        self._set_datos_red_estado(self.red_estado)

    def _exportar_datos_red_csv(self):
        """Exporta los datos de red (celulares) a CSV."""
        try:
            from tkinter import filedialog
            import csv

            ruta = filedialog.asksaveasfilename(
                title='Guardar CSV de Datos Red',
                defaultextension='.csv',
                filetypes=[('CSV', '*.csv')],
                parent=self.panel,
            )
            if not ruta:
                return

            # Extraer datos del treeview
            filas = []
            for child in self.celulares_tree.get_children():
                vals = self.celulares_tree.item(child, 'values')
                filas.append(vals)

            with open(ruta, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["ID", "Nombre", "Rubro", "IP", "Última Actualización", "Estado"])
                for fila in filas:
                    writer.writerow(fila)

            messagebox.showinfo('Exportar', f'Exportados {len(filas)} dispositivos a:\n{ruta}', parent=self.panel)
        except Exception as e:
            messagebox.showerror('Exportar', f'Error exportando: {e}', parent=self.panel)

    def _cargar_registro(self):
        from .services.ingenieria_extras import listar_operaciones_por_fecha, listar_operaciones_todas

        fecha_text = self.reg_fecha_var.get().strip() if hasattr(self, 'reg_fecha_var') else ''
        try:
            if not fecha_text or fecha_text.lower() == 'todos':
                ops = listar_operaciones_todas()
            else:
                fecha = datetime.fromisoformat(fecha_text).date()
                ops = listar_operaciones_por_fecha(fecha)
        except Exception:
            messagebox.showwarning('Aviso', 'Fecha inválida. Use YYYY-MM-DD o Todos.')
            return

        maq_sel = self.reg_maquina_var.get().strip() if hasattr(self, 'reg_maquina_var') else 'Todos'
        maq_id = None
        if maq_sel and maq_sel != 'Todos':
            try:
                maq_id = int(maq_sel.split(maxsplit=1)[0])
            except Exception:
                maq_id = None

        self.reg_tree.delete(*self.reg_tree.get_children())
        now = datetime.now()
        now_str = now.strftime('%Y-%m-%d %H:%M:%S')
        for o in ops:
            if maq_id is not None and o.maquina_id != maq_id:
                continue
            self.reg_tree.insert('', tk.END, iid=str(o.id), values=(
                o.fecha.isoformat() if getattr(o, 'fecha', None) else '',
                o.maquina_id,
                o.horometro_inicial,
                o.horometro_final,
                o.hora_inicio.strftime('%Y-%m-%d %H:%M'),
                o.hora_fin.strftime('%Y-%m-%d %H:%M'),
                now_str,
                o.operador,
                getattr(o, 'combustible', ''),
                'sí' if getattr(o, 'hubo_falla', 0) else ''
            ))

        self.reg_tree.bind('<Double-1>', self._on_registro_double_click)

        # Actualiza lista de máquinas en combo (puede probarse al iniciar o recargar)
        if hasattr(self, 'reg_maquina_combo') and hasattr(self, 'maquinas'):
            vals = ['Todos'] + [f'{m.id} - {m.nombre}' for m in self.maquinas if m]
            self.reg_maquina_combo['values'] = vals
            if self.reg_maquina_var.get() not in vals:
                self.reg_maquina_var.set('Todos')

    def _on_registro_double_click(self, event=None):
        tree = getattr(self, 'reg_tree', None)
        if not tree:
            return

        selected = tree.selection()
        if not selected:
            selected_row = tree.identify_row(event.y) if event else None
            selected = (selected_row,) if selected_row else ()
        if not selected:
            return

        row_id = selected[0]
        try:
            operacion_id = int(row_id)
        except Exception:
            return

        from .services.ingenieria_extras import obtener_operacion_diaria, actualizar_operacion_diaria
        operacion = obtener_operacion_diaria(operacion_id)
        if not operacion:
            messagebox.showwarning('Aviso', 'Operación diaria no encontrada')
            return

        def _format_dt(value):
            return value.strftime('%Y-%m-%d %H:%M') if value else ''

        def _build(content):
            tk.Label(content, text='Editar Operación diaria', font=('Arial', 10, 'bold')).pack(padx=10, pady=5)

            tk.Label(content, text='Fecha (YYYY-MM-DD)').pack(anchor='w', padx=10)
            fecha_var = tk.StringVar(value=operacion.fecha.isoformat() if operacion.fecha else '')
            tk.Entry(content, textvariable=fecha_var).pack(fill=tk.X, padx=10, pady=2)

            tk.Label(content, text='Hora inicio (YYYY-MM-DD HH:MM)').pack(anchor='w', padx=10)
            h_ini_var = tk.StringVar(value=_format_dt(operacion.hora_inicio))
            tk.Entry(content, textvariable=h_ini_var).pack(fill=tk.X, padx=10, pady=2)

            tk.Label(content, text='Hora fin (YYYY-MM-DD HH:MM)').pack(anchor='w', padx=10)
            h_fin_var = tk.StringVar(value=_format_dt(operacion.hora_fin))
            tk.Entry(content, textvariable=h_fin_var).pack(fill=tk.X, padx=10, pady=2)

            tk.Label(content, text='Horómetro inicial').pack(anchor='w', padx=10)
            horo_ini_var = tk.StringVar(value=str(operacion.horometro_inicial or ''))
            tk.Entry(content, textvariable=horo_ini_var).pack(fill=tk.X, padx=10, pady=2)

            tk.Label(content, text='Horómetro final').pack(anchor='w', padx=10)
            horo_fin_var = tk.StringVar(value=str(operacion.horometro_final or ''))
            tk.Entry(content, textvariable=horo_fin_var).pack(fill=tk.X, padx=10, pady=2)

            tk.Label(content, text='Operario').pack(anchor='w', padx=10)
            oper_var = tk.StringVar(value=operacion.operador or '')
            tk.Entry(content, textvariable=oper_var).pack(fill=tk.X, padx=10, pady=2)

            tk.Label(content, text='Combustible').pack(anchor='w', padx=10)
            comb_var = tk.StringVar(value=str(getattr(operacion, 'combustible', '') or ''))
            tk.Entry(content, textvariable=comb_var).pack(fill=tk.X, padx=10, pady=2)

            hubo_var = tk.IntVar(value=1 if getattr(operacion, 'hubo_falla', 0) else 0)
            tk.Checkbutton(content, text='Hubo falla', variable=hubo_var).pack(anchor='w', padx=10, pady=4)

            def guardar():
                try:
                    fecha_val = datetime.fromisoformat(fecha_var.get().strip()).date()
                    h_ini_val = datetime.fromisoformat(h_ini_var.get().strip())
                    h_fin_val = datetime.fromisoformat(h_fin_var.get().strip())
                    h_ini_horo_val = float(horo_ini_var.get().strip())
                    h_fin_horo_val = float(horo_fin_var.get().strip())
                    combust_val = float(comb_var.get().strip()) if comb_var.get().strip() else None
                except Exception as e:
                    messagebox.showerror('Error', f'Datos inválidos: {e}')
                    return

                if h_fin_val <= h_ini_val:
                    messagebox.showerror('Error', 'La hora fin debe ser mayor a la hora inicio')
                    return
                if h_fin_horo_val < h_ini_horo_val:
                    messagebox.showerror('Error', 'Horómetro final no puede ser menor que el inicial')
                    return

                try:
                    actualizar_operacion_diaria(
                        operacion_id,
                        fecha=fecha_val,
                        hora_inicio=h_ini_val,
                        hora_fin=h_fin_val,
                        horometro_inicial=h_ini_horo_val,
                        horometro_final=h_fin_horo_val,
                        operador=oper_var.get().strip(),
                        combustible=combust_val,
                        hubo_falla=1 if hubo_var.get() else 0,
                    )
                    messagebox.showinfo('Guardado', 'Operación diaria actualizada')
                    self._cargar_registro()
                    self.view_mgr.pop()
                except Exception as ex:
                    messagebox.showerror('Error', str(ex))

            _icon_btn(content, '💾', 'Guardar', guardar).pack(fill=tk.X, padx=10, pady=10)

        self.view_mgr.push('Editar Operación', _build)

    def _cargar_datos_espaciales(self):
        # consulta a la API para dispositivos y planos registrados
        import requests
        try:
            resp_dev = requests.get("http://127.0.0.1:8000/devices", timeout=1)
            dispositivos = resp_dev.json() if resp_dev.ok else []
        except Exception:
            dispositivos = []
        try:
            resp_pl = requests.get("http://127.0.0.1:8000/planos", timeout=1)
            planos_api = resp_pl.json() if resp_pl.ok else []
        except Exception:
            planos_api = []

        pd = generar_plano_general()
        resumen = pd.json_summary()

        # Normalizar nombre de dispositivos para texto
        dispositivos_labels = []
        for d in dispositivos:
            if isinstance(d, str):
                dispositivos_labels.append(d)
            elif isinstance(d, dict):
                dispositivos_labels.append(str(d.get('nombre') or d.get('name') or d.get('ip') or d.get('id', '')))
            else:
                dispositivos_labels.append(str(d))

        texto = (
            f"Dispositivos conectados: {', '.join(dispositivos_labels)}\n"
            f"Planos en servidor: {len(planos_api)}\n\n"
            f"Área total: {resumen['area']:.1f} m²\n"
            f"Máquinas: {resumen['maquinas']}\n"
            f"Densidad: {resumen['densidad']:.3f} /m²\n"
            f"Fingerprint: {resumen['fingerprint'][:5]}...\n\n"
        )
        texto += "Polylines:\n"
        for poly in pd.polylines:
            texto += f"- {poly.nombre} ({len(poly.puntos)} pts)\n"
        self.espacial_text.delete("1.0", tk.END)
        self.espacial_text.insert(tk.END, texto)

    def _generar_plano_celulares(self):
        """Genera un plano con los celulares conectados y lo persiste."""
        from datetime import datetime
        try:
            dispositivos = listar_dispositivos() or []
        except Exception:
            dispositivos = []

        if not dispositivos:
            messagebox.showinfo("Plano Celulares", "No hay dispositivos registrados para generar el plano.", parent=self.panel)
            return

        nombre = f"Plano Celulares {datetime.now().strftime('%Y%m%d_%H%M%S')}"
        descripcion = "Plano generado automáticamente desde dispositivos móviles conectados"

        try:
            plano_data = PlanoData.crear_vacio(nombre, descripcion)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo crear el plano: {e}", parent=self.panel)
            return

        # Generación de coordenadas sintéticas basadas en hash de IP/nombre
        import hashlib
        import math

        def _fake_coord(value, idx):
            if not value:
                value = f"unknown-{idx}"
            h = hashlib.sha256(f"{value}-{idx}".encode('utf-8')).hexdigest()
            x = (int(h[0:8], 16) % 1000) / 10.0
            y = (int(h[8:16], 16) % 1000) / 10.0
            return x, y

        markers = []
        poly_points = []

        for i, d in enumerate(dispositivos):
            nombre_disp = getattr(d, 'nombre', None) if not isinstance(d, dict) else d.get('nombre') or d.get('name') or f"disp_{i}"
            ip = getattr(d, 'ip', None) if not isinstance(d, dict) else d.get('ip')
            x, y = _fake_coord(ip or nombre_disp, i)
            marcador = MarcadorMaquina(maquina_id=0, x=x, y=y, etiqueta=str(nombre_disp))
            markers.append(marcador)
            poly_points.append({'x': x, 'y': y})

        # Agregar polylines y marcadores al plano
        plano_data.marcadores.extend(markers)
        if len(poly_points) > 1:
            plano_data.polylines.append(Polyline(puntos=poly_points, nombre="Recorrido celulares", color_hex="#FF0077FF", grosor=2.0, es_cerrada=False, tipo="linea"))

        try:
            plano_data.guardar()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar el plano generado: {e}", parent=self.panel)
            return

        messagebox.showinfo("Plano Celulares", f"Plano '{nombre}' generado y guardado exitosamente con {len(markers)} marcadores.", parent=self.panel)
        self._cargar_datos_espaciales()  # refresca la sección con resumen actualizado

    def _cargar_celulares(self):
        import requests
        from datetime import datetime, timedelta, timezone

        # 1) vez que vienen de API de dispositivos conectados (temp runtime)
        conectados = []
        try:
            resp = requests.get("http://127.0.0.1:8000/devices", timeout=1)
            if resp.ok:
                conectados = resp.json() or []
        except Exception:
            conectados = []

        # 2) dispositivos registrados en DB (persistentes)
        try:
            registrados = listar_dispositivos() or []
        except Exception:
            registrados = []

        # Normalizar lista de objetos completos, manteniendo estado conectado/inactivo
        ahora = datetime.now(timezone.utc)

        # Mapa de nombres de conectados si viene lista de strings o dicts
        conectados_por_nombre = {}
        for item in conectados:
            if isinstance(item, str):
                conectados_por_nombre[item] = {'nombre': item, 'source': 'API', 'connected': True}
            elif isinstance(item, dict):
                key = item.get('nombre') or item.get('name') or str(item.get('id', ''))
                conectados_por_nombre[key] = {**item, 'source': 'API', 'connected': True}

        # Agrega dispositivos de DB con marca de si está conectado ahora
        filas = []
        for d in registrados:
            nombre = getattr(d, 'nombre', '')
            connected = bool(nombre in conectados_por_nombre)

            ultima = getattr(d, 'ultima_actualizacion', None)
            if isinstance(ultima, str):
                try:
                    ultima_dt = datetime.fromisoformat(ultima)
                except Exception:
                    ultima_dt = None
            else:
                ultima_dt = ultima

            if ultima_dt and getattr(ultima_dt, 'tzinfo', None) is None:
                try:
                    ultima_dt = ultima_dt.replace(tzinfo=timezone.utc)
                except Exception:
                    pass

            if ultima_dt:
                delta = ahora - ultima_dt
                estado = 'Activo' if delta <= timedelta(minutes=10) else 'Inactivo'
            else:
                estado = 'Desconocido'

            filas.append({
                'id': getattr(d, 'id', ''),
                'nombre': nombre,
                'rubro': getattr(d, 'rubro', ''),
                'ip': getattr(d, 'ip', ''),
                'ultima_actualizacion': getattr(d, 'ultima_actualizacion', '') or '-',
                'estado': 'Conectado' if connected else estado,
                'registrado': True,
            })

        # Añadir los conectados que no están en DB como filas temporales
        for nombre, info in conectados_por_nombre.items():
            if any(f['nombre'] == nombre for f in filas):
                continue
            filas.append({
                'id': '',
                'nombre': nombre,
                'rubro': info.get('rubro', ''),
                'ip': info.get('ip', ''),
                'ultima_actualizacion': info.get('ultima_actualizacion', ''),
                'estado': 'Conectado',
                'registrado': False,
            })

        if hasattr(self, 'celulares_tree'):
            self.celulares_tree.delete(*self.celulares_tree.get_children())
            for f in filas:
                self.celulares_tree.insert('', tk.END, values=(
                    f['id'],
                    f['nombre'],
                    f['rubro'],
                    f['ip'],
                    f['ultima_actualizacion'],
                    f['estado'],
                ))
        else:
            # Fallback: listbox (no debería ocurrir)
            self.celulares_listbox.delete(0, tk.END)
            for f in filas:
                self.celulares_listbox.insert(tk.END, f['nombre'])

    def _dialogo_agregar_celular(self):
        nombre = simpledialog.askstring("Agregar celular", "Nombre del dispositivo:", parent=self.panel)
        if not nombre:
            return

        rubro = simpledialog.askstring("Agregar celular", "Rubro (opcional):", parent=self.panel)
        ip = simpledialog.askstring("Agregar celular", "IP (opcional):", parent=self.panel)

        try:
            crear_dispositivo(nombre=nombre, rubro=rubro, ip=ip)
            messagebox.showinfo("Celular", "Celular registrado en la base de datos", parent=self.panel)
            self._cargar_celulares()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar el celular: {e}", parent=self.panel)

    def _eliminar_celular_seleccionado(self):
        if not hasattr(self, 'celulares_tree'):
            return

        sel = self.celulares_tree.selection()
        if not sel:
            messagebox.showwarning("Eliminar celular", "Seleccione un dispositivo", parent=self.panel)
            return

        item_id = self.celulares_tree.item(sel[0], 'values')[0]
        try:
            device_id = int(item_id)
        except Exception:
            messagebox.showerror("Error", "ID de dispositivo inválido", parent=self.panel)
            return

        if not messagebox.askyesno("Eliminar", "¿Eliminar el dispositivo seleccionado?", parent=self.panel):
            return

        try:
            eliminado = eliminar_dispositivo(device_id)
            if not eliminado:
                messagebox.showwarning("Eliminar", "Dispositivo no encontrado", parent=self.panel)
            else:
                messagebox.showinfo("Eliminar", "Dispositivo eliminado", parent=self.panel)
                self._cargar_celulares()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo eliminar el dispositivo: {e}", parent=self.panel)

    def _exportar_registro(self):
        from .services.ingenieria_extras import listar_operaciones_por_fecha, listar_operaciones_todas
        from tkinter import filedialog
        fecha_raw = self.reg_fecha_var.get().strip()
        if not fecha_raw:
            messagebox.showerror("Error", "Seleccione una fecha válida")
            return
        if fecha_raw.lower() == 'todos':
            ops = listar_operaciones_todas()
        else:
            try:
                fecha = datetime.fromisoformat(fecha_raw).date()
            except Exception:
                messagebox.showerror("Error", "Formato de fecha inválido. Use yyyy-mm-dd o 'Todos'")
                return
            ops = listar_operaciones_por_fecha(fecha)
        if not ops:
            messagebox.showinfo("Exportar CSV", "No hay operaciones para la fecha seleccionada")
            return
        path = filedialog.asksaveasfilename(defaultextension='.csv', filetypes=[('CSV','*.csv')])
        if not path:
            return
        import csv
        with open(path, 'w', newline='', encoding='utf-8-sig') as csvf:
            writer = csv.writer(csvf, delimiter=';')
            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            writer.writerow(['Fecha', 'Máquina','Hora Inicio','Hora Fin','Horómetro Inicial','Horómetro Final','Hora informe','Operador','Combustible','Falla'])
            for o in ops:
                writer.writerow([
                    o.fecha.isoformat() if getattr(o, 'fecha', None) else '',
                    o.maquina_id,
                    o.hora_inicio.strftime('%H:%M'),
                    o.hora_fin.strftime('%H:%M'),
                    o.horometro_inicial,
                    o.horometro_final,
                    now,
                    o.operador,
                    getattr(o,'combustible',''),
                    'sí' if getattr(o,'hubo_falla',0) else 'no',
                ])

    def refrescar_lista(self, filter_ids=None, sort_by_categoria=False):
        """Cargar la lista de máquinas en el listbox.

        Si se pasan ``filter_ids`` solo se mostrarán esas máquinas; el
        parámetro ``sort_by_categoria`` ordena por categoría alfabéticamente.

        Si la base de datos está vacía no se inserta ningún registro de forma
        automática; el usuario debe usar el diálogo de "Agregar" para crear
        sus propias máquinas.  El botón de ejemplo en ese diálogo puede
        precargar los valores de la máquina de demostración si así se desea.
        """

        self.listbox.delete(0, tk.END)
        self.maquinas = listar_maquinas() or []
        # apply filtering/sorting if requested
        if filter_ids is not None:
            self.maquinas = [m for m in self.maquinas if m.id in filter_ids]
        if sort_by_categoria:
            self.maquinas.sort(key=lambda m: ((m.categoria or "").lower(), (m.nombre or "").lower()))

        # reset any column selection since indices may change
        self.comp_selected_machine_id = None
        # Auto-create demo machine if empty
        if not self.maquinas:
            self._ensure_demo()
            self.maquinas = listar_maquinas() or []
            if filter_ids is not None:
                self.maquinas = [m for m in self.maquinas if m.id in filter_ids]
            if sort_by_categoria:
                self.maquinas.sort(key=lambda m: ((m.categoria or "").lower(), (m.nombre or "").lower()))

        for m in self.maquinas:
            try:
                color = getattr(m, 'estado_color', 'verde')
            except Exception:
                # if something goes wrong (detached etc.), fall back gracefully
                color = 'verde'
            emoji = "🟢" if color == "verde" else ("🟡" if color == "amarillo" else "🔴")
            # substitute -- for missing values
            cat = m.categoria or "--"
            est = m.estado or "--"
            ubi = m.ubicacion or "--"
            fab = m.fabricante or "--"
            mod = m.modelo or "--"
            tipo_u = m.tipo_unidad or "--"
            ano = m.año or "--"
            fecha = m.fecha_actualizacion.strftime("%Y-%m-%d") if getattr(m, 'fecha_actualizacion', None) else "--"
            hor = getattr(m, 'horometro_actual', None)
            hor = hor if hor is not None else "--"
            material_name = m.material.nombre if getattr(m, 'material', None) else "--"
            self.listbox.insert(
                tk.END,
                f"{emoji} {m.id} | {m.nombre} | {tipo_u} | {fab} | {mod} | Año:{ano} | {cat} | {est} | {ubi} | Mat:{material_name} | {fecha} | Horómetro: {hor}",
            )
        # auto‑select first Máquina so that componentes/historial se muestran
        if self.maquinas:
            self.listbox.selection_set(0)

        # Actualizar listado de componentes si éste existe
        if hasattr(self, '_actualiza_componente_listado_por_sistema'):
            self._refresh_componentes()
        elif hasattr(self, '_actualiza_componente_listado'):
            try:
                self._actualiza_componente_listado()
            except Exception:
                pass

        # Actualizar filtro de máquinas para que el combobox refleje la lista actual
        try:
            self._refresh_filtro_maquinas()
        except Exception:
            pass

        # garantizar que los demás paneles se actualicen cuando cambie la lista de máquinas
        try:
            self.refrescar_programaciones()
        except Exception:
            pass
        try:
            self.refrescar_historial()
        except Exception:
            pass

    def _on_maquina_actualizada(self, maquina_id):
        """Callback invoked when a maquina record is updated elsewhere."""
        try:
            # Refresh the master list so Horómetro and other fields are current
            self.refrescar_lista()
        except Exception as e:
            print(f"[ERROR] refrescar_lista: {e}")

        # Refresh programaciones so horómetro timeline updates promptly
        try:
            self.refrescar_programaciones()
            self._mostrar_iteraciones()
        except Exception:
            pass

        if self._detalle_maquina_id == maquina_id and self._detalle_hor_entry:
            try:
                # Update horómetro field in the detail view if it's open
                maquina = next((m for m in self.maquinas if m.id == maquina_id), None)
                if maquina:
                    self._detalle_hor_entry.delete(0, tk.END)
                    self._detalle_hor_entry.insert(0, str(getattr(maquina, 'horometro_actual', 0) or 0))
            except Exception:
                pass

    def refrescar_historial(self):
        """Llena el árbol de historial aplicando filtros de máquina, acción y fechas."""
        if not hasattr(self, 'maquinas') or self.maquinas is None:
            self.maquinas = []

        # ── Update machine combo values ──
        maq_values = ["Todas"] + [f"{m.id} - {m.nombre}" for m in self.maquinas]
        self.hist_maq_combo['values'] = maq_values

        # ── Determine machine filter ──
        maq_sel = self.hist_maq_var.get()
        maq_ids = None  # None = todas
        if maq_sel and maq_sel != "Todas":
            try:
                maq_ids = [int(maq_sel.split(" - ")[0])]
            except Exception:
                maq_ids = None
        else:
            # If "Todas" but user has a listbox selection, use that if no explicit combo
            sel = self.listbox.curselection()
            if sel and maq_sel == "Todas":
                pass  # truly all machines

        self.hist_tree.delete(*self.hist_tree.get_children())
        db = SessionLocal()
        from datetime import datetime as _dt

        # ── Build queries ──
        hquery = db.query(Historial)
        cquery = db.query(Calibracion)
        if maq_ids:
            hquery = hquery.filter(Historial.maquina_id.in_(maq_ids))
            cquery = cquery.filter(Calibracion.maquina_id.in_(maq_ids))
        if self.hist_start_var.get():
            try:
                inicio = _dt.fromisoformat(self.hist_start_var.get())
                hquery = hquery.filter(Historial.fecha >= inicio)
                cquery = cquery.filter(Calibracion.fecha >= inicio)
            except ValueError:
                pass
        if self.hist_end_var.get():
            try:
                fin = _dt.fromisoformat(self.hist_end_var.get())
                hquery = hquery.filter(Historial.fecha <= fin)
                cquery = cquery.filter(Calibracion.fecha <= fin)
            except ValueError:
                pass

        hregs = hquery.all()
        cregs = cquery.all()
        db.close()

        # ── Map machine IDs to names ──
        maq_map = {m.id: f"{m.id} - {m.nombre}" for m in self.maquinas}

        # ── Mapping tipo_registro → origen (panel de origen) ──
        _origen_map = {
            "MANTENIMIENTO": "Registrar Mantenimiento",
            "REPARACIÓN": "Registrar Mantenimiento",
            "FALLA_RESUELTA": "Registrar Mantenimiento",
            "FALLA": "Solicitar Reparación",
            "CAMBIO_COMPONENTE": "Cambio de Componente",
            "CREACION_MATERIAL": "Gestión de Materiales",
            "AJUSTE_MATERIAL": "Gestión de Materiales",
            "ELIMINAR_MATERIAL": "Gestión de Materiales",
            "AGREGAR_PERSONAL": "Gestión de Personal",
            "ELIMINAR_PERSONAL": "Gestión de Personal",
        }

        # ── Build unified entry list ──
        entries = []
        for r in hregs:
            tipo_label = r.tipo_registro or r.accion or ""
            origen = _origen_map.get(tipo_label, tipo_label)
            entries.append((r.fecha, 'hist', r, tipo_label, origen))
        for c in cregs:
            entries.append((c.fecha, 'cal', c, f"CALIBRACION: {c.tipo}", "Registrar Calibración"))
        entries.sort(key=lambda x: x[0], reverse=True)

        # ── Apply origen (source panel) filter ──
        origen_sel = self.hist_origen_var.get()
        if origen_sel and origen_sel != "Todos":
            entries = [e for e in entries if e[4] == origen_sel]

        # ── Collect unique action types for the combo (after origen filter) ──
        acciones_unicas = sorted({e[3] for e in entries if e[3]})
        self.hist_accion_combo['values'] = ["Todas"] + acciones_unicas

        # ── Apply action filter ──
        accion_sel = self.hist_accion_var.get()
        if accion_sel and accion_sel != "Todas":
            entries = [e for e in entries if e[3] == accion_sel]

        # ── Populate tree ──
        for date, typ, obj, tipo_label, _origen in entries:
            try:
                lima = date.replace(tzinfo=ZoneInfo("UTC")).astimezone(ZoneInfo("America/Lima"))
            except Exception:
                lima = date
            fecha_str = lima.strftime("%Y-%m-%d %H:%M")
            maq_nombre = maq_map.get(getattr(obj, 'maquina_id', None), str(getattr(obj, 'maquina_id', '')))
            if typ == 'hist':
                descr = (obj.descripcion or "").replace("; ", "\n")
                iid = f"h{obj.id}"
            else:
                descr = f"dur {obj.duracion_minutos}m {(obj.observaciones or '')}".replace("; ", "\n")
                iid = f"c{obj.id}"
            self.hist_tree.insert("", tk.END, iid=iid, values=(fecha_str, maq_nombre, tipo_label, descr))

    def _exportar_historial_panel(self):
        """Exporta el historial visible (con filtros aplicados) a CSV o Excel."""
        rows = self.hist_tree.get_children()
        if not rows:
            messagebox.showinfo("Exportar", "No hay registros para exportar. Aplique filtros primero.")
            return
        headers = [self.hist_tree.heading(c, option='text') for c in self.hist_tree['columns']]
        data = []
        for iid in rows:
            data.append(list(self.hist_tree.item(iid, 'values')))

        fmt = self._ask_option("Formato", "Seleccione formato de exportación:", ["csv", "xlsx"])
        if not fmt:
            return

        if fmt == "csv":
            path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv")])
            if not path:
                return
            import csv
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f, delimiter=";")
                w.writerow(headers)
                for r in data:
                    w.writerow(r)
            messagebox.showinfo("Exportado", f"CSV guardado en:\n{path}")
        elif fmt == "xlsx":
            try:
                import openpyxl
            except ImportError:
                messagebox.showerror("Error", "Instale openpyxl: pip install openpyxl")
                return
            path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
            if not path:
                return
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Historial"
            ws.append(headers)
            for r in data:
                ws.append(r)
            wb.save(path)
            messagebox.showinfo("Exportado", f"Excel guardado en:\n{path}")

    def _actualiza_componente_listado(self):
        # If the old estilo of componentes list (column per machine) isn't available,
        # do nothing.
        if not hasattr(self, 'comp_tree'):
            return
        # rebuild columns: counter "#" + one per machine
        cols = ["_num"]
        headings = ["#"]
        for m in self.maquinas:
            key = f"m{m.id}"
            cols.append(key)
            headings.append(m.nombre or str(m.id))
        self.comp_tree.config(columns=cols)
        self.comp_tree.column("_num", width=35, anchor="center", stretch=False)
        self.comp_tree.heading("_num", text="#")
        for key, head in zip(cols[1:], headings[1:]):
            self.comp_tree.heading(key, text=head)
            self.comp_tree.column(key, width=150, anchor="w")
        aplicar_ordenamiento(self.comp_tree)
        # collect components grouped independently per machine
        comps = listar_componentes(None)
        machine_comps = {}  # {machine_col_key: [(name, desc_text), ...]}
        self._comp_id_map = {}  # {(col_key, row_index): component_id}
        machine_comp_indices = {}  # {col_key: current_row_counter}
        for c in comps:
            colkey = f"m{c.maquina_id}"
            if colkey not in cols:
                continue
            mat = getattr(c, 'material', None)
            extra = f" [{mat.nombre}]" if mat else ""
            text = f"{c.nombre}: {c.descripcion or ''}{extra}"
            machine_comps.setdefault(colkey, []).append(text)
            idx = machine_comp_indices.get(colkey, 0)
            # some test doubles may not have an id; default to None
            self._comp_id_map[(colkey, idx)] = getattr(c, 'id', None)
            machine_comp_indices[colkey] = idx + 1
        # determine max rows needed
        max_rows = max((len(v) for v in machine_comps.values()), default=0)
        self.comp_tree.delete(*self.comp_tree.get_children())
        # configure tags for cell highlighting
        self.comp_tree.tag_configure("highlight", background="#d0eaff")
        self.comp_tree.tag_configure("normal", background="")
        for i in range(max_rows):
            values = [str(i + 1)]  # row counter
            for key in cols[1:]:
                items = machine_comps.get(key, [])
                values.append(items[i] if i < len(items) else "")
            self.comp_tree.insert("", tk.END, values=values)
        # store machine column keys for click handler
        self._comp_cols = cols
        # also refresh historial for selected machine
        self.refrescar_historial()

    def _dialogo_materiales(self):

        from .services.ingenieria_extras import (
            listar_materiales,
            listar_componentes,
            crear_material,
            actualizar_material,
            eliminar_material,
        )

        def _get_sistemas():
            # Usa la lista de sistemas del panel, si existe; si no, usa valores por defecto.
            return getattr(
                self,
                "sistemas_lista",
                [],
            )

        def _build(content):

            # -------- LISTA DE MATERIALES --------
            tk.Label(content, text="Materiales registrados:").pack(pady=(5, 0))

            mat_list = tk.Listbox(content, height=8)
            mat_list.pack(fill=tk.X, padx=5, pady=5)

            def refresh_list():
                # Recalcular stock mínimo en base a cuántos componentes usan cada material.
                comps = listar_componentes(None)
                usage = {}
                for c in comps:
                    mid = getattr(c, "material_id", None)
                    if mid:
                        usage[mid] = usage.get(mid, 0) + 1

                mat_list.delete(0, tk.END)
                for m in listar_materiales():
                    required = usage.get(m.id, 0)
                    if getattr(m, "stock_minimo", None) != required:
                        try:
                            actualizar_material(m.id, stock_minimo=required)
                            m.stock_minimo = required
                        except Exception:
                            pass

                    sistema = getattr(m, "sistema", None) or ""
                    codigo = getattr(m, "codigo", "") or ""
                    tipo = getattr(m, "tipo", "") or ""
                    descripcion = getattr(m, "descripcion", "") or ""
                    texto = (
                        f"{m.id}  {m.nombre} [Código: {codigo}] [Tipo: {tipo}] "
                        f"[Stock: {m.stock_actual}/{m.stock_minimo}] [{sistema}] {descripcion}"
                    )
                    mat_list.insert(tk.END, texto)

            refresh_list()

            def selected_id():
                sel = mat_list.curselection()
                if not sel:
                    return None
                txt = mat_list.get(sel[0])
                try:
                    return int(txt.split(maxsplit=1)[0])
                except Exception:
                    return None

            def importar_desde_archivo():
                path = filedialog.askopenfilename(filetypes=[("Archivos", "*.csv;*.xlsx;*.xls")])
                if not path:
                    return
                try:
                    from .services.ingenieria_extras import importar_materiales_desde_archivo
                    resultado = importar_materiales_desde_archivo(path, update_existing=True)
                    messagebox.showinfo(
                        "Importado",
                        f"Importados {resultado.get('created', 0)} nuevos, actualizados {resultado.get('updated', 0)}.",
                        parent=self.panel,
                    )
                    refresh_list()
                    self._refresh_componentes()
                except Exception as e:
                    messagebox.showerror("Error", f"No se pudo importar:\n{e}", parent=self.panel)

            # -------- BOTONES LISTA --------
            btn_frame = tk.Frame(content)
            btn_frame.pack(pady=5)

            # -------- IMPORTAR --------
            _icon_btn(btn_frame, "📄", "Importar archivo", importar_desde_archivo).pack(side=tk.LEFT, padx=5)

            # -------- EDITAR --------
            def editar_mat():
                mid = selected_id()
                if mid is None:
                    messagebox.showwarning("Aviso", "Seleccione un material")
                    return

                mat = next((m for m in listar_materiales() if m.id == mid), None)
                if not mat:
                    messagebox.showerror("Error", "Material no encontrado")
                    return

                def _build_edit(econtent):
                    fields = {}
                    sistemas = _get_sistemas()

                    for label in ["nombre", "codigo", "tipo"]:
                        tk.Label(econtent, text=label.capitalize()).pack()

                        if label == "tipo":
                            combo = ttk.Combobox(econtent, values=["Fluido", "Sólido", "Gaseoso"], state="readonly")
                            combo.set(getattr(mat, label) or "")
                            combo.pack()
                            fields[label] = combo
                        else:
                            ent = tk.Entry(econtent)
                            ent.insert(0, getattr(mat, label) or "")
                            ent.pack()
                            fields[label] = ent

                    tk.Label(econtent, text="Descripción").pack()
                    desc_ent = tk.Entry(econtent)
                    desc_ent.insert(0, getattr(mat, "descripcion", "") or "")
                    desc_ent.pack()
                    fields["descripcion"] = desc_ent

                    tk.Label(econtent, text="Stock actual").pack()
                    stock_actual_ent = tk.Entry(econtent)
                    stock_actual_ent.insert(0, str(getattr(mat, "stock_actual", 0) or 0))
                    stock_actual_ent.pack()
                    fields["stock_actual"] = stock_actual_ent

                    tk.Label(econtent, text=f"Stock mínimo requerido: {getattr(mat, 'stock_minimo', 0)}").pack()

                    tk.Label(econtent, text="Sistema").pack()
                    sistema_combo = ttk.Combobox(econtent, values=sistemas, state="readonly")
                    actual = getattr(mat, "sistema", None)
                    if actual in sistemas:
                        sistema_combo.set(actual)
                    elif sistemas:
                        sistema_combo.set(sistemas[0])
                    sistema_combo.pack()

                    def save_edit():
                        sistema_val = sistema_combo.get()
                        if not sistema_val:
                            sistemas_list = sistema_combo["values"]
                            sistema_val = sistemas_list[0] if sistemas_list else ""

                        try:
                            nombre_val = fields["nombre"].get().strip()
                            tipo_val = fields["tipo"].get().strip()
                            if not nombre_val or not tipo_val:
                                messagebox.showwarning("Aviso", "Nombre y Tipo son obligatorios")
                                return

                            stock_actual_val = 0.0
                            try:
                                stock_actual_val = float(fields["stock_actual"].get().strip() or 0)
                            except Exception:
                                pass

                            actualizar_material(
                                mat.id,
                                nombre=nombre_val,
                                codigo=fields["codigo"].get().strip(),
                                tipo=tipo_val,
                                unidad="",
                                descripcion=fields["descripcion"].get().strip(),
                                stock_actual=stock_actual_val,
                                sistema=str(sistema_val),
                            )
                            refresh_list()
                            self._refresh_componentes()
                            self.view_mgr.pop()
                        except Exception as e:
                            messagebox.showerror("Error", str(e))

                    _icon_btn(econtent, "💾", "Guardar cambios", save_edit).pack(pady=10)

                self.view_mgr.push("Editar material", _build_edit)

            # -------- ELIMINAR --------
            def eliminar_mat():
                mid = selected_id()
                if mid is None:
                    messagebox.showwarning("Aviso", "Seleccione un material")
                    return

                if not messagebox.askyesno("Confirmar", "¿Eliminar material?"):
                    return

                try:
                    eliminar_material(mid)
                    refresh_list()
                    self._refresh_componentes()
                except Exception as e:
                    messagebox.showerror("Error", str(e))

            _icon_btn(btn_frame, "✏", "Editar", editar_mat).pack(side=tk.LEFT, padx=5)
            _icon_btn(btn_frame, "🗑", "Eliminar", eliminar_mat).pack(side=tk.LEFT, padx=5)

            mat_list.bind("<Double-1>", lambda e: editar_mat())

            # -------- FORMULARIO NUEVO --------
            form = tk.LabelFrame(content, text="Nuevo material")
            form.pack(fill=tk.X, padx=5, pady=10)

            fields = {}
            sistemas = _get_sistemas()

            for label in ["nombre", "codigo", "tipo"]:

                tk.Label(form, text=label.capitalize()).pack()

                if label == "tipo":

                    combo = ttk.Combobox(
                        form,
                        values=["Fluido", "Sólido", "Gaseoso"],
                        state="readonly",
                    )

                    combo.pack()
                    fields[label] = combo

                else:

                    ent = tk.Entry(form)
                    ent.pack()

                    fields[label] = ent

            tk.Label(form, text="Descripción").pack()
            desc_entry = tk.Entry(form)
            desc_entry.pack()
            fields["descripcion"] = desc_entry

            # -------- GUARDAR --------
            def guardar_mat():
                try:
                    nombre_val = fields["nombre"].get().strip()
                    tipo_val = fields["tipo"].get().strip()
                    codigo_val = fields["codigo"].get().strip()

                    if not nombre_val or not tipo_val:
                        messagebox.showwarning("Aviso", "Nombre y Tipo son obligatorios")
                        return
                    
                # 🔹 Validar duplicados
                    materiales_existentes = listar_materiales()
                    if any(m.nombre.lower() == nombre_val.lower() for m in materiales_existentes):
                        messagebox.showwarning("Aviso", f"Ya existe un material con nombre '{nombre_val}'")
                        return
                    if codigo_val and any(getattr(m, "codigo", "").lower() == codigo_val.lower() for m in materiales_existentes):
                        messagebox.showwarning("Aviso", f"Ya existe un material con código '{codigo_val}'")
                        return

                    # 🔹 Crear nuevo material
                    crear_material(
                        nombre_val,
                        codigo_val,
                        tipo_val,
                        "",  # unidad no se usa en esta UI
                        fields.get("descripcion", tk.StringVar()).get().strip(),
                        0,  # stock actual
                        0,  # stock mínimo
                        None,
                    )

                    for f in fields.values():
                        if hasattr(f, "delete"):
                            f.delete(0, tk.END)

                    refresh_list()
                    self._refresh_componentes()

                # Refrescar material en el diálogo de asociación si está abierto
                    if hasattr(self, '_asociar_mm_win') and getattr(self, '_asociar_mm_win', None) and self._asociar_mm_win.winfo_exists():
                        try:
                            self._asociar_mm_win.refresh_materials()
                        except Exception:
                            pass

                    messagebox.showinfo("Éxito", "Material creado")

                except Exception as e:
                    messagebox.showerror("Error", str(e))

            
            _icon_btn(form, "💾", "Guardar material", guardar_mat).pack(pady=10)
       
        self.view_mgr.push("Gestionar Materiales", _build)


    def _dialogo_asociar_maquina_material(self):
        """Permite seleccionar una máquina y un material y crear el componente asociado."""
        from .services.ingenieria_extras import agregar_componente

        if hasattr(self, '_asociar_mm_win') and getattr(self, '_asociar_mm_win', None) and self._asociar_mm_win.winfo_exists():
            win = self._asociar_mm_win
            win.deiconify()
            win.lift()
            return

        win = tk.Toplevel(self.panel)
        self._asociar_mm_win = win
        win.title("Asociar Máquina - Material")
        win.geometry("700x450")
        theme.apply_theme(win)

        sel_mat = {"id": None}
        sel_maq = {"id": None}

        def _update_button_state():
            if sel_mat.get("ids") and sel_maq.get("ids"):
                asociar_btn.config(state=tk.NORMAL)
            else:
                asociar_btn.config(state=tk.DISABLED)

        # --- Máquina / Material (vista compartida) ---
        dual_frame = tk.Frame(win)
        dual_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        left_frame = tk.Frame(dual_frame)
        right_frame = tk.Frame(dual_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0))

        tk.Label(left_frame, text="Materiales", font=(None, 11, "bold")).pack(pady=(0, 5))
        mat_list = tk.Listbox(left_frame, selectmode=tk.EXTENDED, exportselection=False)
        mat_list.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        def refresh_materials():
            mat_list.delete(0, tk.END)
            for m in listar_materiales():
                mat_list.insert(tk.END, f"{m.id}  {m.nombre} [Código:{m.codigo or ''}] [Tipo:{m.tipo or ''}]")

        # Exponer para refrescar desde otro diálogo (p.ej. crear material)
        win.refresh_materials = refresh_materials

        def _on_select_mat(event=None):
            sel = mat_list.curselection()
            if not sel:
                sel_mat["ids"] = []
            else:
                ids = []
                for idx in sel:
                    txt = mat_list.get(idx)
                    try:
                        ids.append(int(txt.split(maxsplit=1)[0]))
                    except Exception:
                        pass
                sel_mat["ids"] = ids
            _update_button_state()

        mat_list.bind("<<ListboxSelect>>", _on_select_mat)
        refresh_materials()

        tk.Label(right_frame, text="Máquinas", font=(None, 11, "bold")).pack(pady=(0, 5))
        maq_list = tk.Listbox(right_frame, selectmode=tk.EXTENDED, exportselection=False)
        maq_list.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        def refresh_maquinas():
            maq_list.delete(0, tk.END)
            for m in self.maquinas:
                maq_list.insert(tk.END, f"{m.id}  {m.nombre}")

        def _on_select_maq(event=None):
            sel = maq_list.curselection()
            if not sel:
                sel_maq["ids"] = []
            else:
                ids = []
                for idx in sel:
                    txt = maq_list.get(idx)
                    try:
                        ids.append(int(txt.split(maxsplit=1)[0]))
                    except Exception:
                        pass
                sel_maq["ids"] = ids
            _update_button_state()

        maq_list.bind("<<ListboxSelect>>", _on_select_maq)
        refresh_maquinas()

        # Asociar button
        asociar_btn = tk.Button(win, text="Asociar", state=tk.DISABLED, command=lambda: _asociar())
        asociar_btn.pack(padx=10, pady=(0, 10), anchor="e")

        def _asociar():
            if not sel_mat.get("ids") or not sel_maq.get("ids"):
                return
            try:
                mats = [m for m in listar_materiales() if m.id in sel_mat["ids"]]
                maquinas = [m for m in self.maquinas if m.id in sel_maq["ids"]]
                created = 0
                skipped = []
                for mat in mats:
                    if not mat:
                        continue
                    nombre_comp = mat.nombre or (mat.codigo or "Componente")
                    for maq in maquinas:
                        try:
                            agregar_componente(
                                maq.id,
                                nombre_comp,
                                descripcion="",
                                material_id=mat.id,
                                evitar_duplicados=True,
                            )
                            created += 1
                        except ValueError as dup_err:
                            # Cuando la asociación ya existe, evitamos duplicar
                            skipped.append(f"Máquina {maq.id} + Material {mat.id}")

                mensaje = f"Se crearon {created} componente(s) correctamente."
                if skipped:
                    mensaje += "\nSe omitieron duplicados:\n" + "\n".join(skipped)

                messagebox.showinfo("Asociado", mensaje, parent=win)
                self._refresh_componentes()
            except Exception as e:
                messagebox.showerror("Error", str(e), parent=win)

        win.protocol("WM_DELETE_WINDOW", win.withdraw)

    def _crea_tab_personal(self, notebook):
        self.tab_personal = ttk.Frame(notebook)
        notebook.add(self.tab_personal, text="Personal")

        # Rubros por defecto + los que existan en la BD
        self._rubros = ["TECNICO", "OPERARIO", "INGENIERO"]
        self._personal_trees = {}  # rol -> Treeview

        # ── Botones ──
        btns = tk.Frame(self.tab_personal)
        btns.pack(fill=tk.X, padx=10, pady=5)
        _icon_btn(btns, "➕", "Agregar persona", self._dialogo_agregar_persona).pack(side=tk.LEFT, padx=5)
        _icon_btn(btns, "🗑", "Eliminar seleccionado", self._eliminar_persona).pack(side=tk.LEFT, padx=5)
        _icon_btn(btns, "📌", "Crear nuevo rubro", self._crear_rubro).pack(side=tk.LEFT, padx=5)

        # ── Contenedor dinámico de columnas por rubro ──
        self._personal_cols_frame = tk.Frame(self.tab_personal)
        self._personal_cols_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

    def _crear_rubro(self):
        """Pide al usuario el nombre de un nuevo rubro y lo añade."""
        nombre = simpledialog.askstring("Nuevo rubro", "Nombre del rubro (ej: SUPERVISOR):", parent=self.panel)
        if not nombre:
            return
        nombre = nombre.strip().upper()
        if not nombre:
            return
        if nombre in self._rubros:
            messagebox.showinfo("Info", f"El rubro \"{nombre}\" ya existe.", parent=self.panel)
            return
        self._rubros.append(nombre)
        self.refrescar_personal()

    def _build_personal_columns(self):
        """Reconstruye las columnas de rubros dinámicamente."""
        # Limpiar contenedor
        for w in self._personal_cols_frame.winfo_children():
            w.destroy()
        self._personal_trees.clear()

        _emoji_map = {"TECNICO": "🔧", "OPERARIO": "👷"}

        for rol in self._rubros:
            emoji = _emoji_map.get(rol, "📌")
            lf = tk.LabelFrame(self._personal_cols_frame, text=f"{emoji} {rol}", padx=5, pady=5)
            lf.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=2)
            tree = ttk.Treeview(lf, columns=("id", "nombre", "fecha"), show="headings", selectmode="extended")
            tree.heading("id", text="ID")
            tree.heading("nombre", text="Nombre")
            tree.heading("fecha", text="Fecha registro")
            tree.column("id", width=40, anchor="center")
            tree.column("nombre", width=150)
            tree.column("fecha", width=120)
            sv = ttk.Scrollbar(lf, orient="vertical", command=tree.yview)
            tree.configure(yscrollcommand=sv.set)
            tree.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)
            sv.pack(fill=tk.Y, side=tk.RIGHT)
            self._personal_trees[rol] = tree

    def _dialogo_agregar_persona(self):
        def _build(content):
            tk.Label(content, text="Nombre").pack()
            name_entry = tk.Entry(content)
            name_entry.pack()
            tk.Label(content, text="Rol").pack()
            role_var = tk.StringVar(value=self._rubros[0] if self._rubros else "TECNICO")
            role_combo = ttk.Combobox(content, values=list(self._rubros), textvariable=role_var, state="readonly")
            role_combo.pack()
            tk.Label(content, text="PIN (4+ dígitos, acceso móvil)").pack()
            pin_entry = tk.Entry(content, show="*")
            pin_entry.pack()
            def save_person():
                nombre = name_entry.get().strip()
                rol = role_var.get().strip()
                pin = pin_entry.get().strip()
                if not nombre:
                    messagebox.showwarning("Aviso","Ingrese un nombre")
                    return
                if pin and len(pin) < 4:
                    messagebox.showwarning("Aviso","El PIN debe tener al menos 4 dígitos")
                    return
                from .services.ingenieria_extras import crear_persona
                from .services.historial_service import registrar_historial
                crear_persona(nombre, rol, pin=pin or None)
                registrar_historial(
                    maquina_id=None,
                    accion="AGREGAR_PERSONAL",
                    antes=None,
                    despues=f"{nombre} ({rol})",
                )
                self.view_mgr.pop()
                self.refrescar_personal()
            _icon_btn(content, "💾", "Guardar persona", save_person).pack(pady=5)
        self.view_mgr.push("Agregar persona", _build)

    def _eliminar_persona(self):
        """Elimina personas seleccionadas de cualquiera de las columnas."""
        all_sel = {}  # iid -> rol
        for rol, tree in self._personal_trees.items():
            for iid in tree.selection():
                all_sel[iid] = rol
        if not all_sel:
            messagebox.showwarning("Aviso", "Seleccione una o más personas para eliminar", parent=self.panel)
            return

        if not messagebox.askyesno("Confirmar", f"¿Está seguro de eliminar {len(all_sel)} persona(s)?", parent=self.panel):
            return

        from .services.ingenieria_extras import eliminar_persona
        from .services.historial_service import registrar_historial

        for iid, rol in all_sel.items():
            tree = self._personal_trees[rol]
            vals = tree.item(iid, "values")
            nombre = vals[1] if len(vals) > 1 else ""
            try:
                persona_id = int(iid)
                eliminar_persona(persona_id)
                registrar_historial(
                    maquina_id=None,
                    accion="ELIMINAR_PERSONAL",
                    antes=f"{nombre} ({rol})",
                    despues=None,
                )
            except Exception as e:
                messagebox.showerror("Error", str(e), parent=self.panel)

        messagebox.showinfo("Eliminado", f"{len(all_sel)} persona(s) eliminada(s)", parent=self.panel)
        self.refrescar_personal()

    def refrescar_personal(self):
        from .services.ingenieria_extras import listar_personas
        personas = listar_personas()
        # Incorporar rubros existentes en BD que no estén en la lista
        for p in personas:
            if p.rol and p.rol not in self._rubros:
                self._rubros.append(p.rol)
        # Reconstruir columnas y poblar
        self._build_personal_columns()
        for p in personas:
            rol = p.rol or "OPERARIO"
            tree = self._personal_trees.get(rol)
            if not tree:
                continue
            fecha_str = p.fecha_creacion.strftime("%Y-%m-%d %H:%M") if p.fecha_creacion else "--"
            tree.insert("", tk.END, iid=str(p.id), values=(p.id, p.nombre, fecha_str))







    def _crea_tab_programaciones(self, notebook):
   # """Pestaña principal de programaciones"""

        self.tab_prog = ttk.Frame(notebook)
        notebook.add(self.tab_prog, text="Programaciones")

        subnotebook = ttk.Notebook(self.tab_prog)
        subnotebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        try:
           self._crea_tab_prog_manual(subnotebook)
        except Exception as e:
           print("Error tab manual:", e)

        try:
            self._crea_tab_prediccion_fallas(subnotebook)
        except Exception as e:
             print("Error tab predicción:", e)


    def _crea_tab_prog_manual(self, notebook):
        tab = ttk.Frame(notebook)
        notebook.add(tab, text="✏️ Manual")

        paned = ttk.PanedWindow(tab, orient=tk.VERTICAL)
        paned.pack(fill=tk.BOTH, expand=True)

        top_frame = ttk.Frame(paned)
        paned.add(top_frame, weight=3)

        # ======================
        # BOTONES
        # ======================

        btn_bar = tk.Frame(top_frame)
        btn_bar.pack(fill=tk.X, padx=5, pady=3)

        _icon_btn(btn_bar, "➕", "Agregar programación",
                  self._dialogo_agregar_programacion).pack(side=tk.LEFT, padx=3)
        _icon_btn(btn_bar, "�", "Sincronizar Programaciones con Contabilidad",
                  self._sincronizar_programaciones_componentes).pack(side=tk.LEFT, padx=3)
        _icon_btn(btn_bar, "✅", "Autorizar/Desautorizar",
                  self._toggle_autorizacion_programacion).pack(side=tk.LEFT, padx=3)
        _icon_btn(btn_bar, "🗑", "Eliminar programación",
                  self._eliminar_programacion).pack(side=tk.LEFT, padx=3)
        _icon_btn(btn_bar, "♻️", "Resetear programaciones",
                  self._reset_programaciones).pack(side=tk.LEFT, padx=3)
        _icon_btn(btn_bar, "▼", "Mostrar iteraciones",
                  self._mostrar_iteraciones).pack(side=tk.LEFT, padx=3)
        tk.Button(btn_bar, text="Exportar CSV", command=self._exportar_programaciones_csv).pack(side=tk.RIGHT, padx=5)
        tk.Button(btn_bar, text="Exportar PDF", command=self._exportar_programaciones_pdf).pack(side=tk.RIGHT, padx=5)

        # ======================
        # TREEVIEW
        # ======================

        tree_frame = tk.Frame(top_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        self._prog_headings = {
            "_num": ("#", 40),
            "_maq": ("Máquina", 150),
            "_cantidad": ("Cantidad", 80),
            "_comp_asoc": ("Componente Asociado", 160),
            "_tarea": ("Tarea", 160),
            "_tipo": ("Tipo", 100),
            "_usado": ("Repuesto usado", 110),
            "_contabilidad": ("Contabilidad", 120),
            "_aut": ("Autorizado", 90),
            "_modalidad": ("Modalidad", 100),
            "_valor": ("Valor / Frecuencia", 140),
            "_horo_actual": ("Horómetro Actual", 120),
        }

        columns = tuple(self._prog_headings.keys())

        self.prog_tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show="headings",
            selectmode="extended"
        )

        # columnas
        for cid, (heading, width) in self._prog_headings.items():
            self.prog_tree.heading(cid, text=heading)
            self.prog_tree.column(
                cid,
                width=width,
                anchor="center" if cid == "_num" else "w"
            )

        # ======================
        # FILTROS
        # ======================

        filter_frame = tk.Frame(tree_frame, bg=theme.current.get("frame_bg"))
        filter_frame.pack(fill=tk.X)

        self.prog_filters = {}
        for cid, (heading, width) in self._prog_headings.items():
            var = tk.StringVar(value="")
            combo = ttk.Combobox(
                filter_frame,
                textvariable=var,
                width=int(width / 10),
                state="readonly",
                style="TCombobox",
            )
            combo.pack(side=tk.LEFT, padx=2)
            combo["values"] = [""]
            self.prog_filters[cid] = (var, combo)
            combo.bind(
                "<<ComboboxSelected>>",
                lambda e: self._safe_refrescar_programaciones()
            )

        # ======================
        # SCROLLBARS
        # ======================

        vscroll = ttk.Scrollbar(tree_frame, orient="vertical",
                                command=self.prog_tree.yview)
        hscroll = ttk.Scrollbar(tree_frame, orient="horizontal",
                                command=self.prog_tree.xview)
        self.prog_tree.configure(
            yscrollcommand=vscroll.set,
            xscrollcommand=hscroll.set
        )
        self.prog_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vscroll.pack(side=tk.RIGHT, fill=tk.Y)
        hscroll.pack(side=tk.BOTTOM, fill=tk.X)

        aplicar_ordenamiento(self.prog_tree)
        self.prog_tree.bind(
            "<Double-1>",
            self._on_prog_tree_double_click
        )
        # cargar datos iniciales
        self._safe_refrescar_programaciones()

        # =========================
        # PANEL INFERIOR
        # =========================

        bot_frame = ttk.Frame(paned)
        paned.add(bot_frame, weight=1)

        ctrl = tk.Frame(bot_frame)
        ctrl.pack(fill=tk.X, padx=5, pady=3)

        tk.Label(ctrl, text="Mostrar:").pack(side=tk.LEFT)

        self.prog_iter_mode = tk.StringVar(value="horometro")
        ttk.Radiobutton(
            ctrl,
            text="Horómetro",
            variable=self.prog_iter_mode,
            value="horometro",
            command=self._mostrar_iteraciones
        ).pack(side=tk.LEFT)
        ttk.Radiobutton(
            ctrl,
            text="Fecha",
            variable=self.prog_iter_mode,
            value="fecha",
            command=self._mostrar_iteraciones
        ).pack(side=tk.LEFT)
        ttk.Radiobutton(
            ctrl,
            text="Ambos",
            variable=self.prog_iter_mode,
            value="ambos",
            command=self._mostrar_iteraciones
        ).pack(side=tk.LEFT)

        tk.Label(ctrl, text=" Iteraciones:").pack(side=tk.LEFT)

        self.prog_iter_count = tk.IntVar(value=10)
        tk.Spinbox(
            ctrl,
            from_=1,
            to=50,
            textvariable=self.prog_iter_count,
            width=4,
            command=self._mostrar_iteraciones
        ).pack(side=tk.LEFT)

        # =========================
        # FILTRO DE MÁQUINAS
        # =========================

        mid_frame = tk.Frame(bot_frame)
        mid_frame.pack(fill=tk.X, padx=5)

        tk.Label(mid_frame, text="Máquinas:").pack(side=tk.LEFT)

        self.iter_maq_listbox = tk.Listbox(
            mid_frame,
            selectmode=tk.EXTENDED,
            height=4,
            exportselection=False
        )
        self.iter_maq_listbox.pack(side=tk.LEFT, fill=tk.X, expand=True)

        maq_sb = ttk.Scrollbar(
            mid_frame,
            orient="vertical",
            command=self.iter_maq_listbox.yview
        )
        self.iter_maq_listbox.configure(yscrollcommand=maq_sb.set)
        maq_sb.pack(side=tk.LEFT, fill=tk.Y)

        # =========================
        # TABLA ITERACIONES
        # =========================

        iter_frame = tk.Frame(bot_frame)
        iter_frame.pack(fill=tk.BOTH, expand=True)

        self.iter_tree = ttk.Treeview(iter_frame, show="headings")

        v2 = ttk.Scrollbar(
            iter_frame,
            orient="vertical",
            command=self.iter_tree.yview
        )
        self.iter_tree.configure(yscrollcommand=v2.set)
        v2.pack(side=tk.RIGHT, fill=tk.Y)
        self.iter_tree.pack(fill=tk.BOTH, expand=True)
        self.iter_tree.bind("<Double-1>", self._on_iteracion_double_click)

        # compatibilidad con código existente
        self.prog_selection_label = tk.Label(bot_frame, text="", anchor="w")
        self.prog_list = tk.Listbox(bot_frame, height=0)

        # =========================
        # CARGA INICIAL
        # =========================

        self._safe_refrescar_programaciones()
    
    


    def _safe_refrescar_programaciones(self):
        """Refresco robusto que evita que desaparezcan datos"""
        
        try:
            # keep machine list current too — import from the service module like
            # everywhere else.  This was previously wrong and raised
            # ImportError during startup.
            from .services.ingenieria_service import listar_maquinas
            self.maquinas = listar_maquinas() or []
            self.refrescar_programaciones()
        
        except Exception as e:
        
            import traceback
            traceback.print_exc()
        
            # si el refresco falla, no borrar el contenido existente
            if hasattr(self, "prog_tree"):
        
                if not self.prog_tree.get_children():
        
                    self.prog_tree.insert(
                        "",
                        "end",
                        values=("ERROR", "No se pudieron cargar datos", "", "", "", "", "", "", "", "")
                    )
        
        
        
        
        
        
    def _crea_tab_prediccion_fallas(self, notebook):
        """Tab para predicción de fallas — redirige al Análisis Avanzado."""
        tab = ttk.Frame(notebook)
        notebook.add(tab, text="🔮 Predicción de Fallas")
        
        frame_info = tk.Label(tab, text="""
╔═══════════════════════════════════════════════════════════╗
║   PREDICCIÓN DE FALLAS — ANÁLISIS AVANZADO                ║
║   Los cálculos ahora se ejecutan en el Sistema de         ║
║   Análisis Avanzado (PyQt6) para mayor rendimiento.       ║
╚═══════════════════════════════════════════════════════════╝

Funcionalidades disponibles en Análisis Avanzado:
• Series Temporales con pyqtgraph (más fluido)
• Regresión lineal y predicción a 180 días
• Análisis Físico (Termodinámica, Torque, Fricción)
• Estimación de Vida Útil con barra de progreso
• Dashboard de Alertas automáticas
• Plano General colaborativo con Fourier

Haz clic en el botón abajo para abrir el análisis.
        """, justify=tk.LEFT, font=("Courier", 10), padx=20, pady=20)
        frame_info.pack(fill=tk.BOTH, expand=True)
        
        btn_frame = tk.Frame(tab)
        btn_frame.pack(pady=20)
        _icon_btn(btn_frame, "🧠", "Abrir Análisis Avanzado", self._abrir_analisis_avanzado,
                 font=("Arial", 12, "bold")).pack(ipadx=16, ipady=6)

    def _iter_select_all_maq(self):
        """Selecciona todas las máquinas en el listbox de iteraciones."""
        self.iter_maq_listbox.selection_set(0, tk.END)
        
    def refrescar_programaciones(self):
        """Recarga Panel 1 — una fila por programación."""

        # Esta función puede ser llamada muy temprano durante init, antes de que
        # se creen los widgets de la pestaña de Programaciones.
        if not hasattr(self, 'prog_tree') or not hasattr(self, 'iter_maq_listbox'):
            return

        try:

            from .services.ingenieria_extras import listar_programaciones, obtener_uso_componentes
            from .services.ingenieria_service import listar_maquinas

            # cache del uso de componentes por máquina para reflejar contabilidad
            uso_componentes = obtener_uso_componentes()
            uso_map = {}
            for u in uso_componentes:
                key = (u.get('maquina_id'), (u.get('componente') or '').strip().lower())
                uso_map.setdefault(key, []).append(u)

            # -------------------------
            # MAQUINAS
            # -------------------------

            self.maquinas = listar_maquinas() or []
            maq_map = {m.id: m for m in self.maquinas}

            # -------------------------
            # PROGRAMACIONES
            # -------------------------

            all_progs = listar_programaciones() or []

            # ordenar por campo orden
            all_progs.sort(key=lambda p: getattr(p, "orden", p.id))

            # -------------------------
            # RECORDAR ORDEN DE COLUMNAS
            # -------------------------

            sort_col = None
            sort_reverse = False

            for c in self.prog_tree["columns"]:

                txt = self.prog_tree.heading(c, "text")

                if txt.endswith(" ▲"):
                    sort_col = c
                    break

                if txt.endswith(" ▼"):
                    sort_col = c
                    sort_reverse = True
                    break

            # -------------------------
            # PANEL 2 MAQUINAS
            # -------------------------

            prev_sel = list(self.iter_maq_listbox.curselection())

            self.iter_maq_listbox.delete(0, tk.END)

            maq_ids = sorted({p.maquina_id for p in all_progs})

            self._iter_maq_ids = maq_ids

            for mid in maq_ids:

                maq = maq_map.get(mid)

                nombre = maq.nombre if maq else "[Eliminada]"

                self.iter_maq_listbox.insert(
                    tk.END,
                    f"{mid} - {nombre}"
                )

            if prev_sel:
                for idx in prev_sel:
                    if idx < self.iter_maq_listbox.size():
                        self.iter_maq_listbox.selection_set(idx)
            else:
                self.iter_maq_listbox.selection_set(0, tk.END)

            # -------------------------
            # LIMPIAR TREE
            # -------------------------

            self.prog_tree.delete(*self.prog_tree.get_children())

            filter_values = {cid: set() for cid in self.prog_tree["columns"]}

            # -------------------------
            # INSERTAR PROGRAMACIONES
            # -------------------------

            for i, p in enumerate(all_progs, 1):

                maq = maq_map.get(p.maquina_id)

                maq_nombre = (
                    f"{p.maquina_id} - {maq.nombre}"
                    if maq else f"{p.maquina_id} - [Eliminada]"
                )

                horo_actual = (
                    maq.horometro_actual
                    if maq and maq.horometro_actual else "--"
                )

                freq = getattr(p, "frecuencia_horas", None)
                horo_obj = getattr(p, "horometro_objetivo", None)

                modalidad = getattr(p, "horometro_modalidad", "--")

                valor = horo_obj if horo_obj else (freq if freq else "--")

                comp_asoc = "--"
                comp_tipo = "--"
                mantenimiento_tiene_uso = "No"

                if getattr(p, "componente", None):
                    comp = p.componente
                    if getattr(comp, "material", None):
                        comp_asoc = comp.material.nombre
                        tipo_mat = (getattr(comp.material, "tipo", "") or "").strip().lower()
                        if "gase" in tipo_mat:
                            comp_tipo = "Gaseoso"
                        elif "flu" in tipo_mat or "líqu" in tipo_mat:
                            comp_tipo = "Líquido"
                        else:
                            comp_tipo = "Sólido"
                    else:
                        comp_asoc = getattr(comp, "nombre", "--")

                # Marcar si ya hay componente contabilizado en el historial de piezas
                key = (p.maquina_id, (comp_asoc or "").strip().lower())
                if key in uso_map and uso_map[key]:
                    mantenimiento_tiene_uso = "Sí"
                    contabilidad_count = len(uso_map[key])
                else:
                    contabilidad_count = 0

                cantidad_val = getattr(p, "cantidad", None)

                cantidad_val = getattr(p, "cantidad", None)
                estado = getattr(p, "estado_repuesto", "")

                if cantidad_val is not None:
                    if estado == "FLUIDO":
                        cantidad_str = f"{cantidad_val:g} L"
                    else:
                        cantidad_str = f"{cantidad_val:g}"
                else:
                    cantidad_str = "--"

                autorizado_str = "Sí" if getattr(p, "autorizado", 0) else "No"

                row = (
                    i,
                    maq_nombre,
                    cantidad_str,
                    comp_asoc,
                    mantenimiento_tiene_uso,
                    contabilidad_count,
                    p.tipo or "--",
                    comp_tipo,
                    autorizado_str,
                    modalidad,
                    valor,
                    horo_actual,
                )

                # recolectar valores filtros

                for idx, cid in enumerate(self.prog_tree["columns"]):
                    filter_values[cid].add(str(row[idx]))

                # aplicar filtros

                show = True

                for idx, cid in enumerate(self.prog_tree["columns"]):

                    if cid not in self.prog_filters:
                        continue

                    var, _ = self.prog_filters[cid]

                    val = var.get()

                    heading = self._prog_headings[cid][0]

                    if val != heading and str(row[idx]) != val:
                        show = False
                        break

                if show:

                    self.prog_tree.insert(
                        "",
                        tk.END,
                        iid=str(p.id),  # ID real
                        values=row
                    )

            # -------------------------
            # ACTUALIZAR FILTROS
            # -------------------------

            for cid, (var, combo) in self.prog_filters.items():

                values = sorted(filter_values[cid])

                combo["values"] = [self._prog_headings[cid][0]] + values

                if var.get() not in combo["values"]:
                    var.set(self._prog_headings[cid][0])

            # -------------------------
            # REAPLICAR ORDEN
            # -------------------------

            if sort_col:

                from .treeview_utils import treeview_sort_column

                treeview_sort_column(
                    self.prog_tree,
                    sort_col,
                    sort_reverse
                )

        except Exception as err:

            print(f"[ERROR] refrescar_programaciones falló: {err}")

    # ── Iteration matrix ──
    
    
    
    def _get_avg_daily_hours(self, maquina_id):
        """Calcula horas promedio diarias de operación basándose en OperacionDiaria."""
        try:
            from .models import OperacionDiaria
            from .database import SessionLocal
            from sqlalchemy import func
            with SessionLocal() as sess:
                avg = sess.query(func.avg(OperacionDiaria.horas_trabajadas)).filter(
                    OperacionDiaria.maquina_id == maquina_id
                ).scalar()
                return float(avg) if avg else None
        except Exception:
            return None
    
    def _build_iteration_data(self):
        """Construye datos de iteraciones para las programaciones seleccionadas.
        Retorna (headers, rows, row_colors) donde rows = lista de listas de strings
        y row_colors = lista de listas de colores ('verde'|'amarillo'|'rojo'|'realizado')
        por cada celda de iteración.
        """
        import re
        from .services.ingenieria_extras import listar_programaciones, _get_realizados, obtener_uso_componentes
        from datetime import timedelta

        def _normalize_component(comp):
            if not comp:
                return ""
            c = str(comp).strip().lower()
            m = re.search(r"\[([^\]]+)\]", c)
            if m:
                val = m.group(1).strip()
                if val:
                    c = val.lower()
            c = re.sub(r"^\d+\s*-\s*", "", c)
            c = re.sub(r"\s*\(.*$", "", c)
            return c.strip()

        usos_contabilidad = obtener_uso_componentes() or []
        uso_map = {}
        for u in usos_contabilidad:
            maq_id = u.get('maquina_id')
            comp = _normalize_component(u.get('componente'))
            horo = u.get('horometro')
            if maq_id is None or not comp or horo is None:
                continue
            key = (maq_id, comp)
            uso_map.setdefault(key, set()).add(float(horo))
        import json as _json
    
        # Filter by machines selected in Panel 2 listbox
        maq_sel_indices = list(self.iter_maq_listbox.curselection())
        iter_maq_ids = getattr(self, '_iter_maq_ids', [])
        if maq_sel_indices and iter_maq_ids:
            selected_maq_ids = {iter_maq_ids[i] for i in maq_sel_indices if i < len(iter_maq_ids)}
        else:
            selected_maq_ids = None  # all
    
        all_progs = listar_programaciones()
        if selected_maq_ids:
            progs_sel = [p for p in all_progs if p.maquina_id in selected_maq_ids]
        else:
            progs_sel = list(all_progs)
    
        maq_map = {m.id: m for m in (self.maquinas or [])}
        n = self.prog_iter_count.get()
        mode = self.prog_iter_mode.get()
    
        # Build columns
        headers = ["Máquina", "Cantidad", "Componente Asociado", "Tarea"]
        for i in range(n):
            headers.append(f"Iter {i}")
    
        rows = []
        row_colors = []
        milestones_matrix = []
        prog_ids = []

        for p in progs_sel:
            prog_ids.append(getattr(p, 'id', None))
            maq = maq_map.get(p.maquina_id)
            maq_nombre = f"{p.maquina_id} - {maq.nombre}" if maq else str(p.maquina_id)
            freq = getattr(p, 'frecuencia_horas', None) or 0
            horo_obj = getattr(p, 'horometro_objetivo', None) or 0
            modalidad = getattr(p, 'horometro_modalidad', None) or "OBJETIVO"
            horo_ini = maq.horometro_inicial if maq and maq.horometro_inicial else 0
            horo_act = maq.horometro_actual if maq and maq.horometro_actual else horo_ini

            realizados = _get_realizados(p) or set()
            comp_asoc = "--"
            if getattr(p, "componente", None) and getattr(p.componente, "material", None):
                comp_asoc = p.componente.material.nombre
            elif getattr(p, "componente", None):
                comp_asoc = p.componente.nombre

            comp_asoc_key = (p.maquina_id, (comp_asoc or "").strip().lower())
            contab_horometros = uso_map.get(comp_asoc_key, set())
            todos_realizados = set(realizados) | set(contab_horometros)

            # Cantidad / Componente Asociado
            cantidad_val = getattr(p, 'cantidad', None)
            estado = getattr(p, 'estado_repuesto', None) or ""
            if cantidad_val is not None:
                cantidad_str = f"{cantidad_val:g} L" if estado == "FLUIDO" else f"{cantidad_val:g}"
            else:
                cantidad_str = "--"
            comp_asoc = "--"
            if getattr(p, 'componente', None) and getattr(p.componente, 'material', None):
                comp_asoc = p.componente.material.nombre

            # Compute milestones from el horómetro en el que se usó el componente (o el último hito completado).
            # De esta forma el tren parte del último paradero real y no se rehace solo con el horómetro actual.
            todos_realizados = set(realizados) | set(contab_horometros)
            contab_last = max(contab_horometros) if contab_horometros else None
            real_last = max(realizados) if realizados else None
            last_done = contab_last if contab_last is not None else real_last

            if freq > 0 and last_done is not None:
                # Recalcular el siguiente horómetro desde el último mantenimiento de contabilidad/previo.
                base = last_done
            elif horo_obj > 0:
                base = horo_obj
            elif last_done is not None:
                base = last_done
            else:
                base = horo_ini

            milestones = []
            if freq > 0:
                # Si modalidad secuencial, valores relativos al primer punto predeterminado.
                if modalidad == "SECUENCIAL":
                    start = base if base is not None else horo_ini
                    for i in range(n):
                        milestones.append(start + freq * (i + 1))
                else:
                    start = base
                    for i in range(n):
                        milestones.append(start + freq * (i + 1))
            elif horo_obj > 0:
                # Si hay objetivo único, usarlo y generar a partir de ahí para profundidad n
                milestones = [horo_obj]
                for i in range(1, n):
                    milestones.append(horo_obj + i * (freq or 0 or 150))
            else:
                milestones = [horo_act] * n

            # Combinar con los horómetros número registrados en contabilidad para coindicencias exactas.
            contab_key = (p.maquina_id, (comp_asoc or "").strip().lower())
            contab_horometros = uso_map.get(contab_key, set())
            combined = sorted(set(milestones) | set(contab_horometros))

            # Garantizar al menos n hitos (para la tabla de iteraciones)
            while len(combined) < n:
                if combined:
                    step = freq if freq and freq > 0 else (horo_obj if horo_obj > 0 else 0)
                    next_val = combined[-1] + (step or 1)
                else:
                    next_val = horo_act
                combined.append(next_val)

            milestones = combined[:n]

            # Compute estimated dates if needed
            avg_daily = self._get_avg_daily_hours(p.maquina_id) if mode in ("fecha", "ambos") else None
            from datetime import date as _date
            today = _date.today()

            row = [maq_nombre, cantidad_str, comp_asoc, p.tipo or "--"]
            colors = ["", "", "", ""]  # no color for fixed columns
            for idx_ms, ms in enumerate(milestones):
                # Determine color for this milestone cell
                in_transit = False
                if ms in contab_horometros:
                    cell_color = "verde"
                elif ms in realizados:
                    cell_color = "realizado"
                else:
                    if idx_ms == 0:
                        previous_target = horo_ini
                    else:
                        previous_target = milestones[idx_ms - 1]

                    in_transit = horo_act >= previous_target and horo_act < ms

                    if in_transit:
                        cell_color = "actual"
                    elif ms <= horo_act:
                        cell_color = "rojo"
                    else:
                        prev_base = horo_ini
                        for prev_ms in milestones:
                            if prev_ms >= ms:
                                break
                            if prev_ms in realizados:
                                prev_base = prev_ms
                        intervalo = ms - prev_base
                        avance = horo_act - prev_base
                        if intervalo > 0 and avance / intervalo >= 0.7:
                            cell_color = "amarillo"
                        else:
                            cell_color = "verde"
                colors.append(cell_color)

                # Visual line and milestone markers (carril continuo)
                base_line = "──"
                if ms in contab_horometros or ms in realizados:
                    marker = "|"  # parada + mantenimiento realizado (contabilidad/realizado)
                elif in_transit:
                    marker = "▶"  # posición actual
                else:
                    marker = " "  # todavía por recorrer

                if mode == "horometro":
                    cell_text = f"{base_line}{marker}{int(ms)}"
                elif mode == "fecha":
                    if avg_daily and avg_daily > 0:
                        remaining = ms - horo_act
                        if remaining <= 0:
                            cell_text = today.isoformat()
                        else:
                            days = remaining / avg_daily
                            est = today + timedelta(days=days)
                            cell_text = est.strftime("%d/%m/%Y")
                    else:
                        cell_text = "Sin datos"
                    if ms in realizados:
                        cell_text = f"✅ {cell_text}"
                else:  # ambos
                    cell = f"{base_line}{marker}{int(ms)}"
                    if avg_daily and avg_daily > 0:
                        remaining = ms - horo_act
                        if remaining <= 0:
                            cell += f" ({today.strftime('%d/%m/%Y')})"
                        else:
                            days = remaining / avg_daily
                            est = today + timedelta(days=days)
                            cell += f" ({est.strftime('%d/%m/%Y')})"
                    if ms in realizados:
                        cell = f"✅ {cell}"
                    cell_text = cell

                row.append(cell_text)
            rows.append(row)
            row_colors.append(colors)
            milestones_matrix.append(milestones)

        return headers, rows, row_colors, milestones_matrix, prog_ids

    def _mostrar_iteraciones(self):
        """Actualiza Panel 2 con la matriz de iteraciones."""
        headers, rows, row_colors, milestones, prog_ids = self._build_iteration_data()
    
        self.iter_tree.delete(*self.iter_tree.get_children())
        cols = [f"_c{i}" for i in range(len(headers))]
        self.iter_tree["columns"] = cols
        fixed_count = 4  # Máquina, Cantidad, Componente Asociado, Tarea
        fixed_widths = [130, 80, 150, 150]
        for i, h in enumerate(headers):
            cid = cols[i]
            w = fixed_widths[i] if i < fixed_count else 110
            self.iter_tree.heading(cid, text=h)
            self.iter_tree.column(cid, width=w, anchor="center" if i >= fixed_count else "w")
    
        # Configure color tags
        if theme.is_dark():
            self.iter_tree.tag_configure("verde", background="#2E4F2E")
            self.iter_tree.tag_configure("amarillo", background="#5A4F2E")
            self.iter_tree.tag_configure("rojo", background="#5A2E2E")
            self.iter_tree.tag_configure("realizado", background="#37474F")
            self.iter_tree.tag_configure("actual", background="#264f78")
        else:
            self.iter_tree.tag_configure("verde", background="#C8E6C9")
            self.iter_tree.tag_configure("amarillo", background="#FFF9C4")
            self.iter_tree.tag_configure("rojo", background="#FFCDD2")
            self.iter_tree.tag_configure("realizado", background="#B0BEC5")
            self.iter_tree.tag_configure("actual", background="#BBDEFB")
    
        # Keep a map of programacion id -> milestone list so we can mark iterations as done on click
        self._iter_meta = {}

        for idx, r in enumerate(rows):
            # Determine the dominant tag for the row (first non-empty color from iterations)
            # We use the color of the first pending (non-realizado) iteration
            if idx < len(row_colors):
                iter_colors = [c for c in row_colors[idx][fixed_count:] if c and c != "realizado"]
                tag = iter_colors[0] if iter_colors else "verde"
            else:
                tag = "verde"

            item_id = None
            if idx < len(prog_ids) and prog_ids[idx] is not None:
                item_id = f"prog_{prog_ids[idx]}"

            iid = self.iter_tree.insert("", tk.END, iid=item_id, values=r, tags=(tag,))
            if item_id and idx < len(milestones):
                self._iter_meta[item_id] = milestones[idx]

    def _on_iteracion_double_click(self, event):
        """Mark an iteration milestone as completed when user double-clicks a cell."""
        try:
            item = self.iter_tree.identify_row(event.y)
            col = self.iter_tree.identify_column(event.x)
            if not item or not col:
                return
            if item not in getattr(self, '_iter_meta', {}):
                return
            col_idx = int(col.replace('#', '')) - 1
            fixed_count = 4
            if col_idx < fixed_count:
                return
            milestones = self._iter_meta.get(item, [])
            idx = col_idx - fixed_count
            if idx < 0 or idx >= len(milestones):
                return
            horo = milestones[idx]
            if not messagebox.askyesno("Marcar mantenimiento", f"Marcar como realizado el mantenimiento en horómetro {horo}? "):
                return
            from .services.ingenieria_extras import registrar_mantenimiento_programacion
            # derive programacion id from tree item id
            if item.startswith("prog_"):
                prog_id = int(item.split("prog_")[1])
                registrar_mantenimiento_programacion(prog_id, horo)
                self._mostrar_iteraciones()
        except Exception:
            pass

    def _exportar_programaciones_csv(self):
        """Exporta SOLO las filas visibles de Programaciones según los filtros activos a CSV."""
        # Exporta ambas tablas (programaciones e iteraciones) a Excel, cada una en su hoja
        from tkinter import filedialog, messagebox
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Error", "Instale openpyxl: pip install openpyxl")
            return
        try:
            path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
            if not path:
                return
            # Obtener datos de programaciones e iteraciones
            p1_headers, p1_rows = self._get_panel1_data()
            p2_headers, p2_rows, *_ = self._build_iteration_data()
            wb = openpyxl.Workbook()
            # Hoja 1: Programaciones
            ws1 = wb.active
            ws1.title = "Programaciones"
            ws1.append(p1_headers)
            for r in p1_rows:
                ws1.append(list(r))
            # Hoja 2: Iteraciones
            ws2 = wb.create_sheet(title="Iteraciones")
            ws2.append(p2_headers)
            for r in p2_rows:
                ws2.append(r)
            wb.save(path)
            messagebox.showinfo("Exportado", f"Excel guardado en:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _exportar_programaciones_pdf(self):
        """Exporta SOLO las filas visibles de Programaciones según los filtros activos a PDF."""
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
        from reportlab.lib import colors
        from tkinter import filedialog, messagebox

        try:
            path = filedialog.asksaveasfilename(defaultextension=".pdf",
                                                filetypes=[("PDF Files", "*.pdf")])
            if not path:
                return

            columns = self.prog_tree["columns"]
            headers = [self.prog_tree.heading(c, "text") for c in columns]

            # Solo filas visibles
            rows = [list(self.prog_tree.item(iid, "values")) for iid in self.prog_tree.get_children()]

            data = [headers] + rows

            pdf = SimpleDocTemplate(path, pagesize=landscape(letter))
            table = Table(data)
            style = TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightblue),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ])
            table.setStyle(style)
            pdf.build([table])

            messagebox.showinfo("Exportado", f"Programaciones exportadas correctamente a {path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _get_panel1_data(self):
        """Recopila los datos del Panel 1 (lista de programaciones) para exportación."""
        headers = ["#", "Máquina", "Cantidad", "Componente Asociado", "Tarea",
                   "Tipo", "Modalidad", "Valor / Frecuencia", "Horómetro Actual", "Componente"]
        rows = []
        for child in self.prog_tree.get_children():
            rows.append(list(self.prog_tree.item(child, "values")))
        return headers, rows
    
    def _exportar_iteraciones(self, fmt="csv"):
        """Exporta ambos paneles: Panel 1 (programaciones) y Panel 2 (iteraciones)."""
        from tkinter import filedialog
        p1_headers, p1_rows = self._get_panel1_data()
        p2_headers, p2_rows, _colors, _m, _ids = self._build_iteration_data()
        if not p1_rows and not p2_rows:
            messagebox.showinfo("Exportar", "No hay datos para exportar")
            return
    
        if fmt == "csv":
            path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv")])
            if not path:
                return
            import csv
            with open(path, "w", encoding="utf-8-sig", newline="") as f:
                w = csv.writer(f, delimiter=";")
                w.writerow(["=== Programaciones ==="])
                w.writerow(p1_headers)
                for r in p1_rows:
                    w.writerow(r)
                w.writerow([])
                w.writerow(["=== Iteraciones ==="])
                w.writerow(p2_headers)
                for r in p2_rows:
                    w.writerow(r)
            messagebox.showinfo("Exportado", f"CSV guardado en:\n{path}")
        elif fmt == "excel":
            try:
                import openpyxl
            except ImportError:
                messagebox.showerror("Error", "Instale openpyxl: pip install openpyxl")
                return
            path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
            if not path:
                return
            wb = openpyxl.Workbook()
            # Hoja 1: Programaciones
            ws1 = wb.active
            ws1.title = "Programaciones"
            ws1.append(p1_headers)
            for r in p1_rows:
                ws1.append(list(r))
            # Hoja 2: Iteraciones
            ws2 = wb.create_sheet(title="Iteraciones")
            ws2.append(p2_headers)
            for r in p2_rows:
                ws2.append(r)
            wb.save(path)
            messagebox.showinfo("Exportado", f"Excel guardado en:\n{path}")
    
    def _dialogo_agregar_programacion(self):

        from .services.ingenieria_service import listar_maquinas
        from .services.ingenieria_extras import (
            listar_componentes,
            crear_programacion
        )

        def _build(content):

            # -------------------------
            # MAQUINA
            maq_var, maq_combo = self._crear_selector_maquina(content)

            # -------------------------
            # TAREA
            tipo_entry = self._crear_input_tarea(content)

            sugerencias_box, _actualizar_sugerencias, _completar_con_sugerencia, _sugerencias_cache = self._crear_sugerencias(content, maq_var, tipo_entry)

            tipo_entry.bind("<KeyRelease>", _actualizar_sugerencias)
            tipo_entry.bind("<Tab>", _completar_con_sugerencia)

            # -------------------------
            # TIPO MANTENIMIENTO
            frm_tipo = tk.LabelFrame(content, text="Tipo mantenimiento")
            frm_tipo.pack(fill=tk.X, padx=10, pady=4)

            tipo_mant_var = tk.StringVar(value="PREVENTIVO")
            ttk.Radiobutton(frm_tipo, text="Preventivo", variable=tipo_mant_var, value="PREVENTIVO").pack(side=tk.LEFT)
            ttk.Radiobutton(frm_tipo, text="Correctivo", variable=tipo_mant_var, value="CORRECTIVO").pack(side=tk.LEFT)

            # -------------------------
            # PROGRAMACION
            sched_var, fecha_var, hor_entry = self._crear_programacion_ui(content)

            # -------------------------
            # COMPONENTE
            comp_var, comp_combo, actualizar_componentes = self._crear_componente_ui(content, maq_var)

            def _on_maquina_change(event=None):
                actualizar_componentes()
                _actualizar_sugerencias()

            maq_combo.bind("<<ComboboxSelected>>", _on_maquina_change)

            # -------------------------
            # GUARDAR
            # -------------------------

            def guardar():

                sel = maq_var.get()

                if " - " not in sel:

                    messagebox.showerror("Error", "Seleccione máquina")
                    return

                mid = int(sel.split(" - ")[0])

                tipo = tipo_entry.get().strip()

                fecha_prog = None
                hor_obj = None
                freq = None

                if sched_var.get() == "fecha" and fecha_var.get():

                    fecha_prog = datetime.fromisoformat(fecha_var.get())

                if sched_var.get() in ("horometro", "kilometraje"):

                    try:
                        hor_obj = float(hor_entry.get())
                        freq = hor_obj
                    except:
                        messagebox.showerror("Error", "Horómetro/Kilometraje inválido")
                        return

                comp_id = None

                comp_sel = comp_var.get()

                if " - " in comp_sel:
                    comp_id = int(comp_sel.split(" - ")[0])

                crear_programacion(
                    mid,
                    tipo,
                    tipo_mantenimiento=tipo_mant_var.get(),
                    fecha_programada=fecha_prog,
                    horometro_objetivo=hor_obj,
                    frecuencia_horas=freq,
                    componente_id=comp_id,
                )

                # Añadir el tipo recién creado a las sugerencias locales
                if tipo and tipo not in _sugerencias_cache:
                    _sugerencias_cache.append(tipo)

                messagebox.showinfo("OK", "Programación creada")

                self.view_mgr.pop()
                self.refrescar_programaciones()

            _icon_btn(content, "💾", "Guardar", guardar).pack(
                fill=tk.X,
                padx=10,
                pady=10
            )

        self.view_mgr.push("Agregar programación", _build, scrollable=True)
    
    def _on_prog_tree_double_click(self, event):
        """Abrir diálogo de edición al hacer doble click en una programación"""

        item = self.prog_tree.identify_row(event.y)
        if not item:
            return

        values = self.prog_tree.item(item, "values")
        if not values:
            return

        # El Treeview usa el ID del objeto como iid (se asigna en refrescar_programaciones)
        # así que no debemos usar la primera columna (#) para obtener el ID.
        try:
            prog_id = int(item)
        except (ValueError, TypeError):
            return

        self._dialogo_editar_programacion(prog_id)

    def _dialogo_editar_programacion(self, prog_id):
        """Abre un diálogo para editar una programación existente."""
        from .services.ingenieria_extras import listar_programaciones, listar_componentes, actualizar_programacion

        all_progs = listar_programaciones()
        prog = next((p for p in all_progs if p.id == prog_id), None)
        if not prog:
            messagebox.showerror("Error", "Programación no encontrada")
            return

        def _build(content):
            maq_map = {m.id: m for m in (self.maquinas or [])}
            maq = maq_map.get(prog.maquina_id)

            # ── Row number and machine selection ──
            # show the index ("#") for reference but not editable
            row_num = next((i for i,p in enumerate(all_progs,1) if p.id == prog_id), None)
            frm_maq = tk.LabelFrame(content, text="Máquina", padx=8, pady=4)
            frm_maq.pack(fill=tk.X, padx=10, pady=4)
            inner = tk.Frame(frm_maq)
            inner.pack(fill=tk.X)
            if row_num is not None:
                tk.Label(inner, text=f"# {row_num}", width=6, relief="sunken").pack(side=tk.LEFT, padx=(0,4))
            # combobox to allow machine change
            maq_var = tk.StringVar()
            try:
                maq_values = [f"{m.id} - {m.nombre}" for m in self.maquinas]
            except Exception:
                maq_values = []
            maq_combo = ttk.Combobox(inner, values=sorted(maq_values), textvariable=maq_var, width=40)
            maq_combo.pack(fill=tk.X, expand=True)
            # preselect current machine
            if maq:
                maq_var.set(f"{maq.id} - {maq.nombre}")
            # show horómetro actual for selected machine in label
            hor_actual_label = tk.Label(frm_maq, text=f"Horómetro actual: {maq.horometro_actual if maq else '--'}")
            hor_actual_label.pack(anchor="w", pady=(4,0))

            # ── Task description ──
            frm_tarea = tk.LabelFrame(content, text="Descripción de la tarea", padx=8, pady=4)
            frm_tarea.pack(fill=tk.X, padx=10, pady=4)
            tipo_entry = tk.Entry(frm_tarea, width=45)
            tipo_entry.pack(fill=tk.X)
            if prog.tipo:
                tipo_entry.insert(0, prog.tipo)

            # ── Maintenance type ──
            frm_tipo = tk.LabelFrame(content, text="Tipo de mantenimiento", padx=8, pady=4)
            frm_tipo.pack(fill=tk.X, padx=10, pady=4)
            tipo_mant_var = tk.StringVar(value=getattr(prog, 'tipo_mantenimiento', None) or "PREVENTIVO")
            ttk.Radiobutton(frm_tipo, text="PREVENTIVO", variable=tipo_mant_var, value="PREVENTIVO").pack(side=tk.LEFT, padx=5)
            ttk.Radiobutton(frm_tipo, text="CORRECTIVO", variable=tipo_mant_var, value="CORRECTIVO").pack(side=tk.LEFT, padx=5)

            # ── Scheduling method ──
            frm_sched = tk.LabelFrame(content, text="Programado por", padx=8, pady=4)
            frm_sched.pack(fill=tk.X, padx=10, pady=4)
            has_fecha = prog.fecha_programada is not None
            has_horo = (getattr(prog, 'horometro_objetivo', None) or 0) > 0
            sched_var = tk.StringVar(value="horometro" if has_horo and not has_fecha else "fecha")
            ttk.Radiobutton(frm_sched, text="Fecha", variable=sched_var, value="fecha").pack(side=tk.LEFT, padx=5)
            ttk.Radiobutton(frm_sched, text="Horómetro", variable=sched_var, value="horometro").pack(side=tk.LEFT, padx=5)
            ttk.Radiobutton(frm_sched, text="Kilometraje", variable=sched_var, value="kilometraje").pack(side=tk.LEFT, padx=5)

            sched_container = tk.Frame(content)
            sched_container.pack(fill=tk.X, padx=0, pady=0)

            fecha_frame = tk.LabelFrame(sched_container, text="📅 Selección de Fecha", padx=8, pady=4)
            fecha_var = tk.StringVar(value=prog.fecha_programada.date().isoformat() if prog.fecha_programada else "")
            def elegir_fecha():
                cal_win = tk.Toplevel(self.panel)
                cal_win.title("Elegir fecha")
                from tkcalendar import Calendar
                cal = Calendar(cal_win, selectmode="day")
                cal.pack(padx=10, pady=10)
                def _ok():
                    fecha_var.set(cal.selection_get().isoformat())
                    cal_win.destroy()
                btns = tk.Frame(cal_win)
                btns.pack(pady=5)
                _icon_btn(btns, "✔", "Aceptar", _ok).pack(side=tk.LEFT, padx=5)
                _icon_btn(btns, "❌", "Cancelar", cal_win.destroy).pack(side=tk.LEFT, padx=5)
            _icon_btn(fecha_frame, "📅", "Seleccionar fecha", elegir_fecha).pack(side=tk.LEFT, padx=4)
            tk.Label(fecha_frame, textvariable=fecha_var, relief="sunken", width=14).pack(side=tk.LEFT, padx=4)

            hor_frame = tk.LabelFrame(sched_container, text="🔧 Horómetro", padx=8, pady=4)
            hor_label = tk.Label(hor_frame, text="Valor:")
            hor_label.pack(side=tk.LEFT, padx=2)
            hor_obj_entry = tk.Entry(hor_frame, width=10)
            hor_obj_entry.pack(side=tk.LEFT, padx=4)
            if has_horo:
                hor_obj_entry.insert(0, str(prog.horometro_objetivo))

            fecha_frame.grid(row=0, column=0, sticky="ew")
            hor_frame.grid(row=0, column=0, sticky="ew")

            def _toggle_sched(*args):
                if sched_var.get() == "fecha":
                    fecha_frame.tkraise()
                    hor_label.config(text="Horómetro:")
                else:
                    hor_frame.tkraise()
                    if sched_var.get() == "kilometraje":
                        hor_frame.config(text="🚗 Kilometraje")
                        hor_label.config(text="Kilometraje:")
                    else:
                        hor_frame.config(text="🔧 Horómetro")
                        hor_label.config(text="Valor:")

            sched_var.trace_add('write', _toggle_sched)
            _toggle_sched()
            # -------------------------
            # Estado del repuesto y cantidad/litraje
            # -------------------------
            frm_estado = tk.LabelFrame(content, text="Estado del repuesto", padx=8, pady=4)
            frm_estado.pack(fill=tk.X, padx=10, pady=4)
            estado_var = tk.StringVar(value=getattr(prog, 'estado_repuesto', 'SOLIDO'))
            # El estado depende del tipo de material del componente y no debe cambiarse manualmente.
            ttk.Radiobutton(frm_estado, text="Fluido", variable=estado_var, value="FLUIDO", state="disabled").pack(side=tk.LEFT, padx=5)
            ttk.Radiobutton(frm_estado, text="Sólido", variable=estado_var, value="SOLIDO", state="disabled").pack(side=tk.LEFT, padx=5)
            ttk.Radiobutton(frm_estado, text="Gaseoso", variable=estado_var, value="GASEOSO", state="disabled").pack(side=tk.LEFT, padx=5)

            cant_container = tk.Frame(content)
            cant_container.pack(fill=tk.X, padx=0, pady=0)
            litraje_frame = tk.LabelFrame(cant_container, text="💧 Litraje", padx=8, pady=4)
            litraje_label = tk.Label(litraje_frame, text="Litros:")
            litraje_label.pack(side=tk.LEFT, padx=2)
            litraje_entry = tk.Entry(litraje_frame, width=10)
            litraje_entry.pack(side=tk.LEFT, padx=4)
            cantidad_frame = tk.LabelFrame(cant_container, text="📦 Cantidad", padx=8, pady=4)
            tk.Label(cantidad_frame, text="Cantidad:").pack(side=tk.LEFT, padx=2)
            cantidad_entry = tk.Entry(cantidad_frame, width=10)
            cantidad_entry.pack(side=tk.LEFT, padx=4)

            # pre-fill cantidad
            cant_val = getattr(prog, 'cantidad', None)
            if cant_val is not None:
                if estado_var.get() == "FLUIDO":
                    litraje_entry.insert(0, f"{cant_val:g}")
                    litraje_frame.pack(fill=tk.X, padx=10, pady=2)
                else:
                    cantidad_entry.insert(0, f"{cant_val:g}")
                    cantidad_frame.pack(fill=tk.X, padx=10, pady=2)

            def _toggle_estado(*args):
                litraje_frame.pack_forget()
                cantidad_frame.pack_forget()

                if estado_var.get() == "GASEOSO":
        # Cambiar título y etiqueta
                    litraje_frame.config(text="🌬️ Presión")
                    litraje_label.config(text="Presión (PSI):")
                    litraje_frame.pack(fill=tk.X, padx=10, pady=2)

                elif estado_var.get() == "FLUIDO":
        # Restaurar título original
                    litraje_frame.config(text="💧 Litraje")
                    litraje_label.config(text="Litros:")
                    litraje_frame.pack(fill=tk.X, padx=10, pady=2)

                else:
                    cantidad_frame.pack(fill=tk.X, padx=10, pady=2)  

            estado_var.trace_add('write', _toggle_estado)

            # -------------------------
            # Componente asociado
            # -------------------------
            frm_comp_asoc = tk.LabelFrame(content, text="Componente Asociado", padx=8, pady=4)
            frm_comp_asoc.pack(fill=tk.X, padx=10, pady=4)
            comp_asoc_var = tk.StringVar()
            comp_asoc_combo = ttk.Combobox(frm_comp_asoc, values=[], textvariable=comp_asoc_var, width=40)
            comp_asoc_combo.pack(fill=tk.X)
            comps = listar_componentes(prog.maquina_id) or []
            comp_values = []
            for c in comps:
                mat = getattr(c, 'material', None)
                if mat:
                    comp_values.append(f"{c.id} - {mat.nombre}")
                else:
                    comp_values.append(f"{c.id} - {getattr(c, 'nombre', 'Componente')}")
            comp_asoc_combo['values'] = comp_values
            if getattr(prog, 'componente', None):
                comp = prog.componente
                mat = getattr(comp, 'material', None)
                if mat:
                    comp_asoc_var.set(f"{comp.id} - {mat.nombre}")
                else:
                    comp_asoc_var.set(f"{comp.id} - {getattr(comp, 'nombre', 'Componente')}")

            def _ajustar_estado_por_tipo_material(comp_id):
                try:
                    comp = next((c for c in comps if c.id == comp_id), None)
                    if not comp or not getattr(comp, 'material', None):
                        return
                    tipo = (getattr(comp.material, 'tipo', '') or '').strip().lower()
                    if 'gase' in tipo:
                        estado_var.set('GASEOSO')
                    elif 'flu' in tipo or 'líqu' in tipo:
                        estado_var.set('FLUIDO')
                    else:
                        estado_var.set('SOLIDO')
                except Exception:
                    pass

            def _on_comp_asoc_change(event=None):
                sel = comp_asoc_var.get()
                if sel and ' - ' in sel:
                    try:
                        comp_id = int(sel.split(' - ')[0])
                        _ajustar_estado_por_tipo_material(comp_id)
                    except Exception:
                        pass

            comp_asoc_combo.bind('<<ComboboxSelected>>', _on_comp_asoc_change)
            # Ajuste inicial (por si el comp ya está seleccionado)
            _on_comp_asoc_change()

            # -------------------------
            # Notas
            # -------------------------
            tk.Label(content, text="Notas:").pack(anchor="w", padx=10, pady=(8,0))
            desc_entry = tk.Entry(content, width=40)
            desc_entry.pack(fill=tk.X, padx=10, pady=2)
            if prog.descripcion:
                desc_entry.insert(0, prog.descripcion)

            # -------------------------
            # Orden (posición) editable
            # -------------------------
            frm_orden = tk.LabelFrame(content, text="Orden de presentación", padx=8, pady=4)
            frm_orden.pack(fill=tk.X, padx=10, pady=4)
            orden_var = tk.IntVar(value=getattr(prog, 'orden', prog.id))
            orden_spin = tk.Spinbox(frm_orden, from_=1, to=len(all_progs), textvariable=orden_var, width=5)
            orden_spin.pack(side=tk.LEFT, padx=5)
            tk.Label(frm_orden, text="(arriba/abajo en lista)").pack(side=tk.LEFT, padx=5)

            # -------------------------
            # Guardar cambios
            # -------------------------
            def save_edit():
                # máquina
                mid = int(maq_var.get().split(" - ")[0])
                # cantidad/litraje
                cantidad = None
                if estado_var.get() == "FLUIDO":
                    try:
                        cantidad = float(litraje_entry.get())
                    except:
                        cantidad = None
                else:
                    try:
                        cantidad = float(cantidad_entry.get())
                    except:
                        cantidad = None
                # fecha/horómetro
                fecha_prog = None
                hor_obj = None
                freq = None
                if sched_var.get() == "fecha" and fecha_var.get():
                    from datetime import datetime
                    fecha_prog = datetime.fromisoformat(fecha_var.get())
                elif sched_var.get() in ("horometro", "kilometraje"):
                    try:
                        hor_obj = float(hor_obj_entry.get())
                    except:
                        hor_obj = None
                    if hor_obj:
                        freq = hor_obj
                # componente
                comp_id = None
                sel = comp_asoc_var.get()
                if sel and " - " in sel:
                    try:
                        comp_id = int(sel.split(" - ")[0])
                    except:
                        comp_id = None
                # guardar cambios
                actualizar_programacion(
                    prog_id,
                    maquina_id=mid,
                    tipo=tipo_entry.get().strip(),
                    tipo_mantenimiento=tipo_mant_var.get(),
                    fecha_programada=fecha_prog,
                    horometro_objetivo=hor_obj,
                    frecuencia_horas=freq,
                    descripcion=desc_entry.get().strip() or None,
                    componente_id=comp_id,
                    cantidad=cantidad,
                    estado_repuesto=estado_var.get(),
                    orden=orden_var.get()
                )
                messagebox.showinfo("Actualizado", "Programación actualizada con éxito")
                self.view_mgr.pop()
                self.refrescar_programaciones()

            _icon_btn(content, "💾", "Guardar cambios", save_edit, font=("Arial", 10, "bold")).pack(fill=tk.X, padx=10, pady=8)

        self.view_mgr.push("Editar programación", _build, scrollable=True)

    def _eliminar_programacion(self):
        """Elimina la(s) programación(es) seleccionada(s) del Treeview."""
        sel = self.prog_tree.selection()
        if not sel:
            messagebox.showwarning("Aviso", "Seleccione una o más programaciones")
            return
        from .services.ingenieria_extras import eliminar_programacion
        for iid in sel:
            try:
                pid = int(iid)
                eliminar_programacion(pid)
            except Exception as e:
                messagebox.showerror("Error", str(e))
                return
        messagebox.showinfo("Eliminado", f"{len(sel)} programación(es) borrada(s)")
        self.refrescar_programaciones()

    def _reset_programaciones(self):
        """Elimina todas las programaciones para poder agregar desde cero."""
        if not messagebox.askyesno("Confirmar", "¿Desea eliminar todas las programaciones?"): 
            return
        from .services.ingenieria_extras import eliminar_todas_programaciones
        try:
            eliminar_todas_programaciones()
            messagebox.showinfo("Reseteado", "Todas las programaciones han sido eliminadas.")
            self.refrescar_programaciones()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _toggle_autorizacion_programacion(self):
        """Alterna la autorización de las programaciones seleccionadas."""
        sel = self.prog_tree.selection()
        if not sel:
            messagebox.showwarning("Aviso", "Seleccione una o más programaciones")
            return
        from .services.ingenieria_extras import actualizar_programacion
        # Determinar posición real de la columna 'Autorizado' para no depender del orden.
        try:
            auth_idx = list(self.prog_tree["columns"]).index("_aut")
        except ValueError:
            auth_idx = None

        for iid in sel:
            try:
                pid = int(iid)
                vals = self.prog_tree.item(iid, "values")
                current = "No"
                if auth_idx is not None and auth_idx < len(vals):
                    current = vals[auth_idx]
                nuevo = 0 if str(current).strip().lower().startswith("s") else 1
                actualizar_programacion(pid, autorizado=nuevo)
            except Exception as e:
                messagebox.showerror("Error", str(e))
                return
        self.refrescar_programaciones()

    def _abrir_panel_prediccion(self):
        """Abre el panel avanzado de predicción de fallas."""
        from .PanelPrediccionFallas import abrir_panel_prediccion_fallas
        abrir_panel_prediccion_fallas(self.panel, self.maquinas)
    
    def agregar_maquina(self):
        """Vista embebida con scrollbar para agregar máquinas."""
    
        def _build(scrollable_frame):
            main_frame = create_frame(scrollable_frame)
            main_frame.pack(fill=tk.BOTH, expand=True)

            left = create_frame(main_frame)
            left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))

            right = create_frame(main_frame)
            right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0))

            create_label(left, text="Identificación Única", font=("Arial", 10, "bold")).pack(anchor="w", padx=10, pady=(10,0))

            create_label(left, text="Marca / Fabricante", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            fabricante_entry = create_entry(left, width=50)
            fabricante_entry.pack(fill=tk.X, padx=10, pady=2)

            create_label(left, text="Modelo", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            modelo_entry = create_entry(left, width=50)
            modelo_entry.pack(fill=tk.X, padx=10, pady=2)

            create_label(left, text="Número de Serie / PIN", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            serie_entry = create_entry(left, width=50)
            serie_entry.pack(fill=tk.X, padx=10, pady=2)

            acciones_frame = create_frame(left)
            acciones_frame.pack(fill=tk.X, padx=10, pady=4)
            analizar_btn = create_button(acciones_frame, text="Analizar")
            analizar_btn.pack(side=tk.LEFT)

            analisis_label = create_label(left, text="", fg="#096", justify="left", wraplength=450)
            analisis_label.pack(anchor="w", padx=10, pady=(2, 4))





            def extraer_componentes(texto):
                componentes = []

                patrones = re.findall(r"(\d+)\s+([a-zA-Záéíóúñ]+)", texto.lower())

                for cantidad, nombre in patrones:
                    componentes.append({
                    "componente": nombre,
                    "cantidad": int(cantidad)
                })

                return componentes
            

            def extraer_serie(texto):
                patrones = [
                    r"\b[A-Z]{2,}\d{2,}\b",        # ABC123
                    r"\b[A-Z]+-\d+\b",             # XR-200
                    r"\b\d{3,}-[A-Z]+\b",          # 123-ABC
                    r"\b[A-Z0-9]{5,}\b"            # general
                ]

                for patron in patrones:
                    match = re.search(patron, texto.upper())
                    if match:
                        return match.group()

                return ""
            


            try:
                import spacy
                try:
                    nlp = spacy.load("es_core_news_sm")
                except Exception:
                    nlp = None
            except ImportError:
                nlp = None

            def analizar_avanzado(texto):
                if nlp is None:
                    numeros = re.findall(r"\d+", texto)
                    posibles_modelos = [
                        token for token in re.findall(r"[A-Za-z0-9\-]+", texto)
                        if any(char.isdigit() for char in token) and len(token) >= 3
                    ]
                    return {"numeros": numeros, "posibles_modelos": posibles_modelos}

                doc = nlp(texto)

                entidades = {
                    "numeros": [],
                    "posibles_modelos": [],
                }

                for token in doc:
                    if token.like_num:
                        entidades["numeros"].append(token.text)

                for ent in doc.ents:
                    entidades["posibles_modelos"].append(ent.text)

                return entidades
            

            def analizar_completo(texto):
                resultado = _analizar_texto(texto)

                resultado["serie"] = extraer_serie(texto)
                resultado["componentes"] = extraer_componentes(texto)
                avanzado = analizar_avanzado(texto)
                resultado["numeros"] = avanzado.get("numeros", [])
                resultado["posibles_modelos"] = avanzado.get("posibles_modelos", [])
                resultado["analisis_avanzado"] = avanzado

                if not resultado.get("modelo") and avanzado.get("posibles_modelos"):
                    resultado["modelo"] = avanzado["posibles_modelos"][0]

                return resultado


            def _analizar_texto(texto):
                resultado = {"fabricante": "", "modelo": "", "tipo": "", "combustible": "", "nombre": ""}
                try:
                    from .services.ingenieria_service import analizar_maquina
                    return analizar_maquina(texto)
                except Exception:
                    texto_l = texto.lower()
                    if "gasolina" in texto_l or "nafta" in texto_l:
                        resultado["combustible"] = "Gasolina"
                    elif "diesel" in texto_l or "diésel" in texto_l:
                        resultado["combustible"] = "Diésel"
                    elif "eléctrico" in texto_l or "electrico" in texto_l or "eléctrica" in texto_l:
                        resultado["combustible"] = "Eléctrico"

                    if "cort" in texto_l:
                        resultado["tipo"] = "Cortacésped"
                    elif "tractor" in texto_l:
                        resultado["tipo"] = "Tractor"
                    elif "moto" in texto_l and "cult" in texto_l:
                        resultado["tipo"] = "Motocultor"

                    palabras = [p.strip() for p in re.split(r"[\s,;|\\/]+", texto) if p.strip()]
                    if palabras:
                        resultado["nombre"] = palabras[0]
                    if len(palabras) > 1:
                        resultado["fabricante"] = palabras[0]
                        resultado["modelo"] = palabras[1]
                    return resultado

            def analizar():
                texto = " ".join(
                    part for part in (
                        fabricante_entry.get().strip(),
                        modelo_entry.get().strip(),
                        serie_entry.get().strip(),
                        tipo_entry.get().strip(),
                        nombre_entry.get().strip(),
                    ) if part
                )
                if not texto:
                    messagebox.showwarning("Aviso", "Ingrese al menos Marca, Modelo o Serie para analizar")
                    return

                try:
                    resultado = analizar_completo(texto)
                except Exception:
                    resultado = _analizar_texto(texto)

                if resultado.get("fabricante"):
                    fabricante_entry.delete(0, tk.END)
                    fabricante_entry.insert(0, resultado.get("fabricante", ""))
                if resultado.get("modelo"):
                    modelo_entry.delete(0, tk.END)
                    modelo_entry.insert(0, resultado.get("modelo", ""))
                if resultado.get("combustible"):
                    comb_entry.delete(0, tk.END)
                    comb_entry.insert(0, resultado.get("combustible", ""))
                if resultado.get("tipo"):
                    tipo_entry.delete(0, tk.END)
                    tipo_entry.insert(0, resultado.get("tipo", ""))
                if resultado.get("serie"):
                    serie_entry.delete(0, tk.END)
                    serie_entry.insert(0, resultado.get("serie", ""))
                if not nombre_entry.get().strip() and resultado.get("nombre"):
                    nombre_entry.delete(0, tk.END)
                    nombre_entry.insert(0, resultado.get("nombre", ""))

                componentes = resultado.get("componentes") or []
                if componentes:
                    detalles = ", ".join(f"{c['cantidad']}x {c['componente']}" for c in componentes)
                    comb_suggestion_label.config(text=f"Componentes detectados: {detalles}", fg="#096")
                else:
                    comb_suggestion_label.config(text="No se detectaron componentes desde el análisis.", fg="#444")

                resumen = ["Análisis completo ejecutado correctamente."]
                if resultado.get("posibles_modelos"):
                    resumen.append(f"Modelos posibles: {', '.join(resultado['posibles_modelos'])}")
                elif resultado.get("numeros"):
                    resumen.append(f"Números detectados: {', '.join(resultado['numeros'])}")
                analisis_label.config(text=" | ".join(resumen), fg="#096")

            analizar_btn.config(command=analizar)

            create_label(right, text="Categoría", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(10,0))
            nombre_entry = create_entry(right, width=50)
            nombre_entry.pack(fill=tk.X, padx=10, pady=2)

            create_label(right, text="Tipo de unidad", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            tipo_entry = create_entry(right, width=50)
            tipo_entry.pack(fill=tk.X, padx=10, pady=2)

            create_label(right, text="Año", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            ano_entry = create_entry(right, width=15)
            ano_entry.pack(anchor="w", padx=10, pady=2)

            # Motor: Marca, Modelo, Serie en una sola fila
            frm_motor = create_label_frame(right, text="Motor", padx=6, pady=4)
            frm_motor.pack(fill=tk.X, padx=10, pady=(8, 2))
            for col in range(6):
                frm_motor.columnconfigure(col, weight=1 if col % 2 == 1 else 0)
            create_label(frm_motor, text="Marca:", font=("Arial", 9)).grid(row=0, column=0, sticky="w", padx=2)
            motor_marca_entry = create_entry(frm_motor, width=14)
            motor_marca_entry.grid(row=0, column=1, sticky="ew", padx=2)
            create_label(frm_motor, text="Modelo:", font=("Arial", 9)).grid(row=0, column=2, sticky="w", padx=2)
            motor_modelo_entry = create_entry(frm_motor, width=14)
            motor_modelo_entry.grid(row=0, column=3, sticky="ew", padx=2)
            create_label(frm_motor, text="Serie:", font=("Arial", 9)).grid(row=0, column=4, sticky="w", padx=2)
            motor_serie_entry = create_entry(frm_motor, width=14)
            motor_serie_entry.grid(row=0, column=5, sticky="ew", padx=2)

            create_label(right, text="Tipo de Combustible", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            comb_frame = create_frame(right)
            comb_frame.pack(fill=tk.X, padx=10, pady=2)
            comb_entry = create_entry(comb_frame, width=22)
            comb_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
            detectar_comb_btn = create_button(comb_frame, text="Detectar combustible")
            detectar_comb_btn.pack(side=tk.LEFT, padx=(6, 0))
            buscar_web_btn = create_button(comb_frame, text="Buscar en internet")
            buscar_web_btn.pack(side=tk.LEFT, padx=(6, 0))
            comb_suggestion_label = create_label(right, text="", fg="#096", justify="left", wraplength=450)
            comb_suggestion_label.pack(anchor="w", padx=10, pady=(2, 4))

            modo_frame = create_frame(right)
            modo_frame.pack(fill=tk.X, padx=10, pady=(2, 4))
            modo_var = tk.StringVar(value="google")
            create_label(modo_frame, text="Fuente de características:", font=("Arial", 9)).pack(side=tk.LEFT)
            create_radiobutton(modo_frame, text="Google", variable=modo_var, value="google").pack(side=tk.LEFT, padx=(8, 0))
            create_radiobutton(modo_frame, text="Manual", variable=modo_var, value="manual").pack(side=tk.LEFT, padx=(8, 0))

            info_frame = create_label_frame(right, text="Información técnica dinámica", padx=6, pady=4)
            info_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(4, 8))
            info_content = create_frame(info_frame)
            info_content.pack(fill=tk.BOTH, expand=True)

            manual_frame = create_frame(info_frame)
            create_label(manual_frame, text="Ingrese características manuales:", font=("Arial", 9)).pack(anchor="w", padx=4, pady=(2, 2))
            manual_text = create_text(manual_frame, height=6, width=50)
            manual_text.pack(fill=tk.BOTH, expand=True, padx=4, pady=(0, 4))

            def _clear_info_content():
                for widget in info_content.winfo_children():
                    widget.destroy()

            def _add_info_row(label_text, value_text):
                row = create_frame(info_content)
                row.pack(fill=tk.X, pady=1, anchor="w")
                create_label(row, text=f"{label_text}:", font=("Arial", 9, "bold"), width=18, anchor="w").pack(side=tk.LEFT)
                create_label(row, text=value_text, font=("Arial", 9), wraplength=380, justify="left").pack(side=tk.LEFT, fill=tk.X, expand=True)

            def _update_info_view(*args):
                if modo_var.get() == "manual":
                    info_content.pack_forget()
                    manual_frame.pack(fill=tk.BOTH, expand=True)
                else:
                    manual_frame.pack_forget()
                    info_content.pack(fill=tk.BOTH, expand=True)

            modo_var.trace_add("write", _update_info_view)

            def render_dynamic_info(info):
                _clear_info_content()
                if not info:
                    create_label(info_content, text="No hay datos técnicos disponibles. Use Buscar en internet.", fg="#b00", wraplength=420, justify="left").pack(anchor="w")
                    return

                details = []
                if info.get("nombre"):
                    details.append(("Ficha", info["nombre"]))
                if info.get("categoria"):
                    details.append(("Categoría", info["categoria"]))
                if info.get("tipo_combustible"):
                    details.append(("Combustible", info["tipo_combustible"]))
                if info.get("capacidad_combustible"):
                    details.append(("Capacidad combustible", info["capacidad_combustible"]))
                if info.get("tipo_aceite"):
                    details.append(("Tipo de aceite", info["tipo_aceite"]))
                if info.get("cantidad_aceite"):
                    details.append(("Volumen aceite", info["cantidad_aceite"]))
                if info.get("estado"):
                    details.append(("Estado sugerido", info["estado"]))
                if info.get("observaciones"):
                    details.append(("Observaciones", info["observaciones"]))
                if info.get("internet_abstract"):
                    details.append(("Resumen web", info["internet_abstract"]))
                if info.get("internet_related"):
                    details.append(("Temas relacionados", "; ".join(info["internet_related"])))

                if not details:
                    create_label(info_content, text="No se encontraron detalles técnicos en la búsqueda.", fg="#b00", wraplength=420, justify="left").pack(anchor="w")
                    return

                for label_text, value_text in details:
                    _add_info_row(label_text, value_text)

            render_dynamic_info({})
            _update_info_view()

            create_label(right, text="Altura de corte (mm)", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            altura_entry = create_entry(right, width=15)
            altura_entry.pack(anchor="w", padx=10, pady=2)

            create_label(left, text="Estado", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            estado_entry = create_entry(left, width=50)
            estado_entry.pack(fill=tk.X, padx=10, pady=2)

            create_label(left, text="Ubicación", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            ubicacion_entry = create_entry(left, width=50)
            ubicacion_entry.pack(fill=tk.X, padx=10, pady=2)

            create_label(left, text="Observaciones (opcional)", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            obs_entry = create_entry(left, width=50)
            obs_entry.pack(fill=tk.X, padx=10, pady=2)

            # Fecha de entrada en operación (calendario)
            frm_fecha_op = create_label_frame(left, text="📅 Fecha de entrada en operación", padx=8, pady=4)
            frm_fecha_op.pack(fill=tk.X, padx=10, pady=(8, 2))
            fecha_op_var = tk.StringVar()
            create_label(
                frm_fecha_op,
                textvariable=fecha_op_var,
                width=14,
                relief="sunken",
                bg=theme.current.get("entry_bg", "white"),
                fg=theme.current.get("entry_fg", "black"),
            ).pack(side=tk.LEFT, padx=4)
            def elegir_fecha_op():
                cal_win = CTkToplevel(self.panel)
                cal_win.title("Fecha de entrada en operación")
                from tkcalendar import Calendar
                cal = Calendar(cal_win, selectmode="day")
                cal.pack(padx=10, pady=10)
                def _ok():
                    fecha_op_var.set(cal.selection_get().isoformat())
                    cal_win.destroy()
                btns = create_frame(cal_win)
                btns.pack(pady=5)
                _icon_btn(btns, "✔", "Aceptar", _ok).pack(side=tk.LEFT, padx=5)
                _icon_btn(btns, "❌", "Cancelar", cal_win.destroy).pack(side=tk.LEFT, padx=5)
            _icon_btn(frm_fecha_op, "📅", "Seleccionar fecha", elegir_fecha_op).pack(side=tk.LEFT, padx=4)
    
            create_label(left, text="Horómetro inicial", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            horometro_entry = create_entry(left, width=15)
            horometro_entry.pack(anchor="w", padx=10, pady=2)
    
            material_frame = create_label_frame(right, text="Materiales Asociados", padx=6, pady=4)
            material_frame.pack(fill=tk.X, padx=10, pady=(8, 2))
            from .services.ingenieria_extras import listar_materiales
            mat_names = [m.nombre for m in listar_materiales()]
            material_var = tk.StringVar()
            material_combo = ttk.Combobox(material_frame, values=mat_names, textvariable=material_var)
            def _validate_mat(event=None):
                val = material_var.get().strip()
                if val and val not in mat_names:
                    material_var.set("")
            material_combo.bind("<FocusOut>", _validate_mat)
            material_combo.pack(fill=tk.X, padx=10, pady=2)

            def detectar_combustible():
                try:
                    from .services.ingenieria_service import inferir_tipo_combustible
                    sugerido = inferir_tipo_combustible(
                        fabricante=fabricante_entry.get().strip(),
                        modelo=modelo_entry.get().strip(),
                        tipo_unidad=tipo_entry.get().strip(),
                        categoria=nombre_entry.get().strip(),
                        motor_marca=motor_marca_entry.get().strip(),
                        motor_modelo=motor_modelo_entry.get().strip(),
                        descripcion=obs_entry.get().strip(),
                    )
                except Exception:
                    sugerido = None

                if sugerido:
                    comb_entry.delete(0, tk.END)
                    comb_entry.insert(0, sugerido)
                    comb_suggestion_label.config(text=f"Sugerido: {sugerido}", fg="#096")
                else:
                    comb_suggestion_label.config(
                        text="No se pudo inferir el tipo de combustible con los datos actuales.",
                        fg="#b00"
                    )

            detectar_comb_btn.config(command=detectar_combustible)

            def buscar_en_internet():
                try:
                    from .services.ingenieria_service import buscar_info_maquina_online, guardar_en_base
                    info = buscar_info_maquina_online(
                        nombre=nombre_entry.get().strip(),
                        datos_clave=serie_entry.get().strip(),
                        fabricante=fabricante_entry.get().strip(),
                        modelo=modelo_entry.get().strip(),
                        tipo_unidad=tipo_entry.get().strip(),
                        categoria=nombre_entry.get().strip(),
                    )
                    try:
                        if info:
                            guardar_en_base(info)
                    except Exception:
                        pass
                except Exception as e:
                    comb_suggestion_label.config(text=f"Error buscando internet: {e}", fg="#b00")
                    return

                if info.get("tipo_combustible"):
                    comb_entry.delete(0, tk.END)
                    comb_entry.insert(0, info["tipo_combustible"])

                if info.get("categoria") and not nombre_entry.get().strip():
                    nombre_entry.delete(0, tk.END)
                    nombre_entry.insert(0, info["categoria"])
                if info.get("estado") and not estado_entry.get().strip():
                    estado_entry.delete(0, tk.END)
                    estado_entry.insert(0, info["estado"])

                detalles = []
                if info.get("capacidad_combustible"):
                    detalles.append(f"Combustible: {info['capacidad_combustible']}")
                if info.get("tipo_aceite"):
                    detalles.append(f"Aceite: {info['tipo_aceite']}")
                if info.get("cantidad_aceite"):
                    detalles.append(f"Volumen aceite: {info['cantidad_aceite']}")
                if info.get("internet_abstract"):
                    detalles.append(info["internet_abstract"])
                if detalles:
                    comb_suggestion_label.config(text=" | ".join(detalles), fg="#096")
                else:
                    comb_suggestion_label.config(text="No se encontraron datos técnicos adicionales en internet.", fg="#b00")

                render_dynamic_info(info)

            buscar_web_btn.config(command=buscar_en_internet)

            # Botón para cargar datos de ejemplo
            def cargar_ejemplo():
                nombre_entry.delete(0, tk.END); nombre_entry.insert(0, "REELMASTER 5510")
                fabricante_entry.delete(0, tk.END); fabricante_entry.insert(0, "Toro")
                ano_entry.delete(0, tk.END); ano_entry.insert(0, "2020")
                serie_entry.delete(0, tk.END); serie_entry.insert(0, "SN1234")
                motor_marca_entry.delete(0, tk.END); motor_marca_entry.insert(0, "Honda")
                motor_modelo_entry.delete(0, tk.END); motor_modelo_entry.insert(0, "GXV160")
                motor_serie_entry.delete(0, tk.END); motor_serie_entry.insert(0, "MOTOR001")
                comb_entry.delete(0, tk.END); comb_entry.insert(0, "Gasolina")
                altura_entry.delete(0, tk.END); altura_entry.insert(0, "25")
                estado_entry.delete(0, tk.END); estado_entry.insert(0, "Operativa")
                ubicacion_entry.delete(0, tk.END); ubicacion_entry.insert(0, "Campo")
                horometro_entry.delete(0, tk.END); horometro_entry.insert(0, "0")
                fecha_op_var.set(datetime.now().date().isoformat())
    
            _icon_btn(scrollable_frame, "📋", "Usar datos de ejemplo", cargar_ejemplo).pack(fill=tk.X, padx=10, pady=4)
    
            def guardar():
                try:
                    nombre = nombre_entry.get().strip()
                    if not nombre:
                        raise ValueError("El nombre es obligatorio")
    
                    fabricante = fabricante_entry.get().strip() or None
                    modelo = modelo_entry.get().strip() or None
                    tipo_unidad = tipo_entry.get().strip() or None
                    try:
                        año_text = ano_entry.get().strip()
                        año_val = int(año_text) if año_text else None
                    except ValueError:
                        raise ValueError("El año debe ser un número válido")
                    
                    codigo_serie = serie_entry.get().strip() or None
                    motor_marca = motor_marca_entry.get().strip() or None
                    motor_modelo = motor_modelo_entry.get().strip() or None
                    motor_serie = motor_serie_entry.get().strip() or None
                    tipo_comb = comb_entry.get().strip() or None
                    categoria = nombre_entry.get().strip() or None
                    estado = estado_entry.get().strip() or None
                    ubicacion = ubicacion_entry.get().strip() or None
                    if modo_var.get() == "manual":
                        manual_observ = manual_text.get("1.0", "end").strip()
                        observ = manual_observ or obs_entry.get().strip() or None
                    else:
                        observ = obs_entry.get().strip() or None

                    mat_name = material_var.get().strip()
                    mat_id = None
                    if mat_name:
                        mats = listar_materiales()
                        found = next((x for x in mats if x.nombre == mat_name), None)
                        if not found:
                            messagebox.showwarning("Aviso", "El material no existe. Regístrelo en Gestionar Materiales.")
                            return
                        mat_id = found.id
    
                    try:
                        horometro = float(horometro_entry.get().strip() or 0)
                    except ValueError:
                        raise ValueError("El horómetro debe ser numérico")
                    
                    try:
                        altura_val = int(altura_entry.get().strip() or 0)
                    except Exception:
                        altura_val = None
                    fecha_op_text = fecha_op_var.get().strip()
                    fecha_op_val = None
                    if fecha_op_text:
                        fecha_op_val = datetime.fromisoformat(fecha_op_text)
    
                    nueva = crear_maquina(
                        nombre,
                        horometro,
                        None,
                        categoria=categoria,
                        estado=estado,
                        ubicacion=ubicacion,
                        fabricante=fabricante,
                        modelo=modelo,
                        tipo_unidad=tipo_unidad,
                        año=año_val,
                        codigo_serie=codigo_serie,
                        motor_marca=motor_marca,
                        motor_modelo=motor_modelo,
                        motor_serie=motor_serie,
                        tipo_combustible=tipo_comb,
                        altura_corte_mm=altura_val,
                        material_id=mat_id,
                        observaciones=observ,
                        fecha_operacion=fecha_op_val,
                    )
                    nid = nueva.id
                    self.undo.push(f"Crear máquina '{nombre}'", lambda _id=nid: (eliminar_maquina(_id), self.refrescar_lista()))
                    messagebox.showinfo("Éxito", "Máquina creada")
                    self.view_mgr.pop()
                    self.refrescar_lista()
                except Exception as e:
                    messagebox.showerror("Error", str(e))
    
            _icon_btn(scrollable_frame, "✅", "Guardar máquina", guardar, font=("Arial", 10, "bold")).pack(fill=tk.X, padx=10, pady=(10, 20))
    
        self.view_mgr.push("Agregar Máquina", _build, scrollable=True)
    
    def editar_maquina(self):
        sel = self.listbox.curselection()
        if not sel:
            messagebox.showwarning("Aviso", "Seleccione una máquina")
            return
        maquina = self.maquinas[sel[0]]
    
        def _build(scrollable_frame):
            tk.Label(scrollable_frame, text="Número de la Unidad (ID)", font=("Arial", 10)).pack(anchor="w", padx=10, pady=(10,0))
            id_entry = tk.Entry(scrollable_frame, width=50)
            id_entry.insert(0, str(maquina.id))
            id_entry.pack(padx=10, pady=2)
        
            tk.Label(scrollable_frame, text="Nombre *", font=("Arial", 10, "bold")).pack(anchor="w", padx=10, pady=(8,0))
            nombre_entry = tk.Entry(scrollable_frame, width=50)
            nombre_entry.insert(0, maquina.nombre)
            nombre_entry.pack(padx=10, pady=2)
        
            tk.Label(scrollable_frame, text="Tipo de unidad", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            tipo_entry = tk.Entry(scrollable_frame, width=50)
            tipo_entry.insert(0, getattr(maquina, 'tipo_unidad', '') or '')
            tipo_entry.pack(padx=10, pady=2)
        
            # Campos principales
            tk.Label(scrollable_frame, text="Fabricante (marca)", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            fabricante_entry = tk.Entry(scrollable_frame, width=50)
            fabricante_entry.insert(0, getattr(maquina, 'fabricante', '') or '')
            fabricante_entry.pack(padx=10, pady=2)
        
            tk.Label(scrollable_frame, text="Modelo", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            modelo_entry = tk.Entry(scrollable_frame, width=50)
            modelo_entry.insert(0, getattr(maquina, 'modelo', '') or '')
            modelo_entry.pack(padx=10, pady=2)
        
            tk.Label(scrollable_frame, text="Año", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            ano_entry = tk.Entry(scrollable_frame, width=50)
            ano_entry.insert(0, getattr(maquina, 'año', '') or '')
            ano_entry.pack(padx=10, pady=2)
        
            tk.Label(scrollable_frame, text="Categoría", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            categoria_entry = tk.Entry(scrollable_frame, width=50)
            categoria_entry.insert(0, getattr(maquina, 'categoria', '') or '')
            categoria_entry.pack(padx=10, pady=2)
        
            tk.Label(scrollable_frame, text="Estado", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            estado_entry = tk.Entry(scrollable_frame, width=50)
            estado_entry.insert(0, getattr(maquina, 'estado', '') or '')
            estado_entry.pack(padx=10, pady=2)
        
            # material asociado
            tk.Label(scrollable_frame, text="Material asociado", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            from .services.ingenieria_extras import listar_materiales
            mat_names = [m.nombre for m in listar_materiales()]
            material_var = tk.StringVar()
            material_combo = ttk.Combobox(scrollable_frame, values=mat_names, textvariable=material_var)
            # preselect current material name if set
            if getattr(maquina, 'material', None):
                material_var.set(maquina.material.nombre)
            def _validate_mat_edit(event=None):
                val = material_var.get().strip()
                if val and val not in mat_names:
                    material_var.set("")
            material_combo.bind("<FocusOut>", _validate_mat_edit)
            material_combo.pack(padx=10, pady=2)
    
            # código/serie after material
            tk.Label(scrollable_frame, text="Código/Serie N°", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            serie_entry = tk.Entry(scrollable_frame, width=50)
            serie_entry.insert(0, getattr(maquina, 'codigo_serie', '') or '')
            serie_entry.pack(padx=10, pady=2)
    
            tk.Label(scrollable_frame, text="Ubicación", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            ubicacion_entry = tk.Entry(scrollable_frame, width=50)
            ubicacion_entry.insert(0, getattr(maquina, 'ubicacion', '') or '')
            ubicacion_entry.pack(padx=10, pady=2)
        
            tk.Label(scrollable_frame, text="Motor - Marca", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            motor_marca_entry = tk.Entry(scrollable_frame, width=50)
            motor_marca_entry.insert(0, getattr(maquina, 'motor_marca', '') or '')
            motor_marca_entry.pack(padx=10, pady=2)
        
            tk.Label(scrollable_frame, text="Motor - Modelo", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            motor_modelo_entry = tk.Entry(scrollable_frame, width=50)
            motor_modelo_entry.insert(0, getattr(maquina, 'motor_modelo', '') or '')
            motor_modelo_entry.pack(padx=10, pady=2)
        
            tk.Label(scrollable_frame, text="Motor - Serie", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            motor_serie_entry = tk.Entry(scrollable_frame, width=50)
            motor_serie_entry.insert(0, getattr(maquina, 'motor_serie', '') or '')
            motor_serie_entry.pack(padx=10, pady=2)
        
            tk.Label(scrollable_frame, text="Tipo de Combustible", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            comb_entry = tk.Entry(scrollable_frame, width=50)
            comb_entry.insert(0, getattr(maquina, 'tipo_combustible', '') or '')
            comb_entry.pack(padx=10, pady=2)
        
            tk.Label(scrollable_frame, text="Altura de corte (mm)", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            altura_edit = tk.Entry(scrollable_frame, width=50)
            altura_edit.insert(0, getattr(maquina, 'altura_corte_mm', '') or '')
            altura_edit.pack(padx=10, pady=2)
        
            tk.Label(scrollable_frame, text="Observaciones", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            obs_entry = tk.Entry(scrollable_frame, width=50)
            obs_entry.insert(0, getattr(maquina, 'observaciones', '') or '')
            obs_entry.pack(padx=10, pady=2)
        
            tk.Label(scrollable_frame, text="Horómetro actual", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            hor_entry = tk.Entry(scrollable_frame, width=50)
            hor_entry.insert(0, getattr(maquina, 'horometro_actual', 0))
            hor_entry.pack(padx=10, pady=2)
        
            tk.Label(scrollable_frame, text="Horómetro inicial (fijo)", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8,0))
            ho = tk.Entry(scrollable_frame, width=50)
            ho.insert(0, maquina.horometro_inicial)
            ho.config(state="readonly")
            ho.pack(padx=10, pady=2)
    
            def guardar():
                try:
                    nombre = nombre_entry.get().strip()
                    if not nombre:
                        raise ValueError("El nombre es obligatorio")
                    tipo_unidad = tipo_entry.get().strip() or None
                    try:
                        año_val = int(ano_entry.get().strip()) if ano_entry.get().strip() else None
                    except ValueError:
                        raise ValueError("Año debe ser un número")
                    codigo_serie = serie_entry.get().strip() or None
                    motor_marca = motor_marca_entry.get().strip() or None
                    motor_modelo = motor_modelo_entry.get().strip() or None
                    motor_serie = motor_serie_entry.get().strip() or None
                    tipo_comb = comb_entry.get().strip() or None
                    # (Material asociado removido; usamos componentes para esa relación)
                    mat_id = None
                    categoria = categoria_entry.get().strip() or None
                    est = estado_entry.get().strip() or None
                    ubi = ubicacion_entry.get().strip() or None
                    try:
                        altura_val = int(altura_edit.get().strip() or 0)
                    except Exception:
                        altura_val = None
                    observ = obs_entry.get().strip() or None
                    hor_act = None
                    try:
                        hor_act = float(hor_entry.get() or 0)
                    except ValueError:
                        raise ValueError("Horómetro actual debe ser número")
                    nuevo_id = None
                    try:
                        tid = id_entry.get().strip()
                        if tid and int(tid) != maq.id:
                            nuevo_id = int(tid)
                    except ValueError:
                        raise ValueError("ID debe ser un número")
                    actualizar_maquina(
                        maq.id,
                        nombre=nombre,
                        tipo_unidad=tipo_unidad,
                        fabricante=fabricante,
                        modelo=modelo,
                        año=año_val,
                        codigo_serie=codigo_serie,
                        motor_marca=motor_marca,
                        motor_modelo=motor_modelo,
                        motor_serie=motor_serie,
                        tipo_combustible=tipo_comb,
                        material_id=mat_id,
                        categoria=categoria,
                        estado=est,
                        ubicacion=ubi,
                        altura_corte_mm=altura_val,
                        horometro_actual=hor_act,
                        observaciones=observ,
                        nuevo_id=nuevo_id,
                    )
                    messagebox.showinfo("Actualizado", "Máquina actualizada")
                    self.view_mgr.pop()
                    self.refrescar_lista()
                except Exception as e:
                    messagebox.showerror("Error", str(e))
    
            _icon_btn(scrollable_frame, "✅", "Guardar cambios", guardar, font=("Arial", 10)).pack(pady=20)
    
        self.view_mgr.push("Editar Máquina", _build, scrollable=True)
    
    def detallar_maquina(self):
        """Vista embebida con detalles de la máquina seleccionada.
    
        Permite editar metadatos, gestionar componentes y adjuntar imágenes.
        Se activa con doble clic en la lista principal.
        """
        sel = self.listbox.curselection()
        if not sel:
            messagebox.showwarning("Aviso", "Seleccione una máquina")
            return
        maq = self.maquinas[sel[0]]
        self._detalle_maquina_id = maq.id
        self._detalle_hor_entry = None
    
        def _build(content):
            # Cabecera moderna con título y acción rápida
            header_bg = theme.current.get("header_bg", theme.current.get("bg"))
            header_fg = theme.current.get("header_fg", theme.current.get("fg"))
            header = tk.Frame(content, bg=header_bg)
            header.pack(fill=tk.X, padx=10, pady=(10, 2))

            tk.Label(
                header,
                font=("Arial", 12, "bold"),
                bg=header_bg,
                fg=header_fg,
            ).pack(side=tk.LEFT, padx=(0, 6))

            def _add_manual_header():
                # Reutiliza el mismo flujo que el botón de la sección de manuales
                _add_manual()

            def _export_detalles_a4():
                # Exporta un PDF o TXT listo para imprimir (A4)
                try:
                    from reportlab.lib.pagesizes import A4
                    from reportlab.pdfgen import canvas
                except ImportError:
                    # Fallback: exportar como texto plano
                    path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Texto", "*.txt")])
                    if not path:
                        return
                    with open(path, "w", encoding="utf-8") as f:
                        f.write(f"DETALLES MÁQUINA - ID {maq.id}\n")
                        f.write("""============================\n\n""")
                        f.write(f"Nombre: {maq.nombre}\n")
                        f.write(f"Categoría: {maq.categoria or ''}\n")
                        f.write(f"Estado: {maq.estado or ''}\n")
                        f.write(f"Ubicación: {maq.ubicacion or ''}\n")
                        f.write(f"Fabricante: {maq.fabricante or ''}\n")
                        f.write(f"Modelo: {maq.modelo or ''}\n")
                        f.write(f"Horómetro actual: {getattr(maq, 'horometro_actual', '')}\n")
                        f.write(f"Observaciones: {getattr(maq, 'observaciones', '') or ''}\n\n")
                        f.write("COMPONENTES:\n")
                        comps = listar_componentes(maq.id) or []
                        for c in comps:
                            mat = getattr(c, 'material', None)
                            f.write(f" - {c.nombre} ({getattr(mat,'nombre','')}) [{getattr(mat,'tipo','')}]\n")
                    messagebox.showinfo("Exportado", f"Detalles guardados en:\n{path}")
                    return

                path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF", "*.pdf")])
                if not path:
                    return

                c = canvas.Canvas(path, pagesize=A4)
                width, height = A4
                margin = 40
                y = height - margin
                c.setFont("Helvetica-Bold", 14)
                c.drawString(margin, y, f"Detalles Máquina - ID {maq.id}")
                y -= 24
                c.setFont("Helvetica", 10)
                def line(text, indent=0):
                    nonlocal y
                    if y < margin + 40:
                        c.showPage()
                        y = height - margin
                        c.setFont("Helvetica", 10)
                    c.drawString(margin + indent, y, text)
                    y -= 14

                line(f"Nombre: {maq.nombre}")
                line(f"Categoría: {maq.categoria or ''}")
                line(f"Estado: {maq.estado or ''}")
                line(f"Ubicación: {maq.ubicacion or ''}")
                line(f"Fabricante: {maq.fabricante or ''}")
                line(f"Modelo: {maq.modelo or ''}")
                line(f"Horómetro actual: {getattr(maq, 'horometro_actual', '')}")
                line(f"Observaciones: {getattr(maq, 'observaciones', '') or ''}")
                line("")
                line("Componentes:")
                comps = listar_componentes(maq.id) or []
                for cpt in comps:
                    mat = getattr(cpt, 'material', None)
                    line(f" - {cpt.nombre} | {getattr(mat,'nombre','')} | {getattr(mat,'tipo','')}", indent=14)

                c.showPage()
                c.save()
                messagebox.showinfo("Exportado", f"PDF guardado en:\n{path}")

            _icon_btn(
                header,
                "📄",
                "Adjuntar manual (PDF)",
                _add_manual_header,
                font=("Arial", 12),
                relief="flat",
                bg=header_bg,
                fg=header_fg,
            ).pack(side=tk.RIGHT)

            _icon_btn(
                header,
                "🖨️",
                "Exportar detalles (A4)",
                _export_detalles_a4,
                font=("Arial", 12),
                relief="flat",
                bg=header_bg,
                fg=header_fg,
            ).pack(side=tk.RIGHT, padx=(0, 6))

            # --- Layout: Izquierda=metadatos | Derecha=componentes ---
            main_frame = tk.Frame(content)
            main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

            # Izquierda: datos de máquina
            left_frame = tk.Frame(main_frame)
            left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))

            entries = {}
            tk.Label(left_frame, text="ID").pack(anchor='w')
            id_entry = tk.Entry(left_frame)
            id_entry.insert(0, str(maq.id))
            id_entry.pack(fill=tk.X)

            for campo in ["nombre", "categoria", "estado", "ubicacion", "fabricante", "modelo"]:
                tk.Label(left_frame, text=campo.capitalize()).pack(anchor='w', pady=(6, 0))
                entries[campo] = tk.Entry(left_frame)
                entries[campo].insert(0, getattr(maq, campo) or "")
                entries[campo].pack(fill=tk.X)

            tk.Label(left_frame, text="Observaciones").pack(anchor='w', pady=(6, 0))
            obs_entry = tk.Entry(left_frame)
            obs_entry.insert(0, getattr(maq, 'observaciones', '') or '')
            obs_entry.pack(fill=tk.X)

            tk.Label(left_frame, text="Horómetro actual").pack(anchor='w', pady=(6, 0))
            hor_entry = tk.Entry(left_frame)
            hor_entry.insert(0, getattr(maq, 'horometro_actual', 0))
            hor_entry.pack(fill=tk.X)
            self._detalle_hor_entry = hor_entry

            def guardar_meta():
                try:
                    nombre = entries["nombre"].get().strip()
                    if not nombre:
                        raise ValueError("El nombre es obligatorio")
                    cat = entries["categoria"].get().strip() or None
                    est = entries["estado"].get().strip() or None
                    ubi = entries["ubicacion"].get().strip() or None
                    fab = entries["fabricante"].get().strip() or None
                    mod = entries["modelo"].get().strip() or None
                    observ = obs_entry.get().strip() or None
                    hor_act = None
                    try:
                        hor_act = float(hor_entry.get() or 0)
                    except ValueError:
                        raise ValueError("Horómetro actual debe ser número")
                    nuevo_id = None
                    try:
                        tid = id_entry.get().strip()
                        if tid and int(tid) != maq.id:
                            nuevo_id = int(tid)
                    except ValueError:
                        raise ValueError("ID debe ser un número")
                    actualizar_maquina(
                        maq.id,
                        nombre=nombre,
                        categoria=cat,
                        estado=est,
                        ubicacion=ubi,
                        fabricante=fab,
                        modelo=mod,
                        material_id=None,
                        horometro_actual=hor_act,
                        observaciones=observ,
                        nuevo_id=nuevo_id,
                    )
                    messagebox.showinfo("Actualizado", "Máquina actualizada")
                    self.view_mgr.pop()
                    self.refrescar_lista()
                except Exception as e:
                    messagebox.showerror("Error", str(e))

            _icon_btn(left_frame, "💾", "Guardar metadatos", guardar_meta).pack(fill=tk.X, pady=6)
            _icon_btn(left_frame, "⚙️", "Datos Avanzados", lambda: self._dialogo_datos_avanzados(maq)).pack(fill=tk.X)

            # Derecha: Componentes + anexos
            right_frame = tk.Frame(main_frame)
            right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0))

            # componentes section (ficha técnica rápida)
            comp_frame = tk.LabelFrame(right_frame, text="Componentes")
            comp_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 5))

            cols = ("id", "nombre", "material", "tipo", "sistema", "descripcion")
            comp_tree = ttk.Treeview(comp_frame, columns=cols, show="headings", selectmode="browse")
            comp_tree.column("id", width=40, anchor="center", stretch=False)
            comp_tree.column("nombre", width=140, stretch=False)
            comp_tree.column("material", width=130, stretch=False)
            comp_tree.column("tipo", width=100, stretch=False)
            comp_tree.column("sistema", width=110, stretch=False)
            comp_tree.column("descripcion", width=240, stretch=False)
            for c in cols:
                comp_tree.heading(c, text=c.capitalize())

            vsb = ttk.Scrollbar(comp_frame, orient=tk.VERTICAL, command=comp_tree.yview)
            hsb = ttk.Scrollbar(comp_frame, orient=tk.HORIZONTAL, command=comp_tree.xview)
            comp_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            vsb.pack(side=tk.RIGHT, fill=tk.Y)
            hsb.pack(side=tk.BOTTOM, fill=tk.X)
            comp_tree.pack(fill=tk.BOTH, expand=True)

            def refresh_comps():
                comp_tree.delete(*comp_tree.get_children())
                comps = listar_componentes(maq.id) or []
                for c in comps:
                    mat = getattr(c, 'material', None)
                    comp_tree.insert(
                        "",
                        tk.END,
                        iid=str(getattr(c, 'id', '')),
                        values=(
                            getattr(c, 'id', ''),
                            getattr(c, 'nombre', ''),
                            getattr(mat, 'nombre', '') if mat else "",
                            getattr(mat, 'tipo', '') if mat else "",
                            getattr(c, 'sistema', '') or "",
                            getattr(c, 'descripcion', '') or "",
                        ),
                    )

            def _on_comp_double(event=None):
                sel = comp_tree.selection()
                if not sel:
                    return
                try:
                    comp_id = int(sel[0])
                except Exception:
                    return
                self._dialogo_editar_componente(comp_id)

            comp_tree.bind("<Double-1>", _on_comp_double)

            btns = tk.Frame(comp_frame)
            btns.pack(fill=tk.X, pady=2)
            _icon_btn(btns, "🔄", "Refrescar componentes", refresh_comps).pack(side=tk.LEFT, padx=2)
            _icon_btn(btns, "➕", "Agregar componente", lambda: self._dialogo_agregar_componente(maq, callback=refresh_comps)).pack(side=tk.LEFT, padx=2)

            refresh_comps()

            # planos / imágenes section
            plano_frame = tk.LabelFrame(right_frame, text="Imágenes / Planos")
            plano_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 5))
            plano_list = tk.Listbox(plano_frame)
            plano_list.pack(fill=tk.BOTH, expand=True)

            def _add_plano():
                file = filedialog.askopenfilename(title="Seleccionar imagen/plano", parent=self.panel)
                if not file:
                    return
                nombre = simpledialog.askstring("Nombre", "Ingrese nombre:", parent=self.panel) or ""
                from .services.ingenieria_extras import agregar_plano
                agregar_plano(maq.id, nombre, file)
                refresh_planos()

            _icon_btn(plano_frame, "📁", "Adjuntar imagen/plano", _add_plano, state=tk.NORMAL).pack(pady=2)

            def refresh_planos():
                plano_list.delete(0, tk.END)
                from .services.ingenieria_extras import listar_planos
                for p in listar_planos(maq.id):
                    plano_list.insert(tk.END, p.nombre)

            refresh_planos()

            # manuales (PDF/documentos) section
            manual_frame = tk.LabelFrame(right_frame, text="Manuales / Documentos")
            manual_frame.pack(fill=tk.BOTH, expand=True)
            manual_list = tk.Listbox(manual_frame)
            manual_list.pack(fill=tk.BOTH, expand=True)

            def _add_manual():
                file = filedialog.askopenfilename(
                    title="Seleccionar manual (PDF)",
                    filetypes=[("PDF", "*.pdf"), ("Todos", "*")],
                    parent=self.panel,
                )
                if not file:
                    return
                nombre = simpledialog.askstring("Nombre", "Ingrese nombre del manual:", parent=self.panel) or ""
                from .services.ingenieria_service import agregar_manual_maquina
                agregar_manual_maquina(maq.id, file, nombre=nombre)
                refresh_manuales()

            def _open_manual():
                sel = manual_list.curselection()
                if not sel:
                    return
                idx = sel[0]
                manual = actuales_manuales[idx]
                try:
                    if os.name == 'nt':
                        os.startfile(manual.url)
                    else:
                        import webbrowser
                        webbrowser.open(manual.url)
                except Exception:
                    messagebox.showerror("Error", "No se pudo abrir el archivo.")

            def _delete_manual():
                sel = manual_list.curselection()
                if not sel:
                    return
                idx = sel[0]
                manual = actuales_manuales[idx]
                if messagebox.askyesno("Confirmar", f"Eliminar manual '{manual.nombre}'?"):
                    from .services.ingenieria_service import eliminar_manual
                    eliminar_manual(manual.id)
                    refresh_manuales()

            btns_man = tk.Frame(manual_frame)
            btns_man.pack(fill=tk.X, pady=4)
            _icon_btn(btns_man, "📄", "Adjuntar manual", _add_manual).pack(side=tk.LEFT, padx=2)
            _icon_btn(btns_man, "📂", "Abrir manual", _open_manual).pack(side=tk.LEFT, padx=2)
            _icon_btn(btns_man, "🗑", "Eliminar manual", _delete_manual).pack(side=tk.LEFT, padx=2)

            actuales_manuales = []
            def refresh_manuales():
                nonlocal actuales_manuales
                manual_list.delete(0, tk.END)
                from .services.ingenieria_service import listar_manuales_maquina
                try:
                    actuales_manuales = listar_manuales_maquina(maq.id)
                except Exception as e:
                    actuales_manuales = []
                    messagebox.showwarning(
                        "Error al cargar manuales",
                        f"No se pudieron listar los manuales de la máquina:\n{e}",
                        parent=self.panel,
                    )
                for m in actuales_manuales:
                    manual_list.insert(tk.END, m.nombre or os.path.basename(m.url))

            refresh_manuales()

        self.view_mgr.push(f"Detalles Máquina {maq.id}", _build)
        # Forzar el tema actual sobre el nuevo panel (modo oscuro/claro)
        try:
            current_view = self.view_mgr.current_view()
            if current_view is not None:
                theme.apply_theme(current_view)
        except Exception:
            pass
    
    def _dialogo_datos_avanzados(self, maq):
        """Vista para editar los datos avanzados de una máquina (campos extra de 'Agregar Máquina')."""
        def _build(scrollable_frame):
            tk.Label(scrollable_frame, text=f"Datos Avanzados — {maq.nombre}", font=("Arial", 11, "bold")).pack(pady=(10, 5))
    
            tk.Label(scrollable_frame, text="Tipo de unidad", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8, 0))
            tipo_entry = tk.Entry(scrollable_frame, width=50)
            tipo_entry.pack(fill=tk.X, padx=10, pady=2)
            tipo_entry.insert(0, getattr(maq, 'tipo_unidad', '') or '')
    
            tk.Label(scrollable_frame, text="Año", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8, 0))
            ano_entry = tk.Entry(scrollable_frame, width=15)
            ano_entry.pack(anchor="w", padx=10, pady=2)
            ano_entry.insert(0, getattr(maq, 'año', '') or '')
    
            tk.Label(scrollable_frame, text="Código/Serie N°", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8, 0))
            serie_entry = tk.Entry(scrollable_frame, width=50)
            serie_entry.pack(fill=tk.X, padx=10, pady=2)
            serie_entry.insert(0, getattr(maq, 'codigo_serie', '') or '')
    
            # Motor: Marca, Modelo, Serie
            frm_motor = tk.LabelFrame(scrollable_frame, text="Motor", padx=6, pady=4)
            frm_motor.pack(fill=tk.X, padx=10, pady=(8, 2))
            for col in range(6):
                frm_motor.columnconfigure(col, weight=1 if col % 2 == 1 else 0)
            tk.Label(frm_motor, text="Marca:", font=("Arial", 9)).grid(row=0, column=0, sticky="w", padx=2)
            motor_marca_entry = tk.Entry(frm_motor, width=14)
            motor_marca_entry.grid(row=0, column=1, sticky="ew", padx=2)
            tk.Label(frm_motor, text="Modelo:", font=("Arial", 9)).grid(row=0, column=2, sticky="w", padx=2)
            motor_modelo_entry = tk.Entry(frm_motor, width=14)
            motor_modelo_entry.grid(row=0, column=3, sticky="ew", padx=2)
            tk.Label(frm_motor, text="Serie:", font=("Arial", 9)).grid(row=0, column=4, sticky="w", padx=2)
            motor_serie_entry = tk.Entry(frm_motor, width=14)
            motor_serie_entry.grid(row=0, column=5, sticky="ew", padx=2)
    
            tk.Label(scrollable_frame, text="Tipo de Combustible", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8, 0))
            comb_entry = tk.Entry(scrollable_frame, width=50)
            comb_entry.pack(fill=tk.X, padx=10, pady=2)
            comb_entry.insert(0, getattr(maq, 'tipo_combustible', '') or '')
    
            tk.Label(scrollable_frame, text="Altura de corte (mm)", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8, 0))
            altura_entry = tk.Entry(scrollable_frame, width=15)
            altura_entry.pack(anchor="w", padx=10, pady=2)
            altura_entry.insert(0, getattr(maq, 'altura_corte_mm', '') or '')
    
            # Fecha de entrada en operación
            frm_fecha_op = tk.LabelFrame(scrollable_frame, text="📅 Fecha de entrada en operación", padx=8, pady=4)
            frm_fecha_op.pack(fill=tk.X, padx=10, pady=(8, 2))
            fecha_op_var = tk.StringVar()
            if getattr(maq, 'fecha_operacion', None):
                fecha_op_var.set(maq.fecha_operacion.strftime("%Y-%m-%d"))
            tk.Label(
                frm_fecha_op,
                textvariable=fecha_op_var,
                width=14,
                relief="sunken",
                bg=theme.current.get("entry_bg", "white"),
                fg=theme.current.get("entry_fg", "black"),
            ).pack(side=tk.LEFT, padx=4)
            def elegir_fecha_op():
                cal_win = tk.Toplevel(self.panel)
                cal_win.title("Fecha de entrada en operación")
                from tkcalendar import Calendar
                cal = Calendar(cal_win, selectmode="day")
                cal.pack(padx=10, pady=10)
                def _ok():
                    fecha_op_var.set(cal.selection_get().isoformat())
                    cal_win.destroy()
                btns = tk.Frame(cal_win)
                btns.pack(pady=5)
                _icon_btn(btns, "✔", "Aceptar", _ok).pack(side=tk.LEFT, padx=5)
                _icon_btn(btns, "❌", "Cancelar", cal_win.destroy).pack(side=tk.LEFT, padx=5)
            _icon_btn(frm_fecha_op, "📅", "Seleccionar fecha", elegir_fecha_op).pack(side=tk.LEFT, padx=4)
    
            tk.Label(scrollable_frame, text="Horómetro inicial", font=("Arial", 9)).pack(anchor="w", padx=10, pady=(8, 0))
            horometro_ini_entry = tk.Entry(scrollable_frame, width=15)
            horometro_ini_entry.pack(anchor="w", padx=10, pady=2)
            horometro_ini_entry.insert(0, getattr(maq, 'horometro_inicial', 0) or 0)
    
            def guardar_avanzados():
                try:
                    tipo_unidad = tipo_entry.get().strip() or None
                    try:
                        año_val = int(ano_entry.get().strip()) if ano_entry.get().strip() else None
                    except ValueError:
                        raise ValueError("Año debe ser un número")
                    codigo_serie = serie_entry.get().strip() or None
                    motor_marca = motor_marca_entry.get().strip() or None
                    motor_modelo = motor_modelo_entry.get().strip() or None
                    motor_serie = motor_serie_entry.get().strip() or None
                    tipo_comb = comb_entry.get().strip() or None
                    try:
                        altura_val = int(altura_entry.get().strip()) if altura_entry.get().strip() else None
                    except ValueError:
                        raise ValueError("Altura de corte debe ser un número")
                    fecha_op = None
                    if fecha_op_var.get():
                        from datetime import datetime as _dt
                        fecha_op = _dt.fromisoformat(fecha_op_var.get())
                    try:
                        horo_ini = float(horometro_ini_entry.get().strip()) if horometro_ini_entry.get().strip() else None
                    except ValueError:
                        raise ValueError("Horómetro inicial debe ser un número")
    
                    actualizar_maquina(
                        maq.id,
                        tipo_unidad=tipo_unidad,
                        año=año_val,
                        codigo_serie=codigo_serie,
                        motor_marca=motor_marca,
                        motor_modelo=motor_modelo,
                        motor_serie=motor_serie,
                        tipo_combustible=tipo_comb,
                        altura_corte_mm=altura_val,
                        fecha_operacion=fecha_op,
                    )
                    # Horómetro inicial se actualiza directamente
                    if horo_ini is not None:
                        from .database import SessionLocal
                        from .models import Maquina
                        with SessionLocal() as db:
                            m = db.query(Maquina).filter(Maquina.id == maq.id).first()
                            if m:
                                m.horometro_inicial = horo_ini
                                db.commit()
    
                    messagebox.showinfo("Actualizado", "Datos avanzados guardados correctamente")
                    self.view_mgr.pop()
                    self.refrescar_lista()
                except Exception as e:
                    messagebox.showerror("Error", str(e))
    
            _icon_btn(scrollable_frame, "💾", "Guardar datos avanzados", guardar_avanzados, font=("Arial", 10, "bold")).pack(fill=tk.X, padx=10, pady=10)
    
        self.view_mgr.push("Datos Avanzados", _build)
    
    def eliminar(self):
        sel = self.listbox.curselection()
        if not sel:
            messagebox.showwarning("Aviso", "Seleccione una máquina", parent=self.panel)
            return
        maquina = self.maquinas[sel[0]]
        confirmar = messagebox.askyesno("Confirmar", f"¿Eliminar máquina '{maquina.nombre}'?", parent=self.panel)
        if confirmar:
            snap = self._snapshot_maquina(maquina)
            if len(self.maquinas) <= 1:
                # Permitir iniciar de cero: no crear demo automáticamente
                self._disable_demo_on_empty = True
            eliminar_maquina(maquina.id)
            self.undo.push(f"Eliminar máquina '{maquina.nombre}'", lambda s=snap: self._restore_maquina(s))
            self.refrescar_lista()
    def _ask_option(self, title, prompt, options):
        dlg = tk.Toplevel(self.panel)
        dlg.title(title)
        var = tk.StringVar(value=options[0])
        tk.Label(dlg, text=prompt).pack(padx=10, pady=5)
        for opt in options:
            tk.Radiobutton(dlg, text=opt.capitalize(), variable=var, value=opt).pack(anchor="w", padx=20)
        _icon_btn(dlg, "✔", "Aceptar", dlg.destroy).pack(pady=10)
        dlg.transient(self.panel)
        dlg.grab_set()
        self.panel.wait_window(dlg)
        return var.get()
    
    def _dialogo_exportar(self):
        # centralizar exportaciones: máquinas o historial de la(s) máquina(s) seleccionada(s)
        choice = self._ask_option("Exportar", "Seleccione qué desea exportar:",
                                  ["maquinas", "historial", "ficha"])
        if not choice:
            return
        if choice == "historial":
            sel = self.listbox.curselection()
            if not sel:
                messagebox.showwarning("Aviso", "Seleccione al menos una máquina primero")
                return
            ids = [self.maquinas[i].id for i in sel]
            from .services.historial_service import exportar_historial, exportar_historial_excel
            start = None
            end = None
            from datetime import datetime
            if self.hist_start_var.get():
                try:
                    start = datetime.fromisoformat(self.hist_start_var.get())
                except ValueError:
                    pass
            if self.hist_end_var.get():
                try:
                    end = datetime.fromisoformat(self.hist_end_var.get())
                except ValueError:
                    pass
            # if no interval provided the export will include todo el historial
            if start is None and end is None:
                messagebox.showinfo("Exportar", "No se han seleccionado fechas; se exportará todo el historial.")
            if len(ids) == 1:
                maq = next((m for m in self.maquinas if m.id == ids[0]), None)
                if maq and start and maq.fecha_actualizacion and start < maq.fecha_actualizacion:
                    messagebox.showwarning(
                        "Fechas",
                        "El inicio no puede ser anterior a la creación de la máquina. Se ajustará."
                    )
                    start = maq.fecha_actualizacion
            fmt = self._ask_option("Formato", "Seleccione formato:", ["csv", "json", "xlsx"])
            if fmt != "xlsx" and len(ids) > 1:
                messagebox.showwarning("Formato", "Para varias máquinas elija formato Excel (xlsx).")
                return
            if fmt == "xlsx":
                exportar_historial_excel(maquina_ids=ids, rango_inicio=start, rango_fin=end)
                return
            # otherwise csv or json, combine into single text with separators
            combined = []
            for mid in ids:
                part = exportar_historial(maquina_id=mid, rango_inicio=start, rango_fin=end, formato=fmt)
                if part:
                    header = f"\n# Máquina {mid}\n" if len(ids) > 1 else ""
                    combined.append(header + part)
            if not combined:
                messagebox.showinfo("Exportar", "No hay registros en el intervalo seleccionado.")
                return
            data = "\n".join(combined)
            filepath = filedialog.asksaveasfilename(defaultextension="." + fmt,
                                                    filetypes=[("CSV", "*.csv"), ("JSON", "*.json")])
            if not filepath:
                return
            with open(filepath, "w", encoding="utf-8") as f:
                f.write(data)
            messagebox.showinfo("Exportar", f"Historial exportado a {filepath}")
            return
    
        # si se pidió ficha técnica
        if choice == "ficha":
            self.generar_fichas_directo()
            return
    
        # exportar máquinas (posible selección múltiple) sin pedir rango:
        # ofrecemos formato por radio opciones y siempre tomamos todos los datos.
        sel = self.listbox.curselection()
        ids = [self.maquinas[i].id for i in sel] if sel else None
    
        fmt = self._ask_option("Formato", "Seleccione formato:", ["csv", "json", "xlsx"])
        # si eligió csv/json, utilice el servicio estándar; para xlsx podemos
        # reutilizar el mismo método (exportar_maquinas devuelve texto csv, pero
        # la UI nunca pedirá xlsx para máquinas, por ahora sólo sirve para
        # consistencia con historial).
        data = exportar_maquinas(formato=fmt, maquina_ids=ids)
        if not data:
            messagebox.showinfo("Exportar", "No hay datos para exportar.")
            return
        filepath = filedialog.asksaveasfilename(
            defaultextension="." + fmt,
            filetypes=[("CSV", "*.csv"), ("JSON", "*.json")],
        )
        if not filepath:
            return
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(data)
        messagebox.showinfo("Exportar", f"Datos exportados a {filepath}")
    
    def generar_fichas_directo(self):
        # helper that mirrors logic used elsewhere for ficha técnica export
        sel = self.listbox.curselection()
        if sel:
            ids = [self.maquinas[i].id for i in sel]
        else:
            resp = messagebox.askyesno(
                "Fichas técnicas",
                "No hay ninguna máquina seleccionada. ¿Generar para todas?",
                parent=self.panel,
            )
            if resp:
                ids = None
            else:
                messagebox.showwarning("Aviso", "Seleccione al menos una máquina")
                return
        fmt = simpledialog.askstring("Formato", "Ingrese formato (excel/pdf):", parent=self.panel) or "excel"
        from .services.export_service import export_ficha_tecnica
        export_ficha_tecnica(ids, formato=fmt.lower())
    
    
    def _buscar_toro(self):
        termino = self.search_entry.get().strip()
        if not termino:
            return
        # abrir navegador hacia la búsqueda en Toro para que el usuario vea
        # manuales y productos disponibles.
        import webbrowser
        url = f"https://www.toro.com/search?q={termino}"
        try:
            webbrowser.open(url)
        except Exception:
            pass
        # obtener información usando el servicio stub (puede mejorarse con
        # scraping real en toro_utils).
        info = buscar_info_toro(termino)
        # intentar extraer algunos códigos/manuales de la página si bs4 está
        # disponible
        try:
            import requests
            from bs4 import BeautifulSoup
            r = requests.get(url, timeout=5)
            soup = BeautifulSoup(r.text, "html.parser")
            codes = []
            for a in soup.find_all("a"):
                href = a.get("href", "")
                text = a.get_text(strip=True)
                if "/product/" in href or "manual" in text.lower():
                    codes.append(text or href)
            if codes:
                info["candidatos"] = codes
        except Exception:
            # si no hay red o bs4 no está instalado, ignoramos
            pass
        self.result_text.delete("1.0", tk.END)
        for k, v in info.items():
            self.result_text.insert(tk.END, f"{k}: {v}\n")
    
    def _dialogo_agregar_componente(self, maquina=None, callback=None):
        """Diálogo para agregar un componente, ahora con selección múltiple de máquinas.

        `callback` (si se proporciona) se ejecuta después de guardar el componente.
        """

        from .services.ingenieria_extras import agregar_componente, listar_materiales

        all_maquinas = self.maquinas  # Lista de objetos máquina
        mats = listar_materiales() or []        
        mat_names = [m.nombre for m in mats]

        def _build(content):
            tk.Label(content, text="Material asociado (opcional)", font=("Arial", 10, "bold")).pack(anchor="w", padx=10, pady=(10,0))
            comp_mat_var = tk.StringVar()
            comp_mat_combo = ttk.Combobox(content, values=mat_names, textvariable=comp_mat_var, width=38)
            comp_mat_combo.pack(fill=tk.X, padx=10, pady=2)

            tk.Label(content, text="Sistema (opcional)", font=("Arial", 10, "bold")).pack(anchor="w", padx=10, pady=(10,0))
            comp_sist_var = tk.StringVar()
            sist_vals = getattr(self, "sistemas_lista", []) or []
            comp_sist_combo = ttk.Combobox(content, values=sist_vals, textvariable=comp_sist_var, width=38)
            comp_sist_combo.pack(fill=tk.X, padx=10, pady=2)

            def _sugerir_sistema():
                try:
                    from .services.prediction_service import clasificar_componente
                    current_name = comp_mat_var.get().strip() or ""
                    current_desc = ""
                    sugerido = clasificar_componente(current_name, current_desc, sistemas=sist_vals)
                    if sugerido:
                        comp_sist_var.set(sugerido)
                except Exception:
                    pass

            _icon_btn(content, "🤖", "Sugerir sistema", _sugerir_sistema).pack(fill=tk.X, padx=10, pady=(0, 5))

            # -----------------------------------------
            # SELECCIÓN MÚLTIPLE DE MÁQUINAS
            tk.Label(content, text="Máquinas (seleccione una o más con Ctrl/Shift)", font=("Arial", 10, "bold")).pack(anchor="w", padx=10, pady=(8,0))
            lb_maquinas = tk.Listbox(content, selectmode=tk.MULTIPLE, height=6)
            for idx, m in enumerate(all_maquinas):
                lb_maquinas.insert(tk.END, f"{m.nombre} (ID {m.id})")
                if maquina and maquina.id == m.id:
                    lb_maquinas.selection_set(idx)  # Selecciona la máquina por defecto
            lb_maquinas.pack(fill=tk.BOTH, padx=10, pady=2)

            # -----------------------------------------
            def guardar_componente():
                mat_name = comp_mat_var.get().strip()
                mat_id = None
                mat_obj = None
                if mat_name:
                    mat_obj = next((x for x in mats if x.nombre == mat_name), None)
                    if mat_obj:
                        mat_id = mat_obj.id

                nombre = mat_obj.nombre if mat_obj else "Componente"
                descripcion = mat_obj.descripcion if mat_obj and getattr(mat_obj, 'descripcion', None) else ""
                sistema = comp_sist_var.get().strip() or None

                selected_indices = lb_maquinas.curselection()
                if not selected_indices:
                    messagebox.showwarning("Aviso", "Seleccione al menos una máquina")
                    return

                try:
                    for idx in selected_indices:
                        m = all_maquinas[idx]
                        agregar_componente(
                            maquina_id=m.id,
                            nombre=nombre,
                            descripcion=descripcion,
                            material_id=mat_id,
                            sistema=sistema,
                        )
                    messagebox.showinfo("Éxito", f"Componente '{nombre}' agregado a {len(selected_indices)} máquina(s)")
                    self._refresh_componentes()
                    if callback:
                        try:
                            callback()
                        except Exception:
                            pass
                    self.view_mgr.pop()
                except Exception as e:
                    messagebox.showerror("Error", str(e))

            _icon_btn(content, "✅", "Guardar componente", guardar_componente, font=("Arial", 10, "bold")).pack(fill=tk.X, padx=10, pady=10)

        self.view_mgr.push("Agregar Componente", _build, scrollable=True)
    
    
    def _elegir_hist_fecha(self, var):
        # reuse a single calendar window for both 'desde' and 'hasta'.  If it
        # already exists just switch the target variable and bring it to front.
        if hasattr(self, '_hist_cal_win') and getattr(self, '_hist_cal_win', None) and self._hist_cal_win.winfo_exists():
            # change target and show
            self._hist_target = var
            try:
                sel = var.get()
                if sel:
                    self._hist_cal.selection_set(sel)
            except Exception:
                pass
            self._hist_cal_win.lift()
            return
    
        self._hist_target = var
        cal_win = tk.Toplevel(self.panel)
        cal_win.title("Elegir fecha")
        from tkcalendar import Calendar
        cal = Calendar(cal_win, selectmode="day")
        cal.pack(padx=10, pady=10)
        self._hist_cal_win = cal_win
        self._hist_cal = cal
    
        def _ok():
            fecha = cal.selection_get().isoformat()
            self._hist_target.set(fecha)
            # if we just set 'desde' and 'hasta' still empty, prompt to pick end
            if self._hist_target is self.hist_start_var and not self.hist_end_var.get():
                resp = messagebox.askyesno("Rango","¿Elegir fecha de HASTA ahora?", parent=cal_win)
                if resp:
                    self._hist_target = self.hist_end_var
                    return
            # no more selection required, close
            cal_win.destroy()
            self.refrescar_historial()
    
        def _cancel():
            cal_win.destroy()
    
        btns = tk.Frame(cal_win)
        btns.pack(pady=5)
        _icon_btn(btns, "✔", "Aceptar", _ok).pack(side=tk.LEFT, padx=5)
        _icon_btn(btns, "❌", "Cancelar", _cancel).pack(side=tk.LEFT, padx=5)
    
    
    
    
    
    # función auxiliar exportada para el resto de la aplicación
    
def abrir_panel_ingenieria(root):
    """Open or raise the Ingeniería panel singleton.

    If the panel already exists, focus it instead of creating another
    window.  This avoids redundant windows when invoked from Técnico or
    other modules.
    """
    global _current_ingenieria_panel
    if _current_ingenieria_panel and hasattr(_current_ingenieria_panel, "panel"):
        try:
            if _current_ingenieria_panel.panel.winfo_exists():
                _current_ingenieria_panel.panel.deiconify()
                _current_ingenieria_panel.panel.lift()
                _current_ingenieria_panel.panel.focus_force()
                return _current_ingenieria_panel
        except Exception:
            pass
    _current_ingenieria_panel = IngenieriaPanel(root)
    return _current_ingenieria_panel